import logging
import json
import unicodedata
import tempfile
from datetime import datetime
import subprocess
import base64
import requests
import re
from urllib3 import disable_warnings
from urllib3.exceptions import InsecureRequestWarning
from flask import current_app

disable_warnings(InsecureRequestWarning)


# 初始化FFmpeg环境（启动时执行）
def init_ffmpeg_env():
    """初始化FFmpeg环境变量"""
    if not current_app:
        with app.app_context():
            return _init_ffmpeg_env_core()
    else:
        return _init_ffmpeg_env_core()

# 核心逻辑抽离
def _init_ffmpeg_env_core():
    """FFmpeg初始化核心逻辑（需在应用上下文内执行）"""
    # 从系统配置获取FFmpeg路径（优先），无则使用默认路径
    ffmpeg_bin = SystemConfig.get_config('ffmpeg_bin', "D:\\ffmpeg\\bin")
    ffmpeg_exe = SystemConfig.get_config('ffmpeg_exe', "D:\\ffmpeg\\bin\\ffmpeg.exe")
    ffprobe_exe = SystemConfig.get_config('ffprobe_exe', "D:\\ffmpeg\\bin\\ffprobe.exe")

    # 配置环境变量
    os.environ["PATH"] += os.pathsep + ffmpeg_bin
    os.environ["FFMPEG_BINARY"] = ffmpeg_exe
    os.environ["FFPROBE_BINARY"] = ffprobe_exe

    # 验证FFmpeg
    if not os.path.exists(ffmpeg_exe):
        logger.warning(f"FFmpeg文件不存在：{ffmpeg_exe}，语音功能将不可用")
        return False

    try:
        result = subprocess.run(
            [ffmpeg_exe, "-version"],
            capture_output=True,
            text=True,
            timeout=5
        )
        logger.info("✅ FFmpeg 初始化成功")
        return True
    except Exception as e:
        logger.error(f"FFmpeg初始化失败：{str(e)}")
        return False


# 成果类型匹配规则（OCR识别用）
achievement_rules = {
    '期刊论文': {'pattern': r'[学报 | 期刊 | 杂志].*卷.*期 | ISSN:\d+',
                 'keywords': ['期刊', '学报', '论文', '发表', '卷', '期'], 'priority': 1},
    '发明专利': {'pattern': r'ZL\d{4}\d{8}(\.\d+)?|发明专利申请号 | 公开号',
                 'keywords': ['发明', '专利', 'ZL', '申请号', '公开号'], 'priority': 1},
    '实用新型专利': {'pattern': r'实用新型专利|ZL\d{4}2\d{7}', 'keywords': ['实用新型', '专利', 'ZL'], 'priority': 1},
    '会议论文': {'pattern': r'会议论文 | 会议集 |Proceedings', 'keywords': ['会议', '研讨会', '论坛'],
                 'priority': 2},
    '教材': {'pattern': r'教材|ISBN[:：]?\s*\d+|主编 [:：]?|副主编 [:：]?|出版社 [:：]?',
             'keywords': ['教材', '主编', '副主编', 'ISBN', '出版社', '规划教材', '行业教材'],
             'priority': 2},
    '专著': {'pattern': r'专著|ISBN:\d+|独著 | 合著', 'keywords': ['专著', '独著', '合著', 'ISBN'], 'priority': 2},
    '软著': {'pattern': r'计算机软件著作权 | 软著登字第\d+ 号', 'keywords': ['软著', '软件著作权', '著作权'],
             'priority': 3},
    '教学成果获奖': {'pattern': r'成果奖 | 科技奖 | 一等奖 | 二等奖', 'keywords': ['成果', '获奖', '科技奖', '一等奖', '二等奖'],
                 'priority': 3},
    '教学竞赛获奖': {'pattern': r'教学竞赛 | 教学奖 | 课堂教学', 'keywords': ['教学竞赛', '教学奖', '课堂教学'],
                     'priority': 3},
    '指导学生获奖': {'pattern': r'指导老师 | 学生竞赛 | 大学生.*竞赛 | 参赛同学', 'keywords': ['指导老师', '学生', '竞赛', '获奖'],
                     'priority': 2},
    '教研教改和课程建设项目': {'pattern': r'教学改革研究 | 教改 | 课程建设 | 一流本科课程 | 课程思政 | 高等学校教育教学改革',
                                'keywords': ['教学改革', '教改', '课程建设', '一流课程', '课程思政', '立项', '湖南省普通高等学校'],
                                'priority': 1}
}

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ocr_voice.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

from pdf2image import convert_from_path
import tempfile
import os
import json
import csv
from datetime import datetime, date
from io import BytesIO
from flask import Flask, request, redirect, url_for, flash, session, Response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
import uuid
import pandas as pd
from sqlalchemy import or_, func
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import re
import time
import random
import requests
import json
from datetime import datetime


# 新增：定义允许上传的文件扩展名
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'docx', 'doc', 'xlsx', 'xls'}

# ---------------------- 1. 应用初始化配置 ----------------------
app = Flask(__name__)
DB_FILE = 'teaching_achievement.db'  # 数据库文件路径
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_FILE}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'teaching-achievement-2026-key'  # 生产环境需替换为随机密钥
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB上传限制

# 创建上传目录
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# 数据库初始化
db = SQLAlchemy(app)
migrate = Migrate(app, db)  # 可选，用于生产环境迁移


# ---------------------- 2. 数据库模型设计（修正团队负责人关联） ----------------------
class User(db.Model):
    """用户表（包含所有用户信息字段）"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    gender = db.Column(db.String(10))
    birth_date = db.Column(db.Date)
    # 关键修复：将 id_card 的默认值设为 NULL，且空值不触发唯一约束
    id_card = db.Column(db.String(18), unique=True, nullable=True, default=None)
    email = db.Column(db.String(100), unique=True, nullable=False)
    phone = db.Column(db.String(20))
    office_phone = db.Column(db.String(20))
    school = db.Column(db.String(100))
    college = db.Column(db.String(100))
    department = db.Column(db.String(100))
    research_room = db.Column(db.String(100))
    role = db.Column(db.String(20), nullable=False, default='teacher')
    api_config = db.Column(db.Text, default='{}')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # 密码加密/验证
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    # 获取大模型API配置
    def get_api_config(self):
        try:
            return json.loads(self.api_config)
        except:
            return {}

    # 设置大模型API配置
    def set_api_config(self, config):
        self.api_config = json.dumps(config, ensure_ascii=False)

    # 关联关系修正：
    # 1. 反向引用：当前用户管理的所有团队（Team表的leader_id关联）
    managed_teams = db.relationship('Team', backref='leader', foreign_keys='Team.leader_id', lazy='dynamic')
    # 2. 反向引用：当前用户加入的所有团队（UserTeam表关联）
    joined_teams = db.relationship('UserTeam', backref='user', foreign_keys='UserTeam.user_id', lazy='dynamic')


class Team(db.Model):
    """团队表（核心关联团队负责人）"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)  # 团队名称
    leader_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 团队负责人ID（核心外键）
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # 关联关系修正：
    # 1. 正向引用：团队负责人（关联User表）
    # leader = db.relationship('User', backref='managed_teams', foreign_keys=[leader_id])  # 原写法保留也可，二选一
    # 2. 反向引用：团队下的所有成员（通过UserTeam关联）
    members = db.relationship('UserTeam', backref='team', foreign_keys='UserTeam.team_id', lazy='dynamic')


class UserTeam(db.Model):
    """用户-团队关联表（多对多，区分负责人和普通成员）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 关联用户
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)  # 关联团队
    join_time = db.Column(db.DateTime, default=datetime.utcnow)
    is_admin = db.Column(db.Boolean, default=False)  # 扩展：是否为团队管理员（非负责人）

    # 联合唯一索引（一个用户只能加入一个团队一次）
    __table_args__ = (db.UniqueConstraint('user_id', 'team_id', name='_user_team_uc'),)


class AchievementContributor(db.Model):
    """成果关联人表（多对多，支持一个成果有多个关联用户）"""
    id = db.Column(db.Integer, primary_key=True)
    achievement_type = db.Column(db.String(50), nullable=False,
                                 comment='成果类型：journal_paper/conference_paper/textbook/monograph/etc.')
    achievement_id = db.Column(db.Integer, nullable=False, comment='成果 ID')
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='关联用户 ID')
    contributor_role = db.Column(db.String(50), default='author',
                                 comment='贡献角色：author/corresponding_author/editor/compile/etc.')
    is_creator = db.Column(db.Boolean, default=False, comment='是否为录入者')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('achievement_type', 'achievement_id', 'user_id', name='_achievement_user_uc'),
    )

    user = db.relationship('User', backref='achievement_contributions')


class InclusionType(db.Model):
    """论文收录类型表（字典表）"""
    id = db.Column(db.Integer, primary_key=True)
    type_name = db.Column(db.String(100), unique=True, nullable=False, comment='收录类型名称')
    type_code = db.Column(db.String(50), unique=True, nullable=False, comment='收录类型代码')
    description = db.Column(db.Text, comment='描述说明')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # 反向引用
    journal_papers = db.relationship('JournalPaper', secondary='journal_paper_inclusion_relation', back_populates='inclusion_types')


class JournalPaperInclusionRelation(db.Model):
    """期刊论文 - 收录类型关联表（多对多）"""
    id = db.Column(db.Integer, primary_key=True)
    paper_id = db.Column(db.Integer, db.ForeignKey('journal_paper.id'), nullable=False)
    inclusion_type_id = db.Column(db.Integer, db.ForeignKey('inclusion_type.id'), nullable=False)
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    # 联合唯一索引（同一篇论文的同一收录类型只能有一条记录）
    __table_args__ = (db.UniqueConstraint('paper_id', 'inclusion_type_id', name='_paper_inclusion_uc'),)

    # 关联关系
    paper = db.relationship('JournalPaper', back_populates='inclusion_relations', overlaps='journal_papers')
    inclusion_type = db.relationship('InclusionType', backref=db.backref('paper_relations', overlaps='journal_papers'), overlaps='journal_papers')


class JournalPaper(db.Model):
    """期刊论文表（完整字段）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 论文名称
    authors = db.Column(db.Text, nullable=False)  # 论文作者（逗号分隔）
    corresponding_authors = db.Column(db.Text)  # 通讯作者（逗号分隔）
    journal_name = db.Column(db.String(200), nullable=False)  # 期刊名称
    inclusion_status = db.Column(db.Text)  # 收录情况（逗号分隔，保留兼容旧数据）
    inclusion_type_ids = db.Column(db.Text, default='', comment='收录类型 ID（逗号分隔，关联 inclusion_type 表）')
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，包含录入者和共同作者）')
    year = db.Column(db.Integer)  # 年
    volume = db.Column(db.String(50))  # 卷
    issue = db.Column(db.String(50))  # 期
    page_range = db.Column(db.String(50))  # 起止页码
    doi = db.Column(db.String(200))  # DOI
    publish_year = db.Column(db.Integer)  # 发表年份
    publish_date = db.Column(db.Date)  # 发表日期
    attachment = db.Column(db.String(256))  # 论文附件路径
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='journal_papers_owned', foreign_keys=[user_id])
    inclusion_relations = db.relationship('JournalPaperInclusionRelation', back_populates='paper', cascade='all, delete-orphan', overlaps='journal_papers')
    inclusion_types = db.relationship('InclusionType', secondary='journal_paper_inclusion_relation', back_populates='journal_papers', viewonly=True)
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(JournalPaper.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='journal_paper')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')



class ConferencePaper(db.Model):
    """会议论文表（完整字段）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 论文名称
    authors = db.Column(db.Text, nullable=False)  # 论文作者（逗号分隔）
    corresponding_authors = db.Column(db.Text)  # 通讯作者（逗号分隔）
    conference_name = db.Column(db.String(200), nullable=False)  # 会议名称
    conference_time = db.Column(db.Text)  # 会议时间（文本格式，如"2026.03.11-2026.03.13"）
    conference_start_date = db.Column(db.Date)  # 会议开始日期
    conference_end_date = db.Column(db.Date)  # 会议结束日期
    conference_place = db.Column(db.String(200))  # 会议地点
    page_range = db.Column(db.String(50))  # 起止页码
    doi = db.Column(db.String(200))  # DOI
    publish_year = db.Column(db.Integer)  # 发表年份
    attachment = db.Column(db.String(256))  # 论文附件路径
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，包含录入者和共同作者）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='conference_papers_owned', foreign_keys=[user_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(ConferencePaper.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='conference_paper')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class TextbookLevel(db.Model):
    """教材级别表（字典表）"""
    id = db.Column(db.Integer, primary_key=True)
    level_name = db.Column(db.String(100), unique=True, nullable=False, comment='级别名称')
    level_code = db.Column(db.String(50), unique=True, nullable=False, comment='级别代码')
    description = db.Column(db.Text, comment='描述说明')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # 反向引用
    textbooks = db.relationship('Textbook', back_populates='level')


class Textbook(db.Model):
    """教材表（更新字段）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 教材名称
    textbook_series = db.Column(db.String(500))  # 教材系列（顿号分隔）
    chief_editor = db.Column(db.Text)  # 主编（顿号分隔）
    associate_editors = db.Column(db.Text)  # 副主编（顿号分隔）
    editorial_board = db.Column(db.Text)  # 编委（顿号分隔）
    publisher = db.Column(db.String(200))  # 出版社
    isbn = db.Column(db.String(50))  # ISBN
    cip_number = db.Column(db.String(100))  # CIP 核字号
    publication_year = db.Column(db.Integer)  # 出版年份
    publication_month = db.Column(db.Integer)  # 出版月份
    publish_date = db.Column(db.Date)  # 出版日期（兼容原有字段）
    edition = db.Column(db.String(50))  # 版次
    word_count = db.Column(db.String(50))  # 字数（如 318 千字）
    price = db.Column(db.String(20))  # 定价（如 49.00）
    textbook_level_id = db.Column(db.Integer, db.ForeignKey('textbook_level.id'), comment='教材级别 ID（关联 textbook_level 表）')
    textbook_type = db.Column(db.String(50))  # 教材类型（下拉选择）
    applicable_majors = db.Column(db.Text)  # 适用专业
    remarks = db.Column(db.Text)  # 备注
    textbook_attachment = db.Column(db.String(256))  # 教材附件（替换原 attachment）
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，包含录入者和编者）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='textbooks_owned', foreign_keys=[user_id])
    level = db.relationship('TextbookLevel', back_populates='textbooks')
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(Textbook.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='textbook')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')



class Monograph(db.Model):
    """专著表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 专著名称
    textbook_series = db.Column(db.String(500))  # 专著系列（顿号分隔）
    chief_editor = db.Column(db.Text)  # 主编（顿号分隔）
    associate_editors = db.Column(db.Text)  # 副主编（顿号分隔）
    editorial_board = db.Column(db.Text)  # 编委（顿号分隔）
    publisher = db.Column(db.String(200))  # 出版社
    isbn = db.Column(db.String(50))  # ISBN
    cip_number = db.Column(db.String(100))  # CIP 核字号
    publication_year = db.Column(db.Integer)  # 出版年份
    publication_month = db.Column(db.Integer)  # 出版月份
    publish_date = db.Column(db.Date)  # 出版日期（兼容扩展）
    edition = db.Column(db.String(50))  # 版次（如第 1 版、修订版）
    word_count = db.Column(db.String(50))  # 字数（如"318 千字"）
    price = db.Column(db.String(20))  # 定价（如"49.00 元"）
    monograph_type = db.Column(db.String(50))  # 专著类型（下拉选择）
    applicable_majors = db.Column(db.Text)  # 适用专业
    remarks = db.Column(db.Text)  # 备注
    monograph_attachment = db.Column(db.String(256))  # 专著附件
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，包含录入者和著者）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='monographs_owned', foreign_keys=[user_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(Monograph.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='monograph')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class ProjectType(db.Model):
    """项目类型表"""
    id = db.Column(db.Integer, primary_key=True)
    type_name = db.Column(db.String(100), nullable=False, comment='项目类型名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<ProjectType {self.type_name}>'


class ProjectStatus(db.Model):
    """项目状态表"""
    id = db.Column(db.Integer, primary_key=True)
    status_name = db.Column(db.String(50), nullable=False, comment='项目状态名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<ProjectStatus {self.status_name}>'


class ProjectLevel(db.Model):
    """项目级别表"""
    id = db.Column(db.Integer, primary_key=True)
    level_name = db.Column(db.String(50), nullable=False, comment='项目级别名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<ProjectLevel {self.level_name}>'


class ProjectCategory(db.Model):
    """项目类别表"""
    id = db.Column(db.Integer, primary_key=True)
    category_name = db.Column(db.String(100), nullable=False, comment='项目类别名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<ProjectCategory {self.category_name}>'


class TeachingProject(db.Model):
    """教研教改和课程建设项目表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False, comment='项目名称')
    project_code = db.Column(db.String(100), comment='项目编号')
    project_leader = db.Column(db.String(100), comment='项目负责人')
    project_members = db.Column(db.Text, comment='项目参与人（顿号分隔）')
    approval_department = db.Column(db.String(200), comment='项目批准部门')
    approval_date = db.Column(db.Date, comment='项目立项时间')
    project_type_id = db.Column(db.Integer, db.ForeignKey('project_type.id'), comment='项目类型 ID')
    project_level_id = db.Column(db.Integer, db.ForeignKey('project_level.id'), comment='项目级别 ID')
    project_category_id = db.Column(db.Integer, db.ForeignKey('project_category.id'), comment='项目类别 ID')
    project_status_id = db.Column(db.Integer, db.ForeignKey('project_status.id'), comment='项目状态 ID')
    funding = db.Column(db.Numeric(10, 2), comment='项目经费（元）')
    start_date = db.Column(db.Date, comment='项目开始时间')
    end_date = db.Column(db.Date, comment='项目结束时间')
    attachment = db.Column(db.String(256), comment='附件路径')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    # 关联关系
    user = db.relationship('User', backref='teaching_projects_owned', foreign_keys=[user_id])
    project_type = db.relationship('ProjectType', backref='teaching_projects')
    project_level = db.relationship('ProjectLevel', backref='teaching_projects')
    project_category = db.relationship('ProjectCategory', backref='teaching_projects')
    project_status = db.relationship('ProjectStatus', backref='teaching_projects')

    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(TeachingProject.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='teaching_project')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class PatentType(db.Model):
    """专利类型表"""
    id = db.Column(db.Integer, primary_key=True)
    type_name = db.Column(db.String(50), unique=True, nullable=False, comment='专利类型名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<PatentType {self.type_name}>'


class PatentStatus(db.Model):
    """专利状态表"""
    id = db.Column(db.Integer, primary_key=True)
    status_name = db.Column(db.String(50), unique=True, nullable=False, comment='专利状态名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<PatentStatus {self.status_name}>'


class Patent(db.Model):
    """专利表（发明/实用新型/外观）"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 专利名称
    patent_type_id = db.Column(db.Integer, db.ForeignKey('patent_type.id'), nullable=False, comment='专利类型 ID（关联 patent_type 表）')
    patentee = db.Column(db.String(200), comment='专利权人')
    address = db.Column(db.String(500), comment='地址')
    inventors = db.Column(db.Text, comment='发明人（多人用分号分隔）')
    patent_status_id = db.Column(db.Integer, db.ForeignKey('patent_status.id'), comment='专利状态 ID（关联 patent_status 表）')
    patent_number = db.Column(db.String(100), unique=True, comment='专利号')
    grant_announcement_number = db.Column(db.String(100), comment='授权公告号')
    apply_date = db.Column(db.Date, comment='专利申请日')
    grant_announcement_date = db.Column(db.Date, comment='授权公告日')
    applicant_at_apply_date = db.Column(db.String(200), comment='申请日时申请人（发明专利）')
    inventor_at_apply_date = db.Column(db.String(200), comment='申请日时发明人（发明专利）')
    attachment = db.Column(db.String(256), comment='附件路径')
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，发明人）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='patents_owned', foreign_keys=[user_id])
    patent_type = db.relationship('PatentType', backref='patents')
    patent_status = db.relationship('PatentStatus', backref='patents')
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(Patent.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='patent')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')





class SoftwareCopyright(db.Model):
    """软件著作表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 软件名称
    copyright_owner = db.Column(db.Text)  # 著作权人（多人用分号分隔）
    completion_date = db.Column(db.Date)  # 开发完成日期
    first_publication_date = db.Column(db.Date)  # 首次发表日期
    right_acquisition_method = db.Column(db.String(100))  # 权利取得方式
    right_scope = db.Column(db.String(200))  # 权利范围
    copyright_number = db.Column(db.String(100))  # 登记号
    certificate_number = db.Column(db.String(100))  # 证书号
    register_date = db.Column(db.Date)  # 登记日期
    attachment = db.Column(db.String(256))  # 附件路径
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，著作权人）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='software_copyrights_owned', foreign_keys=[user_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(SoftwareCopyright.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='software_copyright')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class TeachingAchievementAward(db.Model):
    """教学成果获奖表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False, comment='成果名称')
    achievement_type_id = db.Column(db.Integer, db.ForeignKey('teaching_achievement_type.id'), nullable=True, comment='教学成果奖类型 ID（关联 teaching_achievement_type 表）')
    achievement_level_id = db.Column(db.Integer, db.ForeignKey('achievement_level.id'), nullable=True, comment='成果等级 ID（关联 achievement_level 表）')
    main_contributors = db.Column(db.Text, comment='主要完成人（多人用分号分隔）')
    completing_units = db.Column(db.Text, comment='成果完成单位（多个用分号分隔）')
    award_year = db.Column(db.Integer, comment='获奖年度')
    award_rank_id = db.Column(db.Integer, db.ForeignKey('award_rank.id'), nullable=True, comment='获奖等级 ID（关联 award_rank 表）')
    certificate_number = db.Column(db.String(100), comment='证书编号')
    awarding_unit = db.Column(db.String(200), comment='颁奖单位')
    award_date = db.Column(db.Date, comment='获奖日期')
    attachment = db.Column(db.String(256), comment='附件路径')
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，获奖完成人）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='teaching_achievement_awards_owned', foreign_keys=[user_id])
    achievement_type = db.relationship('TeachingAchievementType', backref='teaching_achievement_awards')
    achievement_level = db.relationship('AchievementLevel', backref='teaching_achievement_awards')
    award_rank = db.relationship('AwardRank', backref='teaching_achievement_awards')
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(TeachingAchievementAward.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='teaching_achievement_award')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class TeachingAchievementType(db.Model):
    """教学成果奖类型表（管理员可维护）"""
    id = db.Column(db.Integer, primary_key=True)
    type_name = db.Column(db.String(100), unique=True, nullable=False, comment='类型名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<TeachingAchievementType {self.type_name}>'


class AchievementLevel(db.Model):
    """成果等级表（管理员可维护）"""
    id = db.Column(db.Integer, primary_key=True)
    level_name = db.Column(db.String(50), unique=True, nullable=False, comment='等级名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<AchievementLevel {self.level_name}>'


class AwardRank(db.Model):
    """获奖等级表（管理员可维护，多处共用）"""
    id = db.Column(db.Integer, primary_key=True)
    rank_name = db.Column(db.String(50), unique=True, nullable=False, comment='等级名称')
    sort_order = db.Column(db.Integer, default=0, comment='排序顺序')
    is_active = db.Column(db.Boolean, default=True, comment='是否启用')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<AwardRank {self.rank_name}>'


class TeachingCompetitionAward(db.Model):
    """教学竞赛获奖表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    title = db.Column(db.String(500), nullable=False)  # 竞赛名称
    award_year = db.Column(db.String(50))  # 获奖年度
    competition_level_id = db.Column(db.Integer, db.ForeignKey('achievement_level.id'))  # 竞赛等级 id
    award_rank_id = db.Column(db.Integer, db.ForeignKey('award_rank.id'))  # 获奖等级 id
    winners = db.Column(db.Text)  # 获奖人（多个用分号分隔）
    winner_unit = db.Column(db.String(500))  # 获奖人所在单位
    competition_name = db.Column(db.String(200))  # 竞赛主办方
    award_date = db.Column(db.Date)  # 获奖日期
    certificate_number = db.Column(db.String(100))  # 证书编号
    attachment = db.Column(db.String(256))  # 附件路径
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，参赛教师）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='teaching_competition_awards_owned', foreign_keys=[user_id])
    competition_level = db.relationship('AchievementLevel', backref='teaching_competition_awards_competition_level', foreign_keys=[competition_level_id])
    award_rank = db.relationship('AwardRank', backref='teaching_competition_awards_award_rank', foreign_keys=[award_rank_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(TeachingCompetitionAward.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='teaching_competition_award')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')


class StudentGuidanceAward(db.Model):
    """指导学生获奖表"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, comment='录入用户 ID')
    award_year = db.Column(db.String(50))  # 获奖年度
    title = db.Column(db.String(500), nullable=False)  # 获奖名称
    competition_name = db.Column(db.String(200))  # 竞赛名称
    competition_level_id = db.Column(db.Integer, db.ForeignKey('achievement_level.id'))  # 竞赛等级 id
    award_rank_id = db.Column(db.Integer, db.ForeignKey('award_rank.id'))  # 获奖等级 id
    student_name = db.Column(db.String(200))  # 获奖学生
    project_name = db.Column(db.String(500))  # 获奖项目名称
    teacher_name = db.Column(db.String(200))  # 指导教师
    student_unit = db.Column(db.String(500))  # 获奖学生所在单位
    organizer = db.Column(db.String(500))  # 竞赛主办方
    certificate_number = db.Column(db.String(100))  # 证书编号
    award_date = db.Column(db.Date)  # 获奖日期
    attachment = db.Column(db.String(256))  # 附件路径
    related_personnel_ids = db.Column(db.Text, default='', comment='关联人员 ID（逗号分隔，关联 user 表的 id，指导教师）')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    public_team_ids = db.Column(db.Text, default='', comment='公开给的团队 ID（逗号分隔）')

    user = db.relationship('User', backref='student_guidance_awards_owned', foreign_keys=[user_id])
    competition_level = db.relationship('AchievementLevel', backref='student_guidance_awards_competition_level', foreign_keys=[competition_level_id])
    award_rank = db.relationship('AwardRank', backref='student_guidance_awards_award_rank', foreign_keys=[award_rank_id])
    contributors = db.relationship('AchievementContributor',
                                   primaryjoin="and_(StudentGuidanceAward.id==AchievementContributor.achievement_id, AchievementContributor.achievement_type=='student_guidance_award')",
                                   foreign_keys=[AchievementContributor.achievement_id],
                                   viewonly=True,
                                   overlaps='user')

class SystemConfig(db.Model):
    """系统全局配置表（仅管理员可维护）"""
    id = db.Column(db.Integer, primary_key=True)
    config_key = db.Column(db.String(100), unique=True, nullable=False)  # 配置项标识（如system_name、max_upload_size）
    config_value = db.Column(db.Text)  # 配置值（字符串/JSON）
    config_desc = db.Column(db.String(200))  # 配置项描述
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    update_time = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    @classmethod
    def get_config(cls, key, default=''):
        """获取系统配置值（无则返回默认值）"""
        config = cls.query.filter_by(config_key=key).first()
        return config.config_value if config else default

    @classmethod
    def set_config(cls, key, value, desc=''):
        """设置系统配置值（不存在则创建，存在则更新）"""
        config = cls.query.filter_by(config_key=key).first()
        if not config:
            config = cls(config_key=key, config_value=value, config_desc=desc)
            db.session.add(config)
        else:
            config.config_value = value
            config.config_desc = desc
        db.session.commit()
        return config


# ---------------------- 3. 辅助函数 ----------------------
def get_current_user():
    """获取当前登录用户 - 修复SQLAlchemy 2.0警告"""
    if 'user_id' in session:
        # 替换过时的 Query.get() 为 Session.get()
        return db.session.get(User, session['user_id'])
    return None


def generate_nav_menu(user):
    """生成左侧导航菜单（根据角色）"""
    # 基础菜单（所有登录用户可见）
    base_menu = [
        '<li><a href="/">首页</a></li>'
    ]


    if user.role == 'teacher':
        base_menu.append('<li><a href="/user/settings">个人账户设置</a></li>')
        base_menu.append('<li><a href="/stats/dashboard">📊 数据统计仪表盘</a></li>')
        base_menu.append('<li><a href="/achievement/ocr_import">📷 OCR智能导入</a></li>')
        base_menu.append('<li><a href="/achievement/voice_export">🎙️ 语音导出</a></li>')

    elif user.role == 'team_leader':
        base_menu.append('<li><a href="/user/settings">个人账户设置</a></li>')
        base_menu.append('<li><a href="/team/voice_export">🎙️ 团队语音导出</a></li>')

    # 成果管理菜单（仅普通教师可见）
    achievement_menu = [
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">📄 论文管理</span>',  # 添加点击事件
        '  <ul class="submenu">',
        '    <li><a href="/achievement/journal_paper">期刊论文</a></li>',
        '    <li><a href="/achievement/conference_paper">会议论文</a></li>',
        '  </ul>',
        '</li>',
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">📚 教材与专著</span>',  # 添加点击事件
        '  <ul class="submenu">',
        '    <li><a href="/achievement/textbook">教材</a></li>',
        '    <li><a href="/achievement/monograph">专著</a></li>',
        '  </ul>',
        '</li>',
        '<li><a href="/achievement/teaching_project">🔬 教研教改和课程建设项目</a></li>',
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">💡 专利与软著</span>',  # 添加点击事件
        '  <ul class="submenu">',
        '    <li><a href="/achievement/patent">专利</a></li>',
        '    <li><a href="/achievement/software_copyright">软件著作</a></li>',
        '  </ul>',
        '</li>',
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">🏆 获奖管理</span>',  # 添加点击事件
        '  <ul class="submenu">',
        '    <li><a href="/achievement/teaching_achievement_award">教学成果获奖</a></li>',
        '    <li><a href="/achievement/teaching_competition_award">教学竞赛获奖</a></li>',
        '    <li><a href="/achievement/student_guidance_award">指导学生获奖</a></li>',
        '  </ul>',
        '</li>',
    ]

    # 管理员专属菜单（核心修改：移除子菜单，直接显示用户/团队管理）
    admin_menu = [
        '<li><a href="/admin/user_manage">👥 用户管理</a></li>',
        '<li><a href="/admin/team_manage">🏢 团队管理</a></li>',
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">⚙️ 字典管理</span>',
        '  <ul class="submenu">',
        '    <li><a href="/admin/dict_manage/achievement_type">教学成果奖类型</a></li>',
        '    <li><a href="/admin/dict_manage/achievement_level">成果等级</a></li>',
        '    <li><a href="/admin/dict_manage/award_rank">获奖等级</a></li>',
        '  </ul>',
        '</li>'
    ]

    # 团队负责人专属菜单（仅保留带自子菜单的团队管理）
    leader_menu = [
        '<li class="menu-group">',
        '  <span onclick="toggleSubmenu(this)">👥 团队管理</span>',
        '  <ul class="submenu">',
        '    <li><a href="/team/list">📋 查看团队</a></li>',  # 重点：指向新的团队列表页面
        '    <li><a href="/team/achievements">📊 团队成果统计</a></li>',
        '    <li><a href="/team/manage_members">👨‍🏫 团队成员管理</a></li>',
        '    <li><a href="/team/member_achievements">📋 成员成果详情</a></li>',
        '  </ul>',
        '</li>'
    ]

    # 拼接最终菜单
    menu_html = '<ul class="sidebar-menu">'
    menu_html += ''.join(base_menu)

    # 仅普通教师显示成果管理菜单
    if user.role == 'teacher':
        menu_html += ''.join(achievement_menu)
    # 仅团队负责人显示团队管理菜单
    elif user.role == 'team_leader':
        menu_html += ''.join(leader_menu)

    # 管理员显示系统管理菜单
    if user.role == 'admin':
        menu_html += ''.join(admin_menu)

    menu_html += '<li><a href="/logout">🚪 退出登录</a></li></ul>'
    return menu_html


def render_base_layout(title, content, user):
    """渲染基础布局（左右布局，禁止Jinja）"""
    nav_menu = generate_nav_menu(user) if user else ''
    user_info = f'欢迎，{user.username}（{user.role}）' if user else '未登录'

    html = f'''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title} - 教学成果管理系统</title>
    <style>
        /* 全局样式 */
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
            font-family: "Microsoft YaHei", sans-serif;
        }}
        body {{
            background: #f5f7fa;
            min-height: 100vh;
        }}
        /* 左侧导航栏 */
        .sidebar {{
            width: 250px;
            background: #2c3e50;
            color: white;
            min-height: 100vh;
            padding: 20px 0;
            box-shadow: 2px 0 5px rgba(0,0,0,0.1);
            position: fixed;
            top: 0;
            left: 0;
        }}
        .sidebar-header {{
            padding: 0 20px 20px;
            border-bottom: 1px solid rgba(255,255,255,0.1);
            margin-bottom: 20px;
        }}
        .sidebar-header h2 {{
            font-size: 18px;
            font-weight: 600;
        }}
        .sidebar-menu {{
            list-style: none;
            padding: 0 10px;
        }}
        .sidebar-menu li {{
            margin: 5px 0;
        }}
        .sidebar-menu a {{
            display: block;
            padding: 12px 20px;
            color: #ecf0f1;
            text-decoration: none;
            border-radius: 4px;
            transition: all 0.3s;
        }}
        .sidebar-menu a:hover {{
            background: #34495e;
            color: #3498db;
        }}
        .menu-group span {{
            display: block;
            padding: 12px 20px;
            color: #bdc3c7;
            font-weight: 600;
            cursor: pointer;
            transition: color 0.3s;
        }}
        .menu-group span:hover {{
            color: #3498db;
        }}
        .submenu {{
            list-style: none;
            padding-left: 20px;
            display: none; /* 默认隐藏子菜单 */
        }}
        .submenu.active {{
            display: block; /* 激活时显示 */
        }}
        .submenu a {{
            padding: 8px 20px;
            font-size: 14px;
        }}
        /* 右侧内容区 */
        .content {{
            padding: 30px;
            margin-left: 250px;
            width: calc(100% - 250px);
        }}
        .content-header {{
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 1px solid #e9ecef;
        }}
        .content-header h1 {{
            font-size: 24px;
            color: #2c3e50;
        }}
        .user-info {{
            text-align: right;
            margin-bottom: 10px;
            color: #7f8c8d;
        }}
        .container {{
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.05);
            padding: 30px;
            min-height: 500px;
        }}
        /* 表单样式 */
        .form-group {{
            margin-bottom: 20px;
        }}
        label {{
            display: block;
            margin-bottom: 8px;
            font-weight: 600;
            color: #495057;
        }}
        input, select {{
            width: 100%;
            padding: 10px 15px;
            border: 1px solid #ced4da;
            border-radius: 4px;
            font-size: 14px;
        }}
        input:focus, select:focus {{
            outline: none;
            border-color: #3498db;
            box-shadow: 0 0 0 3px rgba(52, 152, 219, 0.1);
        }}
        .btn {{
            display: inline-block;
            padding: 10px 20px;
            background: #3498db;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 14px;
        }}
        .btn:hover {{
            background: #2980b9;
        }}
        /* 提示框样式 */
        .alert {{
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 4px;
        }}
        .alert-success {{
            background: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
        }}
        .alert-danger {{
            background: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
        }}
    </style>
</head>
<body>
    <!-- 左侧导航栏 -->
    <div class="sidebar">
        <div class="sidebar-header">
            <h2>教学成果管理系统</h2>
        </div>
        {nav_menu}
    </div>

    <!-- 右侧内容区 -->
    <div class="content">
        <div class="user-info">{user_info}</div>
        <div class="content-header">
            <h1>{title}</h1>
        </div>
        <div class="container">
            {content}
        </div>
    </div>

    <!-- 新增：子菜单切换脚本 -->
    <script>
        function toggleSubmenu(el) {{
            // 获取点击的span的下一个兄弟元素（submenu）
            const submenu = el.nextElementSibling;
            if (submenu && submenu.classList.contains('submenu')) {{
                // 切换active类，实现显示/隐藏
                submenu.classList.toggle('active');
            }}
        }}
    </script>
</body>
</html>
'''
    return html


def allowed_file(filename):
    """校验上传文件扩展名"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def handle_file_upload(file, sub_folder):
    """通用文件上传处理：返回文件存储路径"""
    if file and allowed_file(file.filename):
        # 创建子目录（按成果类型分类存储）
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], sub_folder)
        if not os.path.exists(upload_path):
            os.makedirs(upload_path)

        # 安全文件名 + 时间戳避免重复
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        new_filename = f"{timestamp}_{filename}"
        file_path = os.path.join(upload_path, new_filename)

        # 保存文件
        file.save(file_path)
        return file_path
    return None


def get_team_user_ids(current_user):
    """获取团队内所有用户 ID（团队负责人用）"""
    if current_user.role != 'team_leader':
        return [current_user.id]

    teams = Team.query.filter_by(leader_id=current_user.id).all()
    team_ids = [t.id for t in teams]

    user_teams = UserTeam.query.filter(UserTeam.team_id.in_(team_ids)).all()
    team_user_ids = [ut.user_id for ut in user_teams] + [current_user.id]
    return team_user_ids


def auto_link_contributors(achievement, achievement_type, authors_str, creator_user_id):
    """自动关联成果与系统用户作者"""
    if not authors_str:
        return

    # 修复：同时支持逗号、分号、顿号等多种分隔符
    import re
    author_names = [name.strip() for name in re.split(r'[;,;,,]', authors_str) if name.strip()]
    related_user_ids = set()

    for author_name in author_names:
        users = User.query.filter(
            (User.username == author_name) |
            (User.employee_id == author_name) |
            (User.email.like(f'%{author_name}%'))
        ).all()

        for user in users:
            related_user_ids.add(user.id)

            contributor = AchievementContributor.query.filter_by(
                achievement_type=achievement_type,
                achievement_id=achievement.id,
                user_id=user.id
            ).first()

            if not contributor:
                contributor = AchievementContributor(
                    achievement_type=achievement_type,
                    achievement_id=achievement.id,
                    user_id=user.id,
                    contributor_role='author',
                    is_creator=(user.id == creator_user_id)
                )
                db.session.add(contributor)

    if related_user_ids:
        achievement.related_personnel_ids = ','.join(map(str, related_user_ids))


def render_achievement_list(model, title, fields_config, current_user):
    """通用成果列表页面渲染"""
    # 权限过滤：普通用户看自己的，团队负责人看团队的，管理员无权查看
    if current_user.role == 'admin':
        # 管理员无权查看成果，直接返回提示
        content = '<div class="alert alert-danger">管理员无权查看/操作教师个人成果！</div>'
        return render_base_layout(title, content, current_user)

    if current_user.role == 'team_leader':
        # 1. 获取当前用户管理的所有团队ID（核心：仅能看公开给自己团队的成果）
        managed_teams = Team.query.filter_by(leader_id=current_user.id).all()
        managed_team_ids = [str(t.id) for t in managed_teams]

        # 2. 获取团队内所有成员ID（用于基础筛选）
        user_teams = UserTeam.query.filter(UserTeam.team_id.in_([t.id for t in managed_teams])).all()
        team_user_ids = [ut.user_id for ut in user_teams] + [current_user.id]

        # 3. 核心查询逻辑：
        # - 成果属于团队成员
        # - 且（是自己的成果 OR 成果公开给当前用户管理的任意团队）
        query = model.query.filter(model.user_id.in_(team_user_ids))
        or_conditions = [model.user_id == current_user.id]

        # 遍历当前用户管理的每个团队ID，检查是否在public_team_ids中
        for team_id in managed_team_ids:
            # 处理public_team_ids格式：",1,2,3," 避免部分匹配（如1匹配10）
            or_conditions.append(
                func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
            )

        query = query.filter(or_(*or_conditions))

    elif current_user.role == 'teacher':
        # 普通教师：仅查看自己的成果
        query = model.query.filter_by(user_id=current_user.id)

    # 分页查询
    page = request.args.get('page', 1, type=int)
    per_page = 10
    pagination = query.order_by(model.update_time.desc()).paginate(page=page, per_page=per_page)
    items = pagination.items

    # 生成列表HTML（添加智能导入按钮）
    table_html = '''
    <div style="margin-bottom:20px;">
        <a href="?action=add" class="btn">新增</a>
        <a href="?action=export" class="btn ">导出</a>
        <a href="?action=stats" class="btn ">统计分析</a>
    '''
    # 根据成果类型添加智能导入按钮
    if model.__name__ == 'JournalPaper':
        table_html += '<a href="/achievement/journal_paper/import" class="btn " style="background:#27ae60;">智能导入</a>'
    elif model.__name__ == 'ConferencePaper':
        table_html += '<a href="/achievement/conference_paper/import" class="btn " style="background:#27ae60;">智能导入</a>'
    table_html += '''
    </div>
    <table style="width:100%; border-collapse: collapse; margin-top:20px;">
        <thead>
            <tr style="background:#f5f7fa;">
                <th style="padding:10px; border:1px solid #dee2e6;">序号</th>
    '''
    # 生成表头
    for field in fields_config:
        table_html += f'<th style="padding:10px; border:1px solid #dee2e6;">{field["label"]}</th>'
    table_html += '''
                <th style="padding:10px; border:1px solid #dee2e6;">操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    # 生成表体
    for idx, item in enumerate(items, 1):
        table_html += '<tr>'
        table_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{idx}</td>'

        # 生成字段值
        for field in fields_config:
            field_name = field['name']
            value = getattr(item, field_name, '')

            # 处理外键关联字段（显示名称而非 ID）
            if field_name == 'achievement_type_id' and value:
                achievement_type = db.session.get(TeachingAchievementType, value)
                value = achievement_type.type_name if achievement_type else value
            elif field_name == 'achievement_level_id' and value:
                achievement_level = db.session.get(AchievementLevel, value)
                value = achievement_level.level_name if achievement_level else value
            elif field_name == 'competition_level_id' and value:
                competition_level = db.session.get(AchievementLevel, value)
                value = competition_level.level_name if competition_level else value
            elif field_name == 'award_rank_id' and value:
                award_rank = db.session.get(AwardRank, value)
                value = award_rank.rank_name if award_rank else value
            elif field_name == 'project_type_id' and value:
                project_type = db.session.get(ProjectType, value)
                value = project_type.type_name if project_type else value
            elif field_name == 'project_level_id' and value:
                project_level = db.session.get(ProjectLevel, value)
                value = project_level.level_name if project_level else value
            elif field_name == 'project_category_id' and value:
                project_category = db.session.get(ProjectCategory, value)
                value = project_category.category_name if project_category else value
            elif field_name == 'project_status_id' and value:
                project_status = db.session.get(ProjectStatus, value)
                value = project_status.status_name if project_status else value
            elif field_name == 'patent_type_id' and value:
                patent_type = db.session.get(PatentType, value)
                value = patent_type.type_name if patent_type else value
            elif field_name == 'patent_status_id' and value:
                patent_status = db.session.get(PatentStatus, value)
                value = patent_status.status_name if patent_status else value
            elif field_name == 'textbook_level_id' and value:
                textbook_level = db.session.get(TextbookLevel, value)
                value = textbook_level.level_name if textbook_level else value

            # 处理 NULL 值，转为空字符串
            if value is None:
                value = ''
            # 特殊处理日期字段
            elif isinstance(value, date) or isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d') if value else ''
            # 附件字段显示下载链接
            elif field_name == 'attachment' and value:
                filename = os.path.basename(value)
                value = f'<a href="/download?path={value}" target="_blank">📎 {filename}</a>' if value else ''

            table_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{value}</td>'

        # 操作列（仅自己的成果可修改/删除）【核心修改：删除语音导出按钮】
        ops = ''
        if item.user_id == current_user.id:
            ops = f'''
                <a href="?action=edit&id={item.id}" class="btn" style="padding:5px 10px; font-size:12px;">修改</a>
                <a href="?action=delete&id={item.id}" class="btn " style="padding:5px 10px; font-size:12px;" onclick="return confirm('确定删除？')">删除</a>
            '''
        else:
            ops = '仅查看'

        table_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{ops}</td>'
        table_html += '</tr>'

    table_html += '''
        </tbody>
    </table>
    '''

    # 分页控件
    pagination_html = '''
    <div style="margin-top:20px; text-align:center;">
    '''
    if pagination.has_prev:
        pagination_html += f'<a href="?page={pagination.prev_num}" class="btn " style="margin:0 5px;">上一页</a>'
    if pagination.has_next:
        pagination_html += f'<a href="?page={pagination.next_num}" class="btn " style="margin:0 5px;">下一页</a>'
    pagination_html += f'''
        <span style="margin:0 10px;">第{pagination.page}页 / 共{pagination.pages}页</span>
    </div>
    '''

    content = table_html + pagination_html
    return render_base_layout(title, content, current_user)


def render_achievement_form(model, title, fields_config, item_id=None):
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('index'))

    item = db.session.get(model, item_id) if item_id else None
    if item and item.user_id != current_user.id:
        flash('无权限修改该成果！', 'danger')
        return redirect(request.referrer or url_for('index'))

    form_html = f'''
    <form method="POST" enctype="multipart/form-data">
        <input type="hidden" name="id" value="{item_id or ''}">
    '''

    for field in fields_config:
        field_name = field['name']
        field_label = field['label']
        field_type = field.get('type', 'text')
        required = 'required' if field.get('required', False) else ''
        current_value = getattr(item, field_name, None) if item else None  # 初始值设为 None

        # 核心修改：处理 NULL 值
        if current_value is None:
            current_value = ''
        # 日期字段处理（核心优化）
        elif field_type == 'date':
            current_value = current_value.strftime('%Y-%m-%d') if current_value else ''  # 空值显示为空字符串，但提交时转为 None

        # 修复 Select 下拉框渲染逻辑（重点）
        if field_type == 'select':
            # 特殊处理：教材级别（从 TextbookLevel 表读取）
            if field_name == 'textbook_level_id':
                levels = TextbookLevel.query.filter_by(is_active=True).order_by(TextbookLevel.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for level in levels:
                    selected = 'selected' if str(level.id) == str(current_value) else ''
                    form_html += f'<option value="{level.id}" {selected}>{level.level_name}</option>'
                form_html += '</select></div>'
                continue

            # 特殊处理：竞赛等级（从 AchievementLevel 表读取）
            if field_name == 'competition_level_id':
                levels = AchievementLevel.query.filter_by(is_active=True).order_by(AchievementLevel.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for level in levels:
                    selected = 'selected' if str(level.id) == str(current_value) else ''
                    form_html += f'<option value="{level.id}" {selected}>{level.level_name}</option>'
                form_html += '</select></div>'
                continue

            # 特殊处理：获奖等级（从 AwardRank 表读取）
            if field_name == 'award_rank_id':
                ranks = AwardRank.query.filter_by(is_active=True).order_by(AwardRank.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for rank in ranks:
                    selected = 'selected' if str(rank.id) == str(current_value) else ''
                    form_html += f'<option value="{rank.id}" {selected}>{rank.rank_name}</option>'
                form_html += '</select></div>'
                continue

            # 特殊处理：专利类型（从 PatentType 表读取，选项为元组格式）
            if field_name == 'patent_type_id':
                patent_types = PatentType.query.order_by(PatentType.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for pt in patent_types:
                    selected = 'selected' if str(pt.id) == str(current_value) else ''
                    form_html += f'<option value="{pt.id}" {selected}>{pt.type_name}</option>'
                form_html += '</select></div>'
                continue

            # 特殊处理：专利状态（从 PatentStatus 表读取，选项为元组格式）
            if field_name == 'patent_status_id':
                patent_statuses = PatentStatus.query.order_by(PatentStatus.sort_order).all()
                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" {required}>'
                form_html += '<option value="">请选择</option>'
                for ps in patent_statuses:
                    selected = 'selected' if str(ps.id) == str(current_value) else ''
                    form_html += f'<option value="{ps.id}" {selected}>{ps.status_name}</option>'
                form_html += '</select></div>'
                continue

            options = field.get('options', [])
            form_html += f'<div class="form-group">'
            form_html += f'<label>{field_label} {"*" if required else ""}</label>'
            form_html += f'<select name="{field_name}" {required}>'
            # 先添加默认空选项
            form_html += '<option value="">请选择</option>'
            # 遍历所有选项并正确设置 selected 状态
            for opt in options:
                # 兼容元组格式 (id, name) 和简单字符串格式
                if isinstance(opt, tuple) and len(opt) == 2:
                    opt_id, opt_name = opt
                    selected = 'selected' if str(current_value) == str(opt_id) else ''
                    form_html += f'<option value="{opt_id}" {selected}>{opt_name}</option>'
                else:
                    # 简单字符串格式
                    selected = 'selected' if str(current_value) == str(opt) else ''
                    form_html += f'<option value="{opt}" {selected}>{opt}</option>'
            form_html += '</select></div>'
            continue


        elif field_type == 'select_multiple':
            # 特殊处理：收录类型多选框（从 InclusionType 表读取）
            if field_name == 'inclusion_type_ids':
                inclusion_types = InclusionType.query.filter_by(is_active=True).order_by(InclusionType.sort_order).all()
                selected_ids = []
                if current_value and current_value.strip():
                    selected_ids = [id_str.strip() for id_str in current_value.split(',') if id_str.strip()]

                form_html += f'<div class="form-group">'
                form_html += f'<label>{field_label} {"*" if required else ""}</label>'
                form_html += f'<select name="{field_name}" multiple size="10" {required}>'
                form_html += '<option value="" style="display:none;"></option>'
                for inc_type in inclusion_types:
                    selected = 'selected' if str(inc_type.id) in selected_ids else ''
                    form_html += f'<option value="{inc_type.id}" {selected}>{inc_type.type_name} ({inc_type.type_code})</option>'
                form_html += '</select>'
                form_html += '<p style="margin-top:5px; color:#666;">提示：按住 Ctrl 键可多选/取消选择</p></div>'
                continue

            # 普通多选框（团队等）
            current_user = get_current_user()
            teams = Team.query.all()  # 可根据权限过滤（如仅显示用户加入/管理的团队）
            selected_ids = []
            if current_value and current_value.strip():
                selected_ids = [id_str.strip() for id_str in current_value.split(',') if id_str.strip()]
            form_html += f'<div class="form-group">'
            form_html += f'<label>{field_label} {"*" if required else ""}</label>'
            form_html += f'<select name="{field_name}" multiple size="5" {required}>'
            form_html += '<option value="" style="display:none;"></option>'  # 新增：隐藏空选项
            for team in teams:
                selected = 'selected' if str(team.id) in selected_ids else ''
                form_html += f'<option value="{team.id}" {selected}>{team.name}</option>'
            form_html += '</select>'
            form_html += '<p style="margin-top:5px; color:#666;">提示：按住 Ctrl 键可多选/取消选择</p></div>'
            # 核心添加：跳过后续普通输入框渲染，避免重复
            continue

        # 处理文件上传字段
        elif field_type == 'file':
            form_html += f'<div class="form-group">'
            form_html += f'<label>{field_label}</label>'
            form_html += f'<input type="file" name="{field_name}" accept=".pdf,.docx,.doc,.png,.jpg,.jpeg">'
            # 显示已上传的文件
            if item and getattr(item, field_name, ''):
                filename = os.path.basename(getattr(item, field_name))
                form_html += f'<p style="margin-top:5px;">当前文件：<a href="/download?path={getattr(item, field_name)}" target="_blank">{filename}</a></p>'
            form_html += '</div>'
            continue

        # 普通输入框（文本/整数/日期）
        form_html += f'<div class="form-group">'
        form_html += f'<label>{field_label} {"*" if required else ""}</label>'
        form_html += f'<input type="{field_type}" name="{field_name}" value="{current_value or ""}" {required}>'
        form_html += '</div>'

    form_html += '''
        <div class="form-group">
            <button type="submit" class="btn">保存</button>
            <a href="javascript:history.back()" class="btn" style="background-color:#95a5a6; margin-left:10px;">取消</a>
        </div>
    </form>
    '''

    return render_base_layout(title, form_html, current_user)


def handle_achievement_submit(model, fields_config):
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('index'))

    item_id = request.form.get('id')
    item = db.session.get(model, item_id) if item_id else None

    if item and item.user_id != current_user.id:
        flash('无权限修改该成果！', 'danger')
        return redirect(request.referrer or url_for('index'))

    if not item:
        item = model()
        item.user_id = current_user.id
        item.create_time = datetime.now()

    for field in fields_config:
        field_name = field['name']
        field_type = field.get('type', 'text')
        if field_type == 'file':
            continue

        value = request.form.get(field_name, '').strip()

        # 处理外键字段（select 类型且字段名以_id 结尾）- 只保存 ID，不保存对象
        if field_type == 'select' and field_name.endswith('_id'):
            if value == '' or value is None:
                value = None
            else:
                try:
                    value = int(value)  # 直接转为整数保存
                    # 验证外键是否存在（可选）
                    if field_name == 'achievement_type_id':
                        related_obj = db.session.get(TeachingAchievementType, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'achievement_level_id':
                        related_obj = db.session.get(AchievementLevel, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'competition_level_id':
                        related_obj = db.session.get(AchievementLevel, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'award_rank_id':
                        related_obj = db.session.get(AwardRank, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'project_type_id':
                        related_obj = db.session.get(ProjectType, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'project_level_id':
                        related_obj = db.session.get(ProjectLevel, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'project_category_id':
                        related_obj = db.session.get(ProjectCategory, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'project_status_id':
                        related_obj = db.session.get(ProjectStatus, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'patent_type_id':
                        related_obj = db.session.get(PatentType, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'patent_status_id':
                        related_obj = db.session.get(PatentStatus, value)
                        if not related_obj:
                            value = None
                    elif field_name == 'textbook_level_id':
                        related_obj = db.session.get(TextbookLevel, value)
                        if not related_obj:
                            value = None
                except Exception:
                    value = None

        if field_type == 'number':
            if value == '':
                value = None
            else:
                try:
                    value = float(value)
                except ValueError:
                    flash(f'{field["label"]}必须是数字！', 'danger')
                    return redirect(request.referrer)

        if field_type == 'integer':
            if value == '':
                value = None
            else:
                try:
                    value = int(value)
                except ValueError:
                    flash(f'{field["label"]}必须是数字！', 'danger')
                    return redirect(request.referrer)


        elif field_type == 'select_multiple':
            selected_ids = request.form.getlist(field_name)
            selected_ids = [id_str.strip() for id_str in selected_ids if id_str.strip() and id_str != '']
            value = ','.join(selected_ids) if selected_ids else ''

            if model == JournalPaper and field_name == 'inclusion_type_ids':
                inclusion_names = []
                for inc_id in selected_ids:
                    inc_type = InclusionType.query.get(int(inc_id))
                    if inc_type:
                        inclusion_names.append(inc_type.type_name)
                item.inclusion_status = ','.join(inclusion_names)

        elif field_type == 'date':
            if value == '':
                value = None
            else:
                try:
                    value = datetime.strptime(value, '%Y-%m-%d').date()
                except ValueError:
                    flash(f'{field["label"]}格式错误（需为 YYYY-MM-DD）！', 'danger')
                    return redirect(request.referrer)

        elif value == '':
            value = None

        setattr(item, field_name, value)


    for field in fields_config:
        if field.get('type') == 'file':
            file = request.files.get(field['name'])
            if file and file.filename:
                old_path = getattr(item, field['name'], '')
                if old_path and os.path.exists(old_path):
                    os.remove(old_path)
                sub_folder = model.__tablename__
                new_path = handle_file_upload(file, sub_folder)
                setattr(item, field['name'], new_path)

    item.update_time = datetime.now()

    try:
        if not item_id:
            db.session.add(item)
        db.session.flush()

        if not item_id:
            achievement_type_map = {
                'journal_paper': JournalPaper,
                'conference_paper': ConferencePaper,
                'textbook': Textbook,
                'monograph': Monograph,
                'teaching_project': TeachingProject,
                'patent': Patent,
                'software_copyright': SoftwareCopyright,
                'teaching_achievement_award': TeachingAchievementAward,
                'teaching_competition_award': TeachingCompetitionAward,
                'student_guidance_award': StudentGuidanceAward
            }

            for ach_type, ach_model in achievement_type_map.items():
                if model == ach_model:
                    authors_field = 'authors' if hasattr(item, 'authors') else 'chief_editor'
                    authors_str = getattr(item, authors_field, '')
                    if authors_str:
                        auto_link_contributors(item, ach_type, authors_str, current_user.id)
                    break

        db.session.commit()

        if model == JournalPaper:
            inclusion_type_ids = request.form.getlist('inclusion_type_ids')
            JournalPaperInclusionRelation.query.filter_by(paper_id=item.id).delete()
            for inc_id in inclusion_type_ids:
                if inc_id.strip():
                    relation = JournalPaperInclusionRelation(paper_id=item.id, inclusion_type_id=int(inc_id.strip()))
                    db.session.add(relation)
            db.session.commit()

        flash(f'{"修改" if item_id else "新增"}成功！', 'success')
        return redirect(url_for(request.endpoint, action='list'))
    except Exception as e:
        db.session.rollback()
        flash(f'操作失败：{str(e)}', 'danger')
        return redirect(request.referrer or url_for('index'))


def handle_achievement_delete(model, item_id):
    """通用成果删除"""
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('index'))

    item = db.session.get(model, item_id)
    if not item or item.user_id != current_user.id:
        flash('无权限删除该成果！', 'danger')
        return redirect(url_for('index'))

    # 删除附件文件
    for field in ['attachment']:
        file_path = getattr(item, field, '')
        if file_path and os.path.exists(file_path):
            os.remove(file_path)

    # 删除数据库记录
    try:
        db.session.delete(item)
        db.session.commit()
        flash('删除成功！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'danger')

    return redirect(request.referrer or url_for('index'))


def render_achievement_stats(model, title, fields_config, current_user):
    """通用成果统计分析页面"""
    user_ids = get_team_user_ids(current_user) if current_user.role != 'admin' else []
    items = model.query.filter(model.user_id.in_(user_ids)).all()

    # 基础统计：总数
    total = len(items)

    # 按年份统计（取publish_year/award_date等日期字段）
    year_stats = {}
    date_field = None
    for field in fields_config:
        if 'year' in field['name'] or 'date' in field['name']:
            date_field = field['name']
            break

    for item in items:
        year = getattr(item, date_field, '')
        if isinstance(year, int) and year:
            year_stats[year] = year_stats.get(year, 0) + 1
        elif isinstance(year, date) and year:
            year = year.year
            year_stats[year] = year_stats.get(year, 0) + 1

    # 生成统计HTML
    stats_html = f'''
    <div class="stats-container">
        <h3>基础统计</h3>
        <p>成果总数：<strong>{total}</strong></p>

        <h3 style="margin-top:20px;">按年份统计</h3>
        <ul>
    '''
    for year, count in sorted(year_stats.items()):
        stats_html += f'<li>{year}年：{count}项</li>'
    stats_html += '''
        </ul>

        <a href="javascript:history.back()" class="btn " style="margin-top:20px;">返回列表</a>
    </div>
    '''

    return render_base_layout(f'{title} - 统计分析', stats_html, current_user)


def export_achievement_excel(model, fields_config, current_user, start_date=None, end_date=None):
    """通用成果Excel导出（支持时间范围筛选）"""
    user_ids = get_team_user_ids(current_user) if current_user.role != 'admin' else []

    # 基础查询：用户权限过滤
    query = model.query.filter(model.user_id.in_(user_ids))

    # 时间范围筛选（核心新增）
    if start_date or end_date:
        # 确定日期字段（根据不同模型的日期字段适配）
        date_field_map = {
            JournalPaper: 'publish_date',
            ConferencePaper: 'conference_time',
            Textbook: 'publish_date',
            Monograph: 'publish_date',
            TeachingProject: 'start_date',
            Patent: 'apply_date',
            SoftwareCopyright: 'register_date',
            TeachingAchievementAward: 'award_date',
            TeachingCompetitionAward: 'award_date',
            StudentGuidanceAward: 'award_date'
        }

        date_field = date_field_map.get(model, None)
        if date_field:
            # 转换字符串日期为date对象
            start_date_obj = None
            end_date_obj = None

            if start_date:
                try:
                    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                except:
                    pass

            if end_date:
                try:
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                except:
                    pass

            # 添加时间筛选条件
            if start_date_obj:
                query = query.filter(getattr(model, date_field) >= start_date_obj)
            if end_date_obj:
                query = query.filter(getattr(model, date_field) <= end_date_obj)

    items = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '成果数据'

    # 表头
    headers = [f['label'] for f in fields_config]
    ws.append(headers)

    # 数据行
    for item in items:
        row = []
        for field in fields_config:
            field_name = field['name']
            value = getattr(item, field_name, '')

            # 处理外键关联字段（显示名称而非 ID）
            if field_name == 'achievement_type_id' and value:
                achievement_type = db.session.get(TeachingAchievementType, value)
                value = achievement_type.type_name if achievement_type else ''
            elif field_name == 'achievement_level_id' and value:
                achievement_level = db.session.get(AchievementLevel, value)
                value = achievement_level.level_name if achievement_level else ''
            elif field_name == 'competition_level_id' and value:
                competition_level = db.session.get(AchievementLevel, value)
                value = competition_level.level_name if competition_level else ''
            elif field_name == 'award_rank_id' and value:
                award_rank = db.session.get(AwardRank, value)
                value = award_rank.rank_name if award_rank else ''
            elif field_name == 'project_type_id' and value:
                project_type = db.session.get(ProjectType, value)
                value = project_type.type_name if project_type else ''
            elif field_name == 'project_level_id' and value:
                project_level = db.session.get(ProjectLevel, value)
                value = project_level.level_name if project_level else ''
            elif field_name == 'project_category_id' and value:
                project_category = db.session.get(ProjectCategory, value)
                value = project_category.category_name if project_category else ''
            elif field_name == 'project_status_id' and value:
                project_status = db.session.get(ProjectStatus, value)
                value = project_status.status_name if project_status else ''
            elif field_name == 'patent_type_id' and value:
                patent_type = db.session.get(PatentType, value)
                value = patent_type.type_name if patent_type else ''
            elif field_name == 'patent_status_id' and value:
                patent_status = db.session.get(PatentStatus, value)
                value = patent_status.status_name if patent_status else ''
            elif field_name == 'textbook_level_id' and value:
                textbook_level = db.session.get(TextbookLevel, value)
                value = textbook_level.level_name if textbook_level else ''

            if value is None:
                value = ''
            elif isinstance(value, (date, datetime)):
                value = value.strftime('%Y-%m-%d') if value else ''
            elif field_name == 'attachment' and value:
                value = os.path.basename(value) if value else ''
            row.append(value)
        ws.append(row)

    # 列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'{model.__tablename__}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
def get_zhipu_api_key(current_user):
    """从用户配置中获取智谱AI API Key"""
    api_config = current_user.get_api_config()
    return api_config.get('zhipu', {}).get('api_key', '')

def ai_analyze_journal_full(citation_text, api_key):
    """AI分析期刊论文引用文本"""
    if not citation_text.strip() or not api_key:
        return {"起止页码": "", "年": "", "卷": "", "期": "", "DOI": ""}

    url = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    prompt = f"""
【任务】从指定的知网期刊论文引用文本（格式：作者.题名[J].刊名,年,卷(期):起止页码.DOI.）中精准提取以下信息：
1. 起止页码
2. 年
3. 卷
4. 期
5. DOI

【输入文本】
{citation_text}

【输出规则】
1. 仅输出标准JSON字符串，无任何多余文字、注释、反引号、说明。
2. JSON必须包含字段："起止页码"、"年"、"卷"、"期"、"DOI"。
3. 起止页码格式：数字-数字，无则为空。
4. 年、卷、期只保留数字，无则为空。
5. DOI只保留编号，去掉DOI:前缀，无则为空。
6. 严格按JSON输出，不要任何多余内容。
    """

    payload = {
        "model": "glm-4-flash",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.0,
        "stream": False
    }

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=15)
        response.raise_for_status()
        result = response.json()
        ai_content = result['choices'][0]['message']['content'].strip()
        ai_content = ai_content.replace('```json', '').replace('```', '').strip()
        data = json.loads(ai_content)

        return {
            "起止页码": data.get("起止页码", "").strip(),
            "年": data.get("年", "").strip(),
            "卷": data.get("卷", "").strip(),
            "期": data.get("期", "").strip(),
            "DOI": data.get("DOI", "").strip()
        }
    except Exception as e:
        print(f"AI分析期刊论文失败：{e}")
        return {"起止页码": "", "年": "", "卷": "", "期": "", "DOI": ""}

def ai_analyze_citation(citation_text, api_key):
    """AI分析会议论文引用文本"""
    if not citation_text.strip() or not api_key:
        return {'会议地点': '', '起止页码': ''}

    url = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    prompt = f"""
【任务】从指定的知网会议论文引用文本（格式：作者.题名[C]//会议主办单位.会议论文集名.会议地点;,出版年:起止页码.DOI.）中精准提取「会议地点」和「起止页码」两类核心信息。
【输入文本】
{citation_text}
【输出规则】
1. 仅输出标准JSON字符串，无任何多余文字；
2. JSON必须包含两个字段："会议地点"和"起止页码"；
3. 会议地点：提取引用文本中的地点/机构信息，无则为空；
4. 起止页码：格式为"数字-数字"，无则为空；
5. 严格按格式输出，示例：{{"会议地点":"湖南中医药大学","起止页码":"10-13"}}。
    """

    payload = {
        "model": "glm-4-flash",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.0,
        "stream": False
    }

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=10)
        response.raise_for_status()
        result = response.json()
        ai_content = result['choices'][0]['message']['content'].strip()
        ai_content = ai_content.replace('```json', '').replace('```', '').strip()
        parsed_ai = json.loads(ai_content)

        return {
            '会议地点': parsed_ai.get('会议地点', '').strip(),
            '起止页码': parsed_ai.get('起止页码', '').strip()
        }
    except Exception as e:
        print(f"AI分析会议论文失败：{e}")
        return {'会议地点': '', '起止页码': ''}

def crawl_cnki_journal(keyword, max_papers=3, driver_path=r'C:\Users\mtlxzmd\OneDrive\桌面\新建文件夹\毕设\msedgedriver.exe'):
    """爬取知网期刊论文"""
    # 浏览器配置
    options = webdriver.EdgeOptions()
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument('--disable-blink-features=AutomationControlled')

    try:
        service = Service(driver_path)
        browser = webdriver.Edge(service=service, options=options)
        browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                Object.defineProperty(navigator, 'languages', {get: () => ['zh-CN', 'zh']});
            """
        })
        browser.implicitly_wait(3)
        actions = ActionChains(browser)
        results = []

        # 提取论文ID
        def extract_paper_id(link):
            try:
                id_match = re.search(r'id=([^&]+)', link)
                return id_match.group(1) if id_match else ""
            except:
                return ""

        # 提取引用格式
        def extract_quote_manual(row):
            citation = ""
            try:
                quote_btn = WebDriverWait(row, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.icon-quote'))
                )
                browser.execute_script("arguments[0].scrollIntoView({block: 'center'});", quote_btn)
                quote_btn.click()

                quote_elem = WebDriverWait(browser, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "td.quote-r"))
                )
                citation = quote_elem.text.strip()

                close_btn = WebDriverWait(browser, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.layui-layer-close.layui-layer-close1'))
                )
                close_btn.click()
            except:
                pass
            return citation

        # 爬取逻辑
        browser.get("https://kns.cnki.net/kns8s/AdvSearch")
        WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'a[name="classify"][resource="JOURNAL"]'))
        )
        browser.find_element(By.CSS_SELECTOR, 'a[name="classify"][resource="JOURNAL"]').click()

        search_input = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-tipid="gradetxt-2"]'))
        )
        search_input.clear()
        search_input.send_keys(keyword)

        search_btn = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input.btn-search'))
        )
        search_btn.click()

        table = WebDriverWait(browser, 60).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'table.result-table-list'))
        )

        rows = table.find_elements(By.CSS_SELECTOR, 'tbody tr')[:max_papers]

        for row in rows:
            paper_data = {
                '论文ID': "", '论文名称': "", '论文作者': "", '通讯作者': "",
                '期刊名称': "", '论文收录情况': "", '年': "", '卷': "", '期': "",
                '起止页码': "", '发表年份': "", '发表日期': "", '引用格式': "", 'DOI': ""
            }
            try:
                title_elem = WebDriverWait(row, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'a.fz14'))
                )
                paper_data['论文名称'] = title_elem.text.strip()
                paper_data['论文ID'] = extract_paper_id(title_elem.get_attribute('href'))
                paper_data['论文作者'] = row.find_element(By.CSS_SELECTOR, 'td.author').text.strip().replace('；', ',')
                paper_data['期刊名称'] = row.find_element(By.CSS_SELECTOR, 'td.source').text.strip()

                # 发表日期
                try:
                    date_elem = row.find_element(By.CSS_SELECTOR, 'td.date')
                    paper_data['发表日期'] = date_elem.text.strip()
                    if paper_data['发表日期']:
                        paper_data['发表年份'] = paper_data['发表日期'].split('-')[0]
                except:
                    pass

                # 提取引用格式
                paper_data['引用格式'] = extract_quote_manual(row)
                results.append(paper_data)
                time.sleep(random.uniform(0.5, 1))

            except Exception as e:
                print(f"爬取单篇期刊论文失败：{e}")
                continue

        browser.quit()
        return results
    except Exception as e:
        print(f"爬取知网期刊论文失败：{e}")
        return []

def crawl_cnki_conference(keyword, max_papers=3, driver_path=r'C:\Users\mtlxzmd\OneDrive\桌面\新建文件夹\毕设\msedgedriver.exe'):
    """爬取知网会议论文"""
    # 浏览器配置
    options = webdriver.EdgeOptions()
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--disable-images')

    try:
        service = Service(driver_path)
        browser = webdriver.Edge(service=service, options=options)
        browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                Object.defineProperty(navigator, 'languages', {get: () => ['zh-CN', 'zh']});
            """
        })
        browser.implicitly_wait(1)
        results = []

        # 提取引用格式
        def extract_quote_manual(row):
            citation = ""
            try:
                quote_btn = WebDriverWait(row, 8).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.icon-quote[title="引用"]'))
                )
                browser.execute_script("arguments[0].scrollIntoView({block: 'center'});", quote_btn)
                browser.execute_script("arguments[0].click();", quote_btn)

                quote_elem = WebDriverWait(browser, 8).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "td.quote-r"))
                )
                citation = quote_elem.text.strip().replace('\n', '').replace('  ', ' ')

                close_btn = WebDriverWait(browser, 8).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.layui-layer-close.layui-layer-close1'))
                )
                close_btn.click()
            except:
                pass
            return citation

        # 爬取逻辑
        browser.get("https://kns.cnki.net/kns8s/AdvSearch")
        WebDriverWait(browser, 8).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'a[name="classify"]'))
        )

        classify_elem = WebDriverWait(browser, 8).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[name="classify"][resource="CONFERENCE"]'))
        )
        classify_elem.click()

        search_input = WebDriverWait(browser, 8).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-tipid="gradetxt-2"]'))
        )
        search_input.clear()
        search_input.send_keys(keyword)

        search_btn = WebDriverWait(browser, 8).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input.btn-search'))
        )
        search_btn.click()

        table = WebDriverWait(browser, 60).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'table.result-table-list'))
        )

        rows = table.find_elements(By.CSS_SELECTOR, 'tbody tr')[:max_papers]

        for idx, row in enumerate(rows):
            paper_data = {
                '论文名称': "", '论文作者': "", '通讯作者': "", '会议名称': "",
                '会议时间': "", '会议地点': "", '起止页码': "", 'DOI': "",
                '发表年份': "", '引用格式': ""
            }
            try:
                # 论文名称
                title_elem = WebDriverWait(row, 8).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'a.fz14'))
                )
                paper_data['论文名称'] = title_elem.text.strip()

                # 论文作者
                try:
                    author_elem = WebDriverWait(row, 3).until(
                        EC.presence_of_element_located((
                            By.CSS_SELECTOR,
                            'td[aria-describedby*="author"], td.authorname, td:nth-child(3)'
                        ))
                    )
                    paper_data['论文作者'] = author_elem.text.strip().replace('；', ',').replace(' ', '')
                except:
                    paper_data['论文作者'] = ""

                # 通讯作者
                try:
                    author_text = paper_data['论文作者']
                    if '通讯作者：' in author_text:
                        paper_data['通讯作者'] = author_text.split('通讯作者：')[1].split(',')[0].strip()
                        paper_data['论文作者'] = author_text.split('通讯作者：')[0].strip().rstrip(',')
                    else:
                        paper_data['通讯作者'] = ""
                except:
                    paper_data['通讯作者'] = ""

                # 会议名称
                try:
                    source_elem = WebDriverWait(row, 3).until(
                        EC.presence_of_element_located((
                            By.CSS_SELECTOR,
                            'td[aria-describedby*="source"], td.conferencename, td:nth-child(4)'
                        ))
                    )
                    paper_data['会议名称'] = source_elem.text.strip()
                except:
                    paper_data['会议名称'] = ""

                # 会议时间
                try:
                    time_elem = WebDriverWait(row, 3).until(
                        EC.presence_of_element_located((
                            By.CSS_SELECTOR,
                            'td[aria-describedby*="meetetime"], td.meetingtime, td.date'
                        ))
                    )
                    paper_data['会议时间'] = time_elem.text.strip()
                except:
                    paper_data['会议时间'] = ""

                # 发表年份
                try:
                    meeting_time = paper_data['会议时间']
                    if meeting_time and len(meeting_time) >= 4:
                        paper_data['发表年份'] = meeting_time[:4]
                    else:
                        paper_data['发表年份'] = ""
                except:
                    paper_data['发表年份'] = ""

                # 引用格式
                paper_data['引用格式'] = extract_quote_manual(row)

                # DOI解析
                try:
                    doi_pattern = r'DOI[:：]?\s*(\d+\.\d+/[\w\-\.]+)'
                    doi_match = re.search(doi_pattern, paper_data['引用格式'], re.IGNORECASE)
                    if doi_match:
                        paper_data['DOI'] = doi_match.group(1).strip()
                except:
                    paper_data['DOI'] = ""

                results.append(paper_data)
                time.sleep(random.uniform(0.5, 1))

            except Exception as e:
                print(f"爬取单篇会议论文失败：{e}")
                continue

        browser.quit()
        return results
    except Exception as e:
        print(f"爬取知网会议论文失败：{e}")
        return []


# ---------------------- OCR/语音核心函数 ----------------------
def get_baidu_token(current_user):
    """从用户配置获取百度API Token"""
    api_config = current_user.get_api_config()
    baidu_api_key = api_config.get('baidu', {}).get('api_key', '')
    baidu_secret_key = api_config.get('baidu', {}).get('secret_key', '')

    if not baidu_api_key or not baidu_secret_key:
        return None, "未配置百度API Key/Secret Key！请先前往个人设置 > 大模型API配置 中配置百度 OCR API。"

    url = f"https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id={baidu_api_key}&client_secret={baidu_secret_key}"
    try:
        response = requests.post(url, verify=False, timeout=10)
        response.raise_for_status()  # 检查 HTTP 状态码
        token_data = response.json()

        # 检查是否成功获取 token
        if "error" in token_data:
            return None, f"百度API 认证失败：{token_data.get('error_description', '未知错误')}"

        return token_data.get("access_token"), None
    except requests.exceptions.Timeout:
        return None, "获取百度 Token 超时，请检查网络连接"
    except requests.exceptions.ConnectionError:
        return None, "无法连接到百度API 服务器，请检查网络"
    except Exception as e:
        logger.error(f"获取百度 Token 异常：{str(e)}")
        return None, f"获取百度 Token 失败：{str(e)}"


def baidu_ocr_recognize(image_path, current_user):
    """百度 OCR识别图片文字"""
    token, err = get_baidu_token(current_user)
    if err:
        return "", err

    try:
        with open(image_path, 'rb') as f:
            image_data = f.read()
        image_base64 = base64.b64encode(image_data).decode('utf-8')
    except FileNotFoundError:
        return "", f"图片文件不存在：{image_path}"
    except Exception as e:
        return "", f"读取图片失败：{str(e)}"

    ocr_url = "https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    params = {
        "access_token": token,
        "image": image_base64,
        "language_type": "CHN_ENG"
    }

    try:
        response = requests.post(ocr_url, headers=headers, data=params, timeout=30)
        response.raise_for_status()  # 检查 HTTP 状态码
        result = response.json()

        # 检查百度API 返回的错误
        if "error_code" in result:
            error_code = result.get('error_code', 'unknown')
            error_msg = result.get('error_msg', '未知错误')

            # 常见错误码解释
            error_explanations = {
                '17': 'QPS 超限，请稍后再试',
                '18': '总请求次数超限',
                '19': '图片大小超限',
                '100': 'Token 无效或过期',
                '110': 'Token 已过期',
            }

            explanation = error_explanations.get(str(error_code), '')
            full_error = f"百度 OCR 调用失败 [错误码{error_code}]：{error_msg}"
            if explanation:
                full_error += f"（{explanation}）"

            logger.error(full_error)
            return "", full_error

        # 正常返回 OCR 结果
        if "words_result" not in result:
            logger.error(f"百度 OCR 返回数据格式异常：{result}")
            return "", "百度 OCR 返回数据格式异常"

        ocr_text = "\n".join([item["words"] for item in result.get("words_result", [])])
        return ocr_text.strip(), None

    except requests.exceptions.Timeout:
        return "", "OCR识别超时，请检查网络连接或重试"
    except requests.exceptions.ConnectionError:
        return "", "无法连接到 OCR 服务器，请检查网络"
    except Exception as e:
        logger.error(f"OCR识别异常：{str(e)}")
        return "", f"OCR识别失败：{str(e)}"


def extract_achievement_info(ocr_text):
    """解析 OCR 文本，提取成果信息（仅识别教研教改和课程建设项目）"""
    if not ocr_text.strip():
        return {
            'type_name': '识别失败',
            'title': '',
            'extra_fields': {},
            'raw_text': ocr_text,
            'confidence': 0.0
        }

    # 规则匹配成果类型（仅教研教改和课程建设项目）
    clean_text = unicodedata.normalize('NFKC', ocr_text)
    clean_text = re.sub(r'\s+', ' ', clean_text)
    lines = [line.strip() for line in ocr_text.split('\n') if line.strip()]
    title = lines[0] if lines else ''

    if '教学改革' in clean_text or '教改' in clean_text:
        return {
            'type_name': '教研教改和课程建设项目',
            'title': title,
            'extra_fields': {},
            'raw_text': ocr_text,
            'confidence': 0.95
        }

    # 匹配关键词
    matched_type = None
    confidence = 0.0

    rule = achievement_rules['教研教改和课程建设项目']

    # 关键词匹配（只要包含任意一个关键词即匹配）
    keyword_matched = [kw for kw in rule['keywords'] if kw in clean_text]
    if keyword_matched:
        matched_type = '教研教改和课程建设项目'
        confidence = min(0.7 + len(keyword_matched) * 0.05, 0.95)

    # 正则匹配（兜底）
    if not matched_type and re.search(rule['pattern'], clean_text, re.IGNORECASE | re.MULTILINE):
        matched_type = '教研教改和课程建设项目'
        confidence = 0.9

    # 如果未匹配到，返回"识别失败"，由 AI 进一步分析
    if not matched_type:
        return {
            'type_name': '识别失败',
            'title': title,
            'extra_fields': {},
            'raw_text': ocr_text,
            'confidence': 0.0
        }

    return {
        'type_name': matched_type,
        'title': title,
        'extra_fields': {},
        'raw_text': ocr_text,
        'confidence': round(confidence, 2)
    }

def audio_to_text(audio_data, current_user):
    """音频转文字（百度语音识别）"""
    # 获取Token
    token, err = get_baidu_token(current_user)
    if err:
        return "", f"获取语音识别Token失败：{err}"

    # WebM转WAV
    f_in_name = None
    f_out_name = None
    try:
        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix=".webm", delete=False) as f_in:
            f_in.write(audio_data)
            f_in_name = f_in.name

        with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as f_out:
            f_out_name = f_out.name

        # FFmpeg路径
        ffmpeg_exe = SystemConfig.get_config('ffmpeg_exe', "D:\\ffmpeg\\bin\\ffmpeg.exe")

        # 转换格式
        cmd = [
            ffmpeg_exe,
            "-i", f_in_name,
            "-ar", "16000",
            "-ac", "1",
            "-sample_fmt", "s16",
            "-y",
            f_out_name
        ]
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=30
        )
        if result.returncode != 0:
            raise RuntimeError(f"FFmpeg转换失败：{result.stderr}")

        # 读取WAV数据
        with open(f_out_name, "rb") as f:
            wav_data = f.read()

    except Exception as e:
        return "", f"音频格式转换失败：{str(e)}"
    finally:
        # 清理临时文件
        if f_in_name and os.path.exists(f_in_name):
            os.unlink(f_in_name)
        if f_out_name and os.path.exists(f_out_name):
            os.unlink(f_out_name)

    # 调用百度语音识别
    audio_base64 = base64.b64encode(wav_data).decode('utf-8')
    params = {
        "format": "wav",
        "rate": 16000,
        "channel": 1,
        "cuid": f"achievement_{current_user.id}",
        "token": token,
        "speech": audio_base64,
        "len": len(wav_data),
        "dev_pid": 1537
    }

    try:
        response = requests.post("https://vop.baidu.com/server_api",
                                 json=params,
                                 headers={"Content-Type": "application/json"},
                                 timeout=10)
        result = response.json()
        if result.get("err_no") != 0:
            return "", f"语音识别失败：{result.get('err_msg', '未知错误')}"
        text = result.get("result", [""])[0]
        return text.strip(), None
    except Exception as e:
        return "", f"语音识别请求失败：{str(e)}"


def parse_voice_command(text):
    """解析语音指令，适配团队导出（支持识别老师姓名+成果类型）"""
    result = {
        "action": None,
        "start_date": None,
        "end_date": None,
        "type_name": None,
        "teacher_name": None,  # 新增：识别老师姓名
        "achievement_name": None,  # 新增：识别成果名称
        "is_my": True,
        "is_team": False,
        "is_teaching": False,
        "type_names": []
    }

    if not text:
        return result

    # 识别操作类型
    export_keywords = ["导出", "下载", "保存", "导出Excel", "下载Excel"]
    if any(keyword in text for keyword in export_keywords):
        result["action"] = "export"
    elif any(keyword in text for keyword in ["筛选", "查看", "查询"]):
        result["action"] = "filter"
    else:
        result["action"] = "export"

    # 识别教学相关
    if any(keyword in text for keyword in ["教学", "教学成果", "教学类"]):
        result["is_teaching"] = True

    # ========== 核心增强：识别老师姓名 ==========
    # 匹配"XX老师"格式
    teacher_pattern = r'([^，。！？\s]+)老师'
    teacher_match = re.search(teacher_pattern, text)
    if teacher_match:
        result["teacher_name"] = teacher_match.group(1).strip()
        result["is_my"] = False  # 指定了老师，不再是导出自己的

    # ========== 核心增强：识别成果名称（可选） ==========
    # 匹配"的XX项目/的XX论文/的XX专利"等格式
    achievement_pattern = r'的([^，。！？\s]+)(项目|论文|专利|软著|教材|专著|获奖)'
    achievement_match = re.search(achievement_pattern, text)
    if achievement_match:
        result["achievement_name"] = achievement_match.group(1).strip() + achievement_match.group(2).strip()

    # ========== 核心修复：增加对「专利」通用关键词的识别 ==========
    type_names = list(achievement_rules.keys())[:-1]  # 排除"其他"
    # 先匹配完整名称（如发明专利）
    for t_name in type_names:
        if t_name in text:
            result["type_name"] = t_name
            break
    # 如果没匹配到，再匹配通用名称
    if not result["type_name"]:
        if "专利" in text:
            result["type_name"] = "专利"  # 匹配通用的"专利"关键词
        elif "软著" in text:
            result["type_name"] = "软著"
        elif "论文" in text:
            if "期刊" in text:
                result["type_name"] = "期刊论文"
            elif "会议" in text:
                result["type_name"] = "会议论文"
            else:
                result["type_name"] = "期刊论文"  # 默认匹配期刊论文
        elif "教材" in text:
            result["type_name"] = "教材"
        elif "专著" in text:
            result["type_name"] = "专著"
        elif "获奖" in text:
            if "教学竞赛" in text:
                result["type_name"] = "教学竞赛获奖"
            elif "指导学生" in text:
                result["type_name"] = "指导学生获奖"
            else:
                result["type_name"] = "教学成果获奖"

    # 识别时间范围
    single_year_pattern = r'(\d{4})年'
    single_year_match = re.search(single_year_pattern, text)
    if single_year_match:
        result["start_date"] = f"{single_year_match.group(1)}-01-01"
        result["end_date"] = f"{single_year_match.group(1)}-12-31"

    # 年份范围
    year_range_pattern = r'(\d{4})年[至|-|到](\d{4})年'
    year_match = re.search(year_range_pattern, text)
    if year_match:
        result["start_date"] = f"{year_match.group(1)}-01-01"
        result["end_date"] = f"{year_match.group(2)}-12-31"

    # 近三年
    if "近三年" in text:
        current_year = datetime.now().year
        result["start_date"] = f"{current_year - 3}-01-01"
        result["end_date"] = f"{current_year}-12-31"

    # 团队/个人
    if any(keyword in text for keyword in ["团队", "集体", "所有成员"]):
        result["is_my"] = False
        result["is_team"] = True

    return result


def create_achievement_from_ocr(ocr_result, current_user):
    """
    根据 OCR+AI 分析结果创建成果记录（填充全量数据库字段）
    """
    zhipu_api_key = get_zhipu_api_key(current_user)

    base_info = extract_achievement_info(ocr_result['raw_text'])

    ai_info = {}
    if zhipu_api_key:
        ai_info = ai_analyze_achievement_text(ocr_result['raw_text'], zhipu_api_key)
        type_name = ai_info.get('type_name', base_info['type_name'])
        title = ai_info.get('title', base_info['title'])
    else:
        type_name = base_info['type_name']
        title = base_info['title']

    type_model_mapping = {
        '期刊论文': JournalPaper,
        '会议论文': ConferencePaper,
        '教材': Textbook,
        '专著': Monograph,
        '发明专利': Patent,
        '实用新型专利': Patent,
        '软著': SoftwareCopyright,
        '教学成果获奖': TeachingAchievementAward,
        '教学竞赛获奖': TeachingCompetitionAward,
        '指导学生获奖': StudentGuidanceAward,
        '教研教改和课程建设项目': TeachingProject
    }

    if type_name not in type_model_mapping:
        return False, f"暂不支持创建{type_name}类型的成果", None, None

    model = type_model_mapping[type_name]
    try:
        achievement = model()
        achievement.user_id = current_user.id
        achievement.title = title if title else f"OCR 识别成果_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        achievement.create_time = datetime.now()
        achievement.update_time = datetime.now()

        if type_name == '期刊论文':
            achievement.authors = ai_info.get('authors', '')
            achievement.corresponding_authors = ai_info.get('corresponding_authors', '')
            achievement.journal_name = ai_info.get('journal_name', '')
            achievement.inclusion_status = ai_info.get('inclusion_status', '')

            if ai_info.get('year'):
                achievement.year = ai_info['year']
            if ai_info.get('publish_year'):
                try:
                    achievement.publish_year = int(ai_info['publish_year'])
                except:
                    pass

            achievement.volume = ai_info.get('volume', '')
            achievement.issue = ai_info.get('issue', '')
            achievement.page_range = ai_info.get('page_range', '')
            achievement.doi = ai_info.get('doi', '')

            if ai_info.get('publish_date'):
                try:
                    achievement.publish_date = datetime.strptime(ai_info['publish_date'], '%Y-%m-%d').date()
                except:
                    pass

        elif type_name == '会议论文':
            achievement.authors = ai_info.get('authors', '')
            achievement.corresponding_authors = ai_info.get('corresponding_authors', '')
            achievement.conference_name = ai_info.get('conference_name', '')

            if ai_info.get('publish_year'):
                try:
                    achievement.publish_year = int(ai_info['publish_year'])
                except:
                    pass

            achievement.page_range = ai_info.get('page_range', '')
            achievement.doi = ai_info.get('doi', '')

            if ai_info.get('conference_time'):
                try:
                    achievement.conference_time = datetime.strptime(ai_info['conference_time'], '%Y-%m-%d').date()
                except:
                    pass
            achievement.conference_place = ai_info.get('conference_place', '')

        elif type_name == '教材':
            achievement.textbook_series = ai_info.get('textbook_series', '')
            achievement.chief_editor = ai_info.get('chief_editor', '')
            achievement.associate_editors = ai_info.get('associate_editors', '')
            achievement.editorial_board = ai_info.get('editorial_board', '')
            achievement.publisher = ai_info.get('publisher', '')
            achievement.isbn = ai_info.get('isbn', '')
            achievement.cip_number = ai_info.get('cip_number', '')

            if ai_info.get('publication_year'):
                try:
                    achievement.publication_year = int(ai_info['publication_year'])
                except:
                    pass
            if ai_info.get('publication_month'):
                try:
                    achievement.publication_month = int(ai_info['publication_month'])
                except:
                    pass

            achievement.edition = ai_info.get('edition', '')
            achievement.word_count = ai_info.get('word_count', '')
            achievement.price = ai_info.get('price', '')
            achievement.textbook_level = ai_info.get('textbook_level', '')
            achievement.textbook_type = ai_info.get('textbook_type', '')
            achievement.applicable_majors = ai_info.get('applicable_majors', '')
            achievement.remarks = ai_info.get('remarks', '')

            if ai_info.get('publish_date'):
                try:
                    achievement.publish_date = datetime.strptime(ai_info['publish_date'], '%Y-%m-%d').date()
                except:
                    pass

        elif type_name == '专著':
            achievement.textbook_series = ai_info.get('textbook_series', '')
            achievement.chief_editor = ai_info.get('chief_editor', '')
            achievement.associate_editors = ai_info.get('associate_editors', '')
            achievement.editorial_board = ai_info.get('editorial_board', '')
            achievement.publisher = ai_info.get('publisher', '')
            achievement.isbn = ai_info.get('isbn', '')
            achievement.cip_number = ai_info.get('cip_number', '')
            achievement.author_role = ai_info.get('author_role', '')

            if ai_info.get('publication_year'):
                try:
                    achievement.publication_year = int(ai_info['publication_year'])
                except:
                    pass
            if ai_info.get('publication_month'):
                try:
                    achievement.publication_month = int(ai_info['publication_month'])
                except:
                    pass

            achievement.edition = ai_info.get('edition', '')
            achievement.word_count = ai_info.get('word_count', '')
            achievement.price = ai_info.get('price', '')
            achievement.monograph_type = ai_info.get('monograph_type', '')
            achievement.applicable_majors = ai_info.get('applicable_majors', '')
            achievement.remarks = ai_info.get('remarks', '')

            if ai_info.get('publish_date'):
                try:
                    achievement.publish_date = datetime.strptime(ai_info['publish_date'], '%Y-%m-%d').date()
                except:
                    pass


        elif type_name in ['发明专利', '实用新型专利', '外观设计专利']:
            # 根据专利类型名称获取类型 ID
            patent_type = PatentType.query.filter_by(type_name=type_name).first()
            if patent_type:
                achievement.patent_type_id = patent_type.id

            # 填充新字段
            achievement.patentee = ai_info.get('patentee', '')
            achievement.address = ai_info.get('address', '')
            achievement.inventors = ai_info.get('inventors', '')
            achievement.grant_announcement_number = ai_info.get('grant_announcement_number', '')
            achievement.applicant_at_apply_date = ai_info.get('applicant_at_apply_date', '')
            achievement.inventor_at_apply_date = ai_info.get('inventor_at_apply_date', '')

            # 基础字段
            achievement.patent_number = ai_info.get('patent_number', '')

            # 根据状态名称获取状态 ID
            status_name = ai_info.get('status', '')
            if status_name:
                patent_status = PatentStatus.query.filter_by(status_name=status_name).first()
                if patent_status:
                    achievement.patent_status_id = patent_status.id

            # 处理日期字段
            if ai_info.get('apply_date'):
                try:
                    achievement.apply_date = datetime.strptime(ai_info['apply_date'], '%Y-%m-%d').date()
                except:
                    pass
            if ai_info.get('grant_announcement_date'):
                try:
                    achievement.grant_announcement_date = datetime.strptime(ai_info['grant_announcement_date'],
                                                                            '%Y-%m-%d').date()
                except:
                    pass

        elif type_name == '软著':
            # 填充新字段
            achievement.copyright_owner = ai_info.get('copyright_owner', '')
            achievement.right_acquisition_method = ai_info.get('right_acquisition_method', '')
            achievement.right_scope = ai_info.get('right_scope', '')
            achievement.certificate_number = ai_info.get('certificate_number', '')
            # 基础字段
            achievement.copyright_number = ai_info.get('copyright_number', '')

            # 处理日期字段
            if ai_info.get('completion_date'):
                try:
                    achievement.completion_date = datetime.strptime(ai_info['completion_date'], '%Y-%m-%d').date()
                except:
                    pass

            if ai_info.get('first_publication_date'):
                try:
                    achievement.first_publication_date = datetime.strptime(ai_info['first_publication_date'],
                                                                           '%Y-%m-%d').date()
                except:
                    pass

            if ai_info.get('register_date'):
                try:
                    achievement.register_date = datetime.strptime(ai_info['register_date'], '%Y-%m-%d').date()
                except:
                    pass



        elif type_name == '教研教改和课程建设项目':
            achievement.project_code = ai_info.get('project_code', '')
            achievement.project_leader = ai_info.get('project_leader', '')
            achievement.project_members = ai_info.get('project_members', '')
            achievement.approval_department = ai_info.get('approval_department', '')
            # 处理项目经费（可以为空）
            funding_str = ai_info.get('funding', '')

            if funding_str and funding_str.strip():
                try:
                    # 去除"万"、"元"等单位，转换为纯数字
                    funding_str = funding_str.replace('万', '0000').replace('元', '').replace(',', '').strip()
                    achievement.funding = float(funding_str)
                except:
                    pass

            # 处理立项时间（支持多种格式）
            if ai_info.get('approval_date'):
                try:
                    achievement.approval_date = datetime.strptime(ai_info['approval_date'], '%Y-%m').date()
                except:
                    pass

            # 处理开始时间
            if ai_info.get('start_date'):
                try:
                    achievement.start_date = datetime.strptime(ai_info['start_date'], '%Y-%m-%d').date()
                except:
                    pass

            # 处理结束时间
            if ai_info.get('end_date'):
                try:
                    achievement.end_date = datetime.strptime(ai_info['end_date'], '%Y-%m-%d').date()
                except:
                    pass

            project_type_name = ai_info.get('project_type_name', '')
            if project_type_name:
                pt = ProjectType.query.filter_by(type_name=project_type_name).first()
                if pt:
                    achievement.project_type_id = pt.id

            project_level_name = ai_info.get('project_level_name', '')
            if project_level_name:
                pl = ProjectLevel.query.filter_by(level_name=project_level_name).first()
                if pl:
                    achievement.project_level_id = pl.id

            project_category_name = ai_info.get('project_category_name', '')
            if project_category_name:
                pc = ProjectCategory.query.filter_by(category_name=project_category_name).first()
                if pc:
                    achievement.project_category_id = pc.id



        elif type_name == '教学成果获奖':

            # 新增字段：主要完成人、成果完成单位、获奖年度、证书编号、颁奖单位

            achievement.main_contributors = ai_info.get('main_contributors', '')

            achievement.completing_units = ai_info.get('completing_units', '')

            achievement.award_year = ai_info.get('award_year')

            achievement.certificate_number = ai_info.get('certificate_number', '')

            achievement.awarding_unit = ai_info.get('awarding_unit', '')

            # 处理外键字段：achievement_type_id

            achievement_type_name = ai_info.get('achievement_type_name', '')

            if achievement_type_name:

                at = TeachingAchievementType.query.filter_by(type_name=achievement_type_name).first()

                if at:
                    achievement.achievement_type_id = at.id

            # 处理外键字段：achievement_level_id

            achievement_level_name = ai_info.get('achievement_level_name', '')

            if achievement_level_name:

                al = AchievementLevel.query.filter_by(level_name=achievement_level_name).first()

                if al:
                    achievement.achievement_level_id = al.id

            # 处理外键字段：award_rank_id

            award_rank_name = ai_info.get('award_rank_name', '')

            if award_rank_name:

                ar = AwardRank.query.filter_by(rank_name=award_rank_name).first()

                if ar:
                    achievement.award_rank_id = ar.id

            # 处理日期

            if ai_info.get('award_date'):

                try:

                    achievement.award_date = datetime.strptime(ai_info['award_date'], '%Y-%m-%d').date()

                except:

                    pass


        elif type_name == '教学竞赛获奖':

            achievement.award_year = ai_info.get('award_year')

            achievement.winners = ai_info.get('winners', '')

            achievement.winner_unit = ai_info.get('winner_unit', '')

            achievement.competition_name = ai_info.get('competition_name', '')

            achievement.certificate_number = ai_info.get('certificate_number', '')

            # 处理外键字段：competition_level_id

            competition_level_name = ai_info.get('competition_level_name', '')

            if competition_level_name:

                cl = AchievementLevel.query.filter_by(level_name=competition_level_name).first()

                if cl:
                    achievement.competition_level_id = cl.id

            # 处理外键字段：award_rank_id

            award_rank_name = ai_info.get('award_rank_name', '')

            if award_rank_name:

                ar = AwardRank.query.filter_by(rank_name=award_rank_name).first()

                if ar:
                    achievement.award_rank_id = ar.id

            if ai_info.get('award_date'):

                try:

                    achievement.award_date = datetime.strptime(ai_info['award_date'], '%Y-%m-%d').date()

                except:

                    pass


        elif type_name == '指导学生获奖':

            achievement.award_year = ai_info.get('award_year')

            achievement.competition_name = ai_info.get('competition_name', '')

            achievement.student_name = ai_info.get('student_name', '')

            achievement.project_name = ai_info.get('project_name', '')

            achievement.teacher_name = ai_info.get('teacher_name', '')

            achievement.student_unit = ai_info.get('student_unit', '')

            achievement.organizer = ai_info.get('organizer', '')

            achievement.certificate_number = ai_info.get('certificate_number', '')

            # 处理外键字段：competition_level_id

            competition_level_name = ai_info.get('competition_level_name', '')

            if competition_level_name:

                cl = AchievementLevel.query.filter_by(level_name=competition_level_name).first()

                if cl:
                    achievement.competition_level_id = cl.id

            # 处理外键字段：award_rank_id

            award_rank_name = ai_info.get('award_rank_name', '')

            if award_rank_name:

                ar = AwardRank.query.filter_by(rank_name=award_rank_name).first()

                if ar:
                    achievement.award_rank_id = ar.id

            if ai_info.get('award_date'):

                try:

                    achievement.award_date = datetime.strptime(ai_info['award_date'], '%Y-%m-%d').date()

                except:

                    pass

        db.session.add(achievement)
        db.session.flush()

        achievement_type_map = {
            '期刊论文': 'journal_paper',
            '会议论文': 'conference_paper',
            '教材': 'textbook',
            '专著': 'monograph',
            '发明专利': 'patent',
            '实用新型专利': 'patent',
            '软著': 'software_copyright',
            '教学成果获奖': 'teaching_achievement_award',
            '教学竞赛获奖': 'teaching_competition_award',
            '指导学生获奖': 'student_guidance_award',
            '教研教改和课程建设项目': 'teaching_project'
        }

        authors_field = 'authors' if hasattr(achievement, 'authors') else 'chief_editor'
        authors_str = getattr(achievement, authors_field, '')
        if authors_str and type_name in achievement_type_map:
            auto_link_contributors(achievement, achievement_type_map[type_name], authors_str, current_user.id)

        db.session.commit()

        return True, f"成功创建{type_name}：{title}（ID：{achievement.id}），已填充{len([k for k, v in ai_info.items() if v])}个字段", type_name, achievement.id
    except Exception as e:
        db.session.rollback()
        logger.error(f"创建 AI 增强版成果失败：{str(e)}")
        return False, f"创建成果失败：{str(e)}", None, None


def ai_analyze_achievement_text(ocr_text, api_key, current_user=None):
    """
    增强版：调用智谱 AI 分析 OCR 文本，提取所有成果类型的全量数据库字段
    :param ocr_text: OCR 识别的原始文本
    :param api_key: 智谱 AI API Key
    :param current_user: 当前登录用户（用于筛选教研教改项目）
    :return: 包含全量字段的结构化字典
    """
    if not ocr_text.strip() or not api_key:
        # 返回全量空字段（匹配数据库模型）
        return {
            # 通用字段
            'type_name': '识别失败',
            'title': '',
            'confidence': 0.0,
            'raw_data': {},

            # 期刊论文专属
            'authors': '',
            'corresponding_authors': '',
            'journal_name': '',
            'inclusion_status': '',
            'year': '',
            'volume': '',
            'issue': '',
            'page_range': '',
            'doi': '',
            'publish_year': '',
            'publish_date': '',

            # 会议论文专属
            'conference_name': '',
            'conference_time': '',
            'conference_place': '',

            # 教材专属
            'textbook_series': '',
            'chief_editor': '',
            'associate_editors': '',
            'editorial_board': '',
            'publisher': '',
            'isbn': '',
            'cip_number': '',
            'publication_year': '',
            'publication_month': '',
            'edition': '',
            'word_count': '',
            'price': '',
            'textbook_level': '',
            'textbook_type': '',
            'applicable_majors': '',
            'remarks': '',

            # 专著专属
            'textbook_series': '',
            'chief_editor': '',
            'associate_editors': '',
            'editorial_board': '',
            'publisher': '',
            'isbn': '',
            'cip_number': '',
            'publication_year': '',
            'publication_month': '',
            'publish_date': '',
            'edition': '',
            'word_count': '',
            'price': '',
            'monograph_type': '',
            'applicable_majors': '',
            'remarks': '',

            # 专利专属
            'patent_type': '',
            'patent_number': '',
            'apply_date': '',
            'grant_date': '',
            'status': '',

            # 软著专属
            'copyright_owner': '',
            'completion_date': '',
            'first_publication_date': '',
            'right_acquisition_method': '',
            'right_scope': '',
            'copyright_number': '',
            'certificate_number': '',
            'register_date': '',

            # 教研教改和课程建设项目专属
            'project_code': '',
            'project_leader': '',
            'project_members': '',
            'approval_department': '',
            'approval_date': '',
            'project_type': '',
            'project_level': '',
            'project_category': '',
            'funding': '',
            'start_date': '',
            'end_date': '',

            # 获奖类专属
            'award_level': '',
            'award_rank': '',
            'award_date': '',
            'competition_name': '',
            'student_name': '',
        }

    url = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    user_name = ''
    if current_user:
        user_name = current_user.username

    # 增强版 Prompt：明确要求提取所有数据库字段
    prompt = f"""
【任务】从以下文本中精准提取教学成果的**所有关键信息**，严格按指定格式输出 JSON 字符串。
【文本内容】
{ocr_text}

【核心要求】
1. 先识别成果类型（必须是以下之一）：
   期刊论文/会议论文/发明专利/实用新型专利/教材/专著/软著/教学成果获奖/教学竞赛获奖/指导学生获奖/教研教改和课程建设项目

2. **重要：如果是教研教改和课程建设项目，且文本中包含多个项目，只提取项目负责人或参与人包含"{user_name}"的项目**
   - 如果表格中有"主持人"或"项目负责人"列，只提取该列包含"{user_name}"的项目
   - 如果表格中有"参加人员"列，也检查是否包含"{user_name}"
   - 如果没有找到"{user_name}"的项目，返回空字段


3. 根据成果类型，提取对应**所有**字段（无信息则为空字符串）：



### 期刊论文字段
- title: 论文名称（必填）
- authors: 作者（多个用逗号分隔）
- corresponding_authors: 通讯作者（多个用逗号分隔）
- journal_name: 期刊名称
- inclusion_status: 收录情况（如 SCI/SSCI/EI/CSSCI/北大核心等）
- year: 发表年（仅数字）
- volume: 卷（仅数字/字符）
- issue: 期（仅数字/字符）
- page_range: 起止页码（如 10-20）
- doi: DOI 编号
- publish_year: 发表年份（仅数字）
- publish_date: 发表日期（格式 YYYY-MM-DD，无则为空）

### 会议论文字段
- title: 论文名称
- authors: 作者
- corresponding_authors: 通讯作者
- conference_name: 会议名称
- conference_time: 会议时间（YYYY-MM-DD）
- conference_place: 会议地点
- page_range: 起止页码
- doi: DOI 编号
- publish_year: 发表年份

### 教材字段
- title: 教材名称
- textbook_series: 教材系列
- chief_editor: 主编
- associate_editors: 副主编
- editorial_board: 编委
- publisher: 出版社
- isbn: ISBN 号
- cip_number: CIP 核字号
- publication_year: 出版年份
- publication_month: 出版月份（仅数字）
- edition: 版次（如第 1 版）
- word_count: 字数（如 318 千字）
- price: 定价（如 49.00）
- textbook_level: 教材级别（国家级规划/全国行业规划/协编/自编/其它）
- textbook_type: 教材类型（纸质/数字）
- applicable_majors: 适用专业
- remarks: 备注

### 专著字段（重点增强）
- title: 专著名称
- textbook_series: 专著系列
- chief_editor: 主编
- associate_editors: 副主编
- editorial_board: 编委
- publisher: 出版社
- isbn: ISBN 号（13 位或 10 位）
- cip_number: CIP 核字号（如"2023 第 XXXXX 号"）
- publication_year: 出版年份（仅数字）
- publication_month: 出版月份（仅数字）
- publish_date: 出版日期（格式 YYYY-MM-DD）
- edition: 版次（如第 1 版、修订版）
- word_count: 字数（如 318 千字）
- price: 定价（如 49.00）
- monograph_type: 专著类型（学术专著/技术专著/科普著作/其它）
- applicable_majors: 适用专业
- remarks: 备注

### 专利字段（发明/实用新型）
- title: 专利名称
- patent_type: 专利类型（发明专利/实用新型专利/外观设计专利）
- patentee: 专利权人
- address: 地址
- inventors: 发明人（多人用分号分隔）
- status: 专利状态（受理/初步审查/公开/实质审查/授权）
- patent_number: 专利号/申请号（如 ZL202412345678.9）
- grant_announcement_number: 授权公告号
- apply_date: 专利申请日（YYYY-MM-DD）
- grant_announcement_date: 授权公告日（YYYY-MM-DD）
- applicant_at_apply_date: 申请日时申请人（发明专利特有）
- inventor_at_apply_date: 申请日时发明人（发明专利特有）


### 软著字段
- title: 软件名称（必填）
- copyright_owner: 著作权人（多人用分号分隔，如"张三；李四；王五"）
- completion_date: 开发完成日期（格式 YYYY-MM-DD）
- first_publication_date: 首次发表日期（格式 YYYY-MM-DD）
- right_acquisition_method: 权利取得方式（如"原始取得"、"受让取得"、"继承取得"等）
- right_scope: 权利范围（如"全部权利"、"部分权利"等）
- copyright_number: 登记号（如"2024SR123456"）
- certificate_number: 证书号（如"软著登字第 1234567 号"）
- register_date: 登记日期（格式 YYYY-MM-DD）


### 教研教改和课程建设项目字段（重点新增）
- title: 项目名称（必填，从项目名单表格中提取）
- project_code: 项目编号（序号或正式编号，如"1"、"2019JG001"等）
- project_leader: 项目负责人/主持人（从表格"主持人"列提取，文本格式，如"李超"）
- project_members: 项目参与人（从表格"参加人员"列提取，多人时用顿号分隔，如"刘增明、黄嘉、赵可、杨华文"）
- approval_department: 项目批准部门（从文件头提取，如"湖南省教育厅"）
- approval_date: 项目立项时间（从文件落款日期提取，格式 YYYY-MM，如"2019-09"）
- project_type_name: 项目类型名称（从以下选择：普通本科高校教学改革研究项目、学位与研究生教育改革研究项目、一流本科课程建设项目、课程思政建设项目、其它）
- project_level_name: 项目级别名称（从以下选择：国家级、省部级、市厅级、校级、院级、其它）
- project_category_name: 项目类别名称（从以下选择：重点项目、一般项目、线上一流课程、线上线下混合式一流课程、线下一流课程、社会实践一流课程、虚拟仿真实验教学一流课程、其它）
- funding: 项目经费（数值型，单位元，如 50000）
- start_date: 项目开始时间（格式 YYYY-MM-DD）
- end_date: 项目结束时间（格式 YYYY-MM-DD）

### 教学成果获奖字段
- title: 成果名称（必填）
- achievement_type_name: 教学成果奖类型（湖南中医药大学教学成果奖/湖南中医药大学研究生教学成果奖/湖南省计算机学会高等教育教学成果奖/其它）
- achievement_level_name: 成果等级（国家级/省部级/市厅级/校级/院级/其它）
- award_rank_name: 获奖等级（特等奖/一等奖/二等奖/三等奖/优秀奖/其它）
- main_contributors: 主要完成人（多人用分号分隔）
- completing_units: 成果完成单位（多个用分号分隔）
- award_year: 获奖年度（仅数字）
- certificate_number: 证书编号
- awarding_unit: 颁奖单位
- award_date: 获奖日期（YYYY-MM-DD）

### 教学竞赛获奖
**特征词**：教师团队、教师竞赛、教学竞赛、特发此证、鼓励
- title: 竞赛名称
- award_year: 获奖年度（仅数字）
- competition_level_name: 竞赛等级（国家级/省部级/市厅级/校级/院级/其它）
- award_rank_name: 获奖等级（特等奖/一等奖/二等奖/三等奖/优秀奖/其它）
- winners: 获奖人（多人用分号分隔）
- winner_unit: 获奖人所在单位
- competition_name: 竞赛主办方
- certificate_number: 证书编号
- award_date: 获奖日期（YYYY-MM-DD）

#### 指导学生获奖
**特征词**：指导老师、学生获奖、指导教师、获奖学生、学生姓名
- title: 获奖名称
- award_year: 获奖年度（仅数字）
- competition_name: 竞赛名称
- competition_level_name: 竞赛等级（国家级/省部级/市厅级/校级/院级/其它）
- award_rank_name: 获奖等级（特等奖/一等奖/二等奖/三等奖/优秀奖/其它）
- student_name: 获奖学生
- project_name: 获奖项目名称
- teacher_name: 指导教师
- student_unit: 获奖学生所在单位
- organizer: 竞赛主办方
- certificate_number: 证书编号
- award_date: 获奖日期（YYYY-MM-DD）

【输出规则】
1. 仅输出标准 JSON 字符串，无任何多余文字、注释、反引号
2. 所有字段值为字符串类型，无信息则为空字符串
3. 必须包含 confidence 字段（0-1，代表识别置信度）
4. type_name 字段必须匹配指定的成果类型列表
5. 对于教研教改项目，重点关注以下特征词：
   - "教学改革研究"、"教改"、"课程建设"、"一流本科课程"、"课程思政"
   - "学校名称"、"项目名称"、"主持人"、"参加人员"、"项目类别"
   - "普通教育"、"湖南省普通高等学校"等
"""

    payload = {
        "model": "glm-4-flash",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "stream": False
    }


    try:
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        result = response.json()
        ai_content = result['choices'][0]['message']['content'].strip()

        # 清理 AI 返回格式
        ai_content = ai_content.replace('json', '').replace('```', '').strip()
        ai_data_raw = json.loads(ai_content)


        # 检查 AI 返回的数据类型
        if isinstance(ai_data_raw, list):
            ai_data = ai_data_raw[0] if len(ai_data_raw) > 0 and isinstance(ai_data_raw[0], dict) else {}
        elif isinstance(ai_data_raw, dict):
            ai_data = ai_data_raw
        else:
            ai_data = {}

        type_name = ai_data.get('type_name', '')
        if type_name == '教学竞赛获奖':
            student_name = ai_data.get('student_name', '')
            teacher_name = ai_data.get('teacher_name', '')
            has_student_keyword = '学生' in ocr_text or '指导' in ocr_text
            if (student_name or teacher_name) and has_student_keyword:
                type_name = '指导学生获奖'
                ai_data['type_name'] = type_name
        elif type_name == '指导学生获奖':
            student_name = ai_data.get('student_name', '')
            teacher_name = ai_data.get('teacher_name', '')
            has_student_keyword = '学生' in ocr_text or '指导' in ocr_text
            if not student_name and not teacher_name and not has_student_keyword:
                type_name = '教学竞赛获奖'
                ai_data['type_name'] = type_name

        # 数据清洗：确保所有字段存在且为字符串（强制转换，避免 None）
        result_data = {
            # 通用字段
            'type_name': str(ai_data.get('type_name', '其他') or '其他'),
            'title': str(ai_data.get('title', '') or ''),
            'confidence': float(ai_data.get('confidence', 0.8) or 0.8),
            'raw_data': ai_data,

            # 期刊论文（强制转换为字符串，避免 None）
            'authors': str(ai_data.get('authors', '') or ''),
            'corresponding_authors': str(ai_data.get('corresponding_authors', '') or ''),
            'journal_name': str(ai_data.get('journal_name', '') or ''),
            'inclusion_status': str(ai_data.get('inclusion_status', '') or ''),
            'year': str(ai_data.get('year', '') or ''),
            'volume': str(ai_data.get('volume', '') or ''),
            'issue': str(ai_data.get('issue', '') or ''),
            'page_range': str(ai_data.get('page_range', '') or ''),
            'doi': str(ai_data.get('doi', '') or ''),
            'publish_year': str(ai_data.get('publish_year', '') or ''),
            'publish_date': str(ai_data.get('publish_date', '') or ''),

            # 会议论文
            'conference_name': str(ai_data.get('conference_name', '') or ''),
            'conference_time': str(ai_data.get('conference_time', '') or ''),
            'conference_place': str(ai_data.get('conference_place', '') or ''),

            # 教材
            'textbook_series': str(ai_data.get('textbook_series', '') or ''),
            'chief_editor': str(ai_data.get('chief_editor', '') or ''),
            'associate_editors': str(ai_data.get('associate_editors', '') or ''),
            'editorial_board': str(ai_data.get('editorial_board', '') or ''),
            'publisher': str(ai_data.get('publisher', '') or ''),
            'isbn': str(ai_data.get('isbn', '') or ''),
            'cip_number': str(ai_data.get('cip_number', '') or ''),
            'publication_year': str(ai_data.get('publication_year', '') or ''),
            'publication_month': str(ai_data.get('publication_month', '') or ''),
            'edition': str(ai_data.get('edition', '') or ''),
            'word_count': str(ai_data.get('word_count', '') or ''),
            'price': str(ai_data.get('price', '') or ''),
            'textbook_level': str(ai_data.get('textbook_level', '') or ''),
            'textbook_type': str(ai_data.get('textbook_type', '') or ''),
            'applicable_majors': str(ai_data.get('applicable_majors', '') or ''),
            'remarks': str(ai_data.get('remarks', '') or ''),

            # 专著（增强字段）
            'textbook_series': str(ai_data.get('textbook_series', '') or ''),
            'chief_editor': str(ai_data.get('chief_editor', '') or ''),
            'associate_editors': str(ai_data.get('associate_editors', '') or ''),
            'editorial_board': str(ai_data.get('editorial_board', '') or ''),
            'publisher': str(ai_data.get('publisher', '') or ''),
            'isbn': str(ai_data.get('isbn', '') or ''),
            'cip_number': str(ai_data.get('cip_number', '') or ''),
            'publication_year': str(ai_data.get('publication_year', '') or ''),
            'publication_month': str(ai_data.get('publication_month', '') or ''),
            'publish_date': str(ai_data.get('publish_date', '') or ''),
            'edition': str(ai_data.get('edition', '') or ''),
            'word_count': str(ai_data.get('word_count', '') or ''),
            'price': str(ai_data.get('price', '') or ''),
            'monograph_type': str(ai_data.get('monograph_type', '') or ''),
            'applicable_majors': str(ai_data.get('applicable_majors', '') or ''),
            'remarks': str(ai_data.get('remarks', '') or ''),

            # 专利
            'patent_type': str(ai_data.get('patent_type', '') or ''),
            'patentee': str(ai_data.get('patentee', '') or ''),
            'address': str(ai_data.get('address', '') or ''),
            'inventors': str(ai_data.get('inventors', '') or ''),
            'status': str(ai_data.get('status', '') or ''),
            'patent_number': str(ai_data.get('patent_number', '') or ''),
            'grant_announcement_number': str(ai_data.get('grant_announcement_number', '') or ''),
            'apply_date': str(ai_data.get('apply_date', '') or ''),
            'grant_announcement_date': str(ai_data.get('grant_announcement_date', '') or ''),
            'applicant_at_apply_date': str(ai_data.get('applicant_at_apply_date', '') or ''),
            'inventor_at_apply_date': str(ai_data.get('inventor_at_apply_date', '') or ''),

            # 软著
            'copyright_owner': str(ai_data.get('copyright_owner', '') or ''),
            'completion_date': str(ai_data.get('completion_date', '') or ''),
            'first_publication_date': str(ai_data.get('first_publication_date', '') or ''),
            'right_acquisition_method': str(ai_data.get('right_acquisition_method', '') or ''),
            'right_scope': str(ai_data.get('right_scope', '') or ''),
            'copyright_number': str(ai_data.get('copyright_number', '') or ''),
            'certificate_number': str(ai_data.get('certificate_number', '') or ''),
            'register_date': str(ai_data.get('register_date', '') or ''),

            # 教研教改和课程建设项目（新增字段）
            'project_code': str(ai_data.get('project_code', '') or ''),
            'project_leader': str(ai_data.get('project_leader', '') or ''),
            'project_members': str(ai_data.get('project_members', '') or ''),
            'approval_department': str(ai_data.get('approval_department', '') or ''),
            'approval_date': str(ai_data.get('approval_date', '') or ''),
            'project_type_name': str(ai_data.get('project_type_name', '') or ''),
            'project_level_name': str(ai_data.get('project_level_name', '') or ''),
            'project_category_name': str(ai_data.get('project_category_name', '') or ''),
            'funding': str(ai_data.get('funding', '') or ''),
            'start_date': str(ai_data.get('start_date', '') or ''),
            'end_date': str(ai_data.get('end_date', '') or ''),

            # 教学成果获奖（新增字段）
            'achievement_type_name': str(ai_data.get('achievement_type_name', '') or ''),
            'achievement_level_name': str(ai_data.get('achievement_level_name', '') or ''),
            'award_rank_name': str(ai_data.get('award_rank_name', '') or ''),
            'main_contributors': str(ai_data.get('main_contributors', '') or ''),
            'completing_units': str(ai_data.get('completing_units', '') or ''),
            'award_year': str(ai_data.get('award_year', '') or ''),
            'certificate_number': str(ai_data.get('certificate_number', '') or ''),
            'awarding_unit': str(ai_data.get('awarding_unit', '') or ''),

            # 获奖类通用
            'award_level': str(ai_data.get('award_level', '') or ''),
            'award_rank': str(ai_data.get('award_rank', '') or ''),
            'award_date': str(ai_data.get('award_date', '') or ''),
            'competition_name': str(ai_data.get('competition_name', '') or ''),
            'student_name': str(ai_data.get('student_name', '') or ''),
        }

        return result_data
    except Exception as e:
        logger.error(f"AI 分析成果信息失败：{str(e)}")
        # 兜底返回空字段（所有字段强制为空字符串，绝对不能为 None）
        return {
            # 通用字段
            'type_name': '错误',
            'title': '',
            'confidence': 0.5,
            'raw_data': {},

            # 期刊论文（所有字段默认为空字符串）
            'authors': '',
            'corresponding_authors': '',
            'journal_name': '',
            'inclusion_status': '',
            'year': '',
            'volume': '',
            'issue': '',
            'page_range': '',
            'doi': '',
            'publish_year': '',
            'publish_date': '',

            # 会议论文
            'conference_name': '',
            'conference_time': '',
            'conference_place': '',

            # 教材
            'textbook_series': '',
            'chief_editor': '',
            'associate_editors': '',
            'editorial_board': '',
            'publisher': '',
            'isbn': '',
            'cip_number': '',
            'publication_year': '',
            'publication_month': '',
            'edition': '',
            'word_count': '',
            'price': '',
            'textbook_level': '',
            'textbook_type': '',
            'applicable_majors': '',
            'remarks': '',

            # 专著
            'textbook_series': '',
            'chief_editor': '',
            'associate_editors': '',
            'editorial_board': '',
            'publisher': '',
            'isbn': '',
            'cip_number': '',
            'publication_year': '',
            'publication_month': '',
            'publish_date': '',
            'edition': '',
            'word_count': '',
            'price': '',
            'monograph_type': '',
            'applicable_majors': '',
            'remarks': '',

            # 专利
            'patent_type': '',
            'patentee': '',
            'address': '',
            'inventors': '',
            'status': '',
            'patent_number': '',
            'grant_announcement_number': '',
            'apply_date': '',
            'grant_announcement_date': '',
            'applicant_at_apply_date': '',
            'inventor_at_apply_date': '',

            # 软著
            'copyright_owner': '',
            'completion_date': '',
            'first_publication_date': '',
            'right_acquisition_method': '',
            'right_scope': '',
            'copyright_number': '',
            'certificate_number': '',
            'register_date': '',

            # 教研教改和课程建设项目（新增字段）
            'project_code': '',
            'project_leader': '',
            'project_members': '',
            'approval_department': '',
            'approval_date': '',
            'project_type': '',
            'project_level': '',
            'project_category': '',
            'funding': '',
            'start_date': '',
            'end_date': '',

            # 获奖类
            'award_level': '',
            'award_rank': '',
            'award_date': '',
            'competition_name': '',
            'student_name': '',
        }


def pdf_to_images(pdf_path, output_dir=None):
    """
    将PDF文件转换为图片（每页一张），优化大PDF处理
    :param pdf_path: PDF文件路径
    :param output_dir: 图片输出目录（默认临时目录）
    :return: 图片文件路径列表
    """
    if not output_dir:
        output_dir = tempfile.mkdtemp()  # 创建临时目录

    try:
        # 适配Windows/Linux/Mac
        poppler_path = None
        if os.name == 'nt':  # Windows系统
            poppler_path = r"D:\Poppler\Library\bin"  # 替换为你的poppler路径

        # 优化：增加参数减少内存占用，分块处理
        images = convert_from_path(
            pdf_path,
            dpi=200,  # 降低分辨率（从300改为200，可根据需要调整）
            output_folder=output_dir,
            fmt='png',
            poppler_path=poppler_path,
            paths_only=True,  # 只返回文件路径，不加载图片对象
            grayscale=True,  # 转为灰度图，减少文件大小
            thread_count=2  # 多线程处理，提升速度
        )
        return images
    except Exception as e:
        logger.error(f"PDF转图片失败：{str(e)}")
        raise Exception(f"PDF转图片失败：{str(e)}")



def init_project_dictionaries():
    """初始化项目字典表数据（首次运行时调用）"""
    try:
        # 先创建所有数据库表
        db.create_all()

        # 初始化项目类型
        project_types = [
            ('普通本科高校教学改革研究项目', 1),
            ('学位与研究生教育改革研究项目', 2),
            ('一流本科课程建设项目', 3),
            ('课程思政建设项目', 4)
        ]
        for type_name, sort in project_types:
            if not ProjectType.query.filter_by(type_name=type_name).first():
                pt = ProjectType(type_name=type_name, sort_order=sort)
                db.session.add(pt)

        # 初始化项目状态
        project_statuses = [
            ('在研', 1),
            ('结题', 2),
            ('延期', 3)
        ]
        for status_name, sort in project_statuses:
            if not ProjectStatus.query.filter_by(status_name=status_name).first():
                ps = ProjectStatus(status_name=status_name, sort_order=sort)
                db.session.add(ps)

        # 初始化项目级别
        project_levels = [
            ('国家级', 1),
            ('省部级', 2),
            ('市厅级', 3),
            ('校级', 4),
            ('院级', 5)
        ]
        for level_name, sort in project_levels:
            if not ProjectLevel.query.filter_by(level_name=level_name).first():
                pl = ProjectLevel(level_name=level_name, sort_order=sort)
                db.session.add(pl)

        # 初始化项目类别
        project_categories = [
            ('重点项目', 1),
            ('一般项目', 2),
            ('线上一流课程', 3),
            ('线上线下混合式一流课程', 4),
            ('线下一流课程', 5),
            ('社会实践一流课程', 6),
            ('虚拟仿真实验教学一流课程', 7)
        ]
        for category_name, sort in project_categories:
            if not ProjectCategory.query.filter_by(category_name=category_name).first():
                pc = ProjectCategory(category_name=category_name, sort_order=sort)
                db.session.add(pc)

        # 初始化专利类型
        patent_types = [
            ('发明专利', 1),
            ('实用新型专利', 2),
            ('外观设计专利', 3)
        ]
        for type_name, sort in patent_types:
            if not PatentType.query.filter_by(type_name=type_name).first():
                pt = PatentType(type_name=type_name, sort_order=sort)
                db.session.add(pt)

        # 初始化专利状态
        patent_statuses = [
            ('受理', 1),
            ('初步审查', 2),
            ('公开', 3),
            ('实质审查', 4),
            ('授权', 5)
        ]
        for status_name, sort in patent_statuses:
            if not PatentStatus.query.filter_by(status_name=status_name).first():
                ps = PatentStatus(status_name=status_name, sort_order=sort)
                db.session.add(ps)

        db.session.commit()
        logger.info("项目字典表初始化完成")
    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ 项目字典表初始化失败：{str(e)}")



# 在应用启动时自动初始化字典表
with app.app_context():
    init_project_dictionaries()

# ---------------------- 4. 核心路由 ----------------------
@app.route('/')
def index():
    """首页"""
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))

    # 首页内容（根据角色显示不同内容）
    if user.role == 'teacher':
        content = '''
        <h2>教师工作台</h2>
        <p>欢迎使用教学成果管理系统！</p>
        <p>您可以通过左侧导航栏管理您的论文、专利、获奖等教学成果。</p>
        <ul>
            <li>📄 录入/编辑个人论文、教材等成果</li>
            <li>📊 查看个人成果统计分析</li>
            <li>📤 导出成果数据用于项目申报</li>
        </ul>
        '''
    elif user.role == 'team_leader':
        content = '''
        <h2>团队负责人工作台</h2>
        <p>您可以管理团队成员并查看团队整体成果数据。</p>
        <ul>
            <li>👨‍🏫 管理团队成员</li>
            <li>📊 查看团队成果统计</li>
            <li>📤 导出团队成果数据</li>
            <li>📄 管理个人教学成果</li>
        </ul>
        '''
    else:  # admin
        content = '''
        <h2>系统管理员工作台</h2>
        <p>您可以管理系统用户、团队和全局配置。</p>
        <ul>
            <li>👥 管理所有用户账号</li>
            <li>🏢 创建/删除团队</li>
            <li>📊 查看系统整体数据统计</li>
        </ul>
        '''

    return render_base_layout('首页', content, user)


@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录页面"""
    # 如果已登录，跳转到首页
    if get_current_user():
        return redirect(url_for('index'))

    # 处理登录提交
    if request.method == 'POST':
        login_id = request.form.get('login_id')  # 用户名/工号
        password = request.form.get('password')

        # 验证用户（支持用户名或工号登录）
        user = User.query.filter(
            (User.username == login_id) | (User.employee_id == login_id)
        ).first()

        if user and user.check_password(password):
            # 登录成功，设置session
            session['user_id'] = user.id
            flash('登录成功！', 'success')
            return redirect(url_for('index'))
        else:
            flash('用户名/工号或密码错误！', 'danger')

    # 登录页面HTML（无Jinja）
    flash_messages = ''
    for category, message in session.pop('_flashes', []):
        flash_messages += f'<div class="alert alert-{category}">{message}</div>'

    login_html = f'''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>登录 - 教学成果管理系统</title>
    <style>
        body {{
            font-family: "Microsoft YaHei", sans-serif;
            background: #f5f7fa;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }}
        .login-box {{
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 20px rgba(0,0,0,0.1);
            width: 400px;
        }}
        .login-box h2 {{
            text-align: center;
            margin-bottom: 30px;
            color: #2c3e50;
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        label {{
            display: block;
            margin-bottom: 8px;
            font-weight: 600;
            color: #495057;
        }}
        input {{
            width: 100%;
            padding: 10px 15px;
            border: 1px solid #ced4da;
            border-radius: 4px;
            font-size: 14px;
        }}
        input:focus {{
            outline: none;
            border-color: #3498db;
            box-shadow: 0 0 0 3px rgba(52, 152, 219, 0.1);
        }}
        .btn {{
            width: 100%;
            padding: 10px;
            background: #3498db;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 16px;
            margin-top: 10px;
        }}
        .btn:hover {{
            background: #2980b9;
        }}
        .alert {{
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 4px;
        }}
        .alert-success {{
            background: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
        }}
        .alert-danger {{
            background: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
        }}
        .register-link {{
            text-align: center;
            margin-top: 20px;
        }}
        .register-link a {{
            color: #3498db;
            text-decoration: none;
        }}
    </style>
</head>
<body>
    <div class="login-box">
        <h2>教学成果管理系统</h2>
        {flash_messages}
        <form method="POST">
            <div class="form-group">
                <label for="login_id">用户名/工号</label>
                <input type="text" id="login_id" name="login_id" required>
            </div>
            <div class="form-group">
                <label for="password">密码</label>
                <input type="password" id="password" name="password" required>
            </div>
            <button type="submit" class="btn">登录</button>
        </form>
        <div class="register-link">
            <a href="/register">还没有账号？点击注册</a>
        </div>
    </div>
</body>
</html>
'''
    return login_html


@app.route('/register', methods=['GET', 'POST'])
def register():
    """注册页面（仅显示必填项）"""
    # 如果已登录，跳转到首页
    if get_current_user():
        return redirect(url_for('index'))

    # 处理注册提交
    if request.method == 'POST':
        try:
            # 获取表单数据（仅保留必填项）
            username = request.form.get('username')
            password = request.form.get('password')
            employee_id = request.form.get('employee_id')
            email = request.form.get('email')
            user_role = request.form.get('role', 'teacher')

            # 检查必填字段唯一性
            if User.query.filter_by(username=username).first():
                flash('用户名已存在！', 'danger')
                return redirect(url_for('register'))

            if User.query.filter_by(employee_id=employee_id).first():
                flash('工号已存在！', 'danger')
                return redirect(url_for('register'))

            if User.query.filter_by(email=email).first():
                flash('邮箱已存在！', 'danger')
                return redirect(url_for('register'))

            # 安全校验：仅允许teacher/team_leader角色
            if user_role not in ['teacher', 'team_leader']:
                user_role = 'teacher'

            # 创建用户（仅初始化必填字段）
            user = User(
                username=username,
                employee_id=employee_id,
                email=email,
                role=user_role
            )
            user.set_password(password)

            # 保存到数据库
            db.session.add(user)
            db.session.commit()

            flash('注册成功！请登录', 'success')
            return redirect(url_for('login'))

        except Exception as e:
            db.session.rollback()
            flash(f'注册失败：{str(e)}', 'danger')

    # 注册页面HTML（仅保留必填项）
    flash_messages = ''
    for category, message in session.pop('_flashes', []):
        flash_messages += f'<div class="alert alert-{category}">{message}</div>'

    register_html = f'''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>注册 - 教学成果管理系统</title>
    <style>
        body {{
            font-family: "Microsoft YaHei", sans-serif;
            background: #f5f7fa;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }}
        .register-box {{
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 20px rgba(0,0,0,0.1);
            width: 400px; /* 缩小宽度，适配少字段 */
        }}
        .register-box h2 {{
            text-align: center;
            margin-bottom: 30px;
            color: #2c3e50;
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        label {{
            display: block;
            margin-bottom: 8px;
            font-weight: 600;
            color: #495057;
        }}
        input, select {{
            width: 100%;
            padding: 10px 15px;
            border: 1px solid #ced4da;
            border-radius: 4px;
            font-size: 14px;
        }}
        input:focus, select:focus {{
            outline: none;
            border-color: #3498db;
            box-shadow: 0 0 0 3px rgba(52, 152, 219, 0.1);
        }}
        .btn {{
            width: 100%;
            padding: 10px;
            background: #3498db;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 16px;
            margin-top: 20px;
        }}
        .btn:hover {{
            background: #2980b9;
        }}
        .alert {{
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 4px;
        }}
        .alert-success {{
            background: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
        }}
        .alert-danger {{
            background: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
        }}
        .login-link {{
            text-align: center;
            margin-top: 20px;
        }}
        .login-link a {{
            color: #3498db;
            text-decoration: none;
        }}
        .required {{
            color: red;
        }}
    </style>
</head>
<body>
    <div class="register-box">
        <h2>用户注册</h2>
        {flash_messages}
        <form method="POST">
            <!-- 仅保留必填字段 -->
            <div class="form-group">
                <label for="username">用户名 <span class="required">*</span></label>
                <input type="text" id="username" name="username" required>
            </div>

            <div class="form-group">
                <label for="employee_id">工号 <span class="required">*</span></label>
                <input type="text" id="employee_id" name="employee_id" required>
            </div>

            <div class="form-group">
                <label for="email">邮箱 <span class="required">*</span></label>
                <input type="email" id="email" name="email" required>
            </div>

            <div class="form-group">
                <label for="role">用户角色 <span class="required">*</span></label>
                <select id="role" name="role" required>
                    <option value="">请选择</option>
                    <option value="teacher">普通教师</option>
                    <option value="team_leader">团队负责人</option>
                </select>
            </div>

            <div class="form-group">
                <label for="password">密码 <span class="required">*</span></label>
                <input type="password" id="password" name="password" required minlength="6">
            </div>

            <button type="submit" class="btn">注册</button>
        </form>
        <div class="login-link">
            <a href="/login">已有账号？返回登录</a>
        </div>
    </div>
</body>
</html>
'''
    return register_html


@app.route('/logout')
def logout():
    """登出（清空所有session数据，包括flash消息）"""
    # 清空整个session，而非仅删除user_id
    session.clear()
    flash('已成功退出登录！', 'success')
    return redirect(url_for('login'))


# ---------------------- 通用成果管理路由 ----------------------
@app.route('/user/settings', methods=['GET', 'POST'])
def user_settings():
    """个人信息修改（邮箱、电话、身份证等）"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 核心新增：管理员禁止访问个人信息设置
    if current_user.role == 'admin':
        flash('管理员无需设置个人账户信息！', 'danger')
        return redirect(url_for('index'))

    # 处理表单提交
    if request.method == 'POST':
        try:
            # 基础信息更新
            current_user.gender = request.form.get('gender') or None
            birth_date_str = request.form.get('birth_date')
            if birth_date_str:
                current_user.birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
            else:
                current_user.birth_date = None

            # 关键修复：空身份证号转为 None
            current_user.id_card = request.form.get('id_card').strip() if request.form.get('id_card') else None
            current_user.email = request.form.get('email', '')  # 必填，前端已校验
            current_user.phone = request.form.get('phone') or None
            current_user.office_phone = request.form.get('office_phone') or None
            current_user.school = request.form.get('school') or None
            current_user.college = request.form.get('college') or None
            current_user.department = request.form.get('department') or None
            current_user.research_room = request.form.get('research_room') or None

            db.session.commit()
            flash('个人信息修改成功！', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'修改失败：{str(e)}', 'danger')

    # 渲染个人信息修改页面
    birth_date = current_user.birth_date.strftime('%Y-%m-%d') if current_user.birth_date else ''
    form_html = f'''
    <h2>个人信息修改</h2>
    <form method="POST">
        <div class="form-group">
            <label>用户名（不可修改）</label>
            <input type="text" value="{current_user.username}" disabled>
        </div>
        <div class="form-group">
            <label>工号（不可修改）</label>
            <input type="text" value="{current_user.employee_id}" disabled>
        </div>
        <div class="form-group">
            <label>性别</label>
            <select name="gender">
                <option value="">请选择</option>
                <option value="男" {"selected" if current_user.gender == '男' else ''}>男</option>
                <option value="女" {"selected" if current_user.gender == '女' else ''}>女</option>
            </select>
        </div>
        <div class="form-group">
            <label>出生年月日</label>
            <input type="date" name="birth_date" value="{birth_date}">
        </div>
        <div class="form-group">
            <label>身份证号码</label>
            <input type="text" name="id_card" value="{current_user.id_card or ''}" maxlength="18">
        </div>
        <div class="form-group">
            <label>邮箱 <span class="required">*</span></label>
            <input type="email" name="email" value="{current_user.email}" required>
        </div>
        <div class="form-group">
            <label>手机号</label>
            <input type="tel" name="phone" value="{current_user.phone or ''}">
        </div>
        <div class="form-group">
            <label>办公电话</label>
            <input type="tel" name="office_phone" value="{current_user.office_phone or ''}">
        </div>
        <div class="form-group">
            <label>学校</label>
            <input type="text" name="school" value="{current_user.school or ''}">
        </div>
        <div class="form-group">
            <label>学院</label>
            <input type="text" name="college" value="{current_user.college or ''}">
        </div>
        <div class="form-group">
            <label>系部</label>
            <input type="text" name="department" value="{current_user.department or ''}">
        </div>
        <div class="form-group">
            <label>教研室</label>
            <input type="text" name="research_room" value="{current_user.research_room or ''}">
        </div>
        <button type="submit" class="btn">保存修改</button>
    </form>
    <div style="margin-top:20px;">
        <a href="/user/change_password" class="btn">修改密码</a>
        <a href="/user/api_config" class="btn">大模型API配置</a>
    </div>
    '''
    return render_base_layout('个人信息设置', form_html, current_user)


@app.route('/user/change_password', methods=['GET', 'POST'])
def change_password():
    """密码修改（个人主动改密码）"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 核心新增：管理员禁止修改密码（需通过数据库重置）
    if current_user.role == 'admin':
        flash('管理员密码请通过数据库手动重置！', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        old_pwd = request.form.get('old_password')
        new_pwd = request.form.get('new_password')

        # 校验
        if not current_user.check_password(old_pwd):
            flash('原密码错误！', 'danger')
        elif len(new_pwd) < 6:
            flash('新密码长度不能少于6位！', 'danger')
        else:
            try:
                current_user.set_password(new_pwd)
                db.session.commit()
                flash('密码修改成功，请重新登录！', 'success')
                return redirect(url_for('logout'))
            except Exception as e:
                db.session.rollback()
                flash(f'修改失败：{str(e)}', 'danger')

    # 渲染密码修改页面
    form_html = '''
    <h2>修改密码</h2>
    <form method="POST">
        <div class="form-group">
            <label>原密码 <span class="required">*</span></label>
            <input type="password" name="old_password" required>
        </div>
        <div class="form-group">
            <label>新密码 <span class="required">*</span></label>
            <input type="password" name="new_password" required minlength="6">
        </div>
        <button type="submit" class="btn">确认修改</button>
    </form>
    '''
    return render_base_layout('修改密码', form_html, current_user)


@app.route('/user/api_config', methods=['GET', 'POST'])
def api_config():
    """大模型API配置（仅保留百度+智谱）"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 核心新增：管理员禁止配置API
    if current_user.role == 'admin':
        flash('管理员无需配置个人账户的API！', 'danger')
        return redirect(url_for('index'))

    # 获取现有API配置
    api_config = current_user.get_api_config() or {}

    if request.method == 'POST':
        try:
            # 仅保留百度+智谱API配置
            new_config = {
                'baidu': {
                    'api_key': request.form.get('baidu_api_key', ''),
                    'secret_key': request.form.get('baidu_secret_key', '')
                },
                'zhipu': {
                    'api_key': request.form.get('zhipu_api_key', '')
                }
            }
            current_user.set_api_config(new_config)
            db.session.commit()
            flash('API配置保存成功！', 'success')
            api_config = new_config  # 更新页面展示数据
        except Exception as e:
            db.session.rollback()
            flash(f'保存失败：{str(e)}', 'danger')

    # 渲染简化后的API配置页面（仅百度+智谱）
    form_html = f'''
    <h2>大模型API配置</h2>
    <div class="alert alert-info">
        配置完成后可用于OCR智能导入、语音导入/导出等功能
    </div>
    <form method="POST">
        <h3 style="margin-top:20px;">百度文心一言API配置</h3>
        <div class="form-group">
            <label>API Key</label>
            <input type="text" name="baidu_api_key" value="{api_config.get('baidu', {}).get('api_key', '')}" >
        </div>
        <div class="form-group">
            <label>Secret Key</label>
            <input type="text" name="baidu_secret_key" value="{api_config.get('baidu', {}).get('secret_key', '')}" >
        </div>

        <h3 style="margin-top:30px;">智谱AI（ZHIPU）API配置</h3>
        <div class="form-group">
            <label>API Key</label>
            <input type="text" name="zhipu_api_key" value="{api_config.get('zhipu', {}).get('api_key', '')}" >
        </div>

        <button type="submit" class="btn" style="margin-top:30px;">保存配置</button>
    </form>
    '''
    return render_base_layout('大模型API配置', form_html, current_user)

@app.route('/admin/user_manage', methods=['GET', 'POST'])
def admin_user_manage():
    """管理员-用户管理（仅查看角色，不可修改）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'admin':
        flash('无管理员权限！', 'danger')
        return redirect(url_for('index'))

    # 处理用户创建/删除（保留创建、删除功能，移除角色编辑）
    if request.method == 'POST':
        action = request.form.get('action')
        user_id = request.form.get('user_id')

        try:
            if action == 'create':
                # 创建新用户
                username = request.form.get('username')
                employee_id = request.form.get('employee_id')
                email = request.form.get('email')
                role = request.form.get('role', 'teacher')
                password = request.form.get('password', '123456')  # 默认密码

                # 校验唯一性
                if User.query.filter_by(username=username).first():
                    flash('用户名已存在！', 'danger')
                    return redirect(url_for('admin_user_manage'))
                if User.query.filter_by(employee_id=employee_id).first():
                    flash('工号已存在！', 'danger')
                    return redirect(url_for('admin_user_manage'))

                # 安全校验：仅允许teacher/team_leader角色
                if user_role not in ['teacher', 'team_leader']:
                    user_role = 'teacher'

                # 创建用户
                new_user = User(
                    username=username,
                    employee_id=employee_id,
                    email=email,
                    role=role
                )
                new_user.set_password(password)
                db.session.add(new_user)
                flash(f'用户{username}创建成功（默认密码：{password}）', 'success')

            elif action == 'delete':
                # 删除用户（非管理员）
                user = db.session.get(User, user_id)
                if user and user.username != 'admin':
                    # 删除关联数据（简化版，实际可保留成果数据）
                    db.session.delete(user)
                    flash(f'用户{user.username}删除成功！', 'success')

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

    # 查询所有用户
    users = User.query.order_by(User.role, User.username).all()

    # 渲染用户管理页面（移除角色编辑下拉框，仅展示角色）
    user_list_html = '''
    <h2>用户管理</h2>
    <div style="margin-bottom:20px;">
        <button onclick="showCreateForm()" class="btn">新增用户</button>
    </div>

    <!-- 新增用户表单 -->
    <div id="createForm" style="display:none; margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h3>新增用户</h3>
        <form method="POST">
            <input type="hidden" name="action" value="create">
            <div class="form-group">
                <label>用户名 <span class="required">*</span></label>
                <input type="text" name="username" required>
            </div>
            <div class="form-group">
                <label>工号 <span class="required">*</span></label>
                <input type="text" name="employee_id" required>
            </div>
            <div class="form-group">
                <label>邮箱 <span class="required">*</span></label>
                <input type="email" name="email" required>
            </div>
            <div class="form-group">
                <label>角色 <span class="required">*</span></label>
                <select name="role">
                    <option value="teacher">普通教师</option>
                    <option value="team_leader">团队负责人</option>
                    <option value="admin">管理员（谨慎）</option>
                </select>
            </div>
            <div class="form-group">
                <label>初始密码（默认：123456）</label>
                <input type="password" name="password" value="123456">
            </div>
            <button type="submit" class="btn">创建</button>
            <button type="button" onclick="hideCreateForm()" class="btn" style="background:#95a5a6;">取消</button>
        </form>
    </div>

    <!-- 用户列表 -->
    <table style="width:100%; border-collapse:collapse;">
        <thead>
            <tr style="background:#f5f7fa;">
                <th style="padding:10px; border:1px solid #dee2e6;">用户名</th>
                <th style="padding:10px; border:1px solid #dee2e6;">工号</th>
                <th style="padding:10px; border:1px solid #dee2e6;">邮箱</th>
                <th style="padding:10px; border:1px solid #dee2e6;">角色</th>
                <th style="padding:10px; border:1px solid #dee2e6;">操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    for user in users:
        # 仅展示角色，移除编辑下拉框
        role_display = {
            'teacher': '普通教师',
            'team_leader': '团队负责人',
            'admin': '管理员'
        }.get(user.role, '未知角色')

        # 角色样式（区分不同角色）
        role_style = ''
        if user.role == 'admin':
            role_style = 'style="color: #e74c3c; font-weight: bold;"'
        elif user.role == 'team_leader':
            role_style = 'style="color: #2980b9; font-weight: bold;"'

        # 删除按钮（超级管理员不可删）
        delete_btn = ''
        if user.username != 'admin':
            delete_btn = f'''
            <form method="POST" style="display:inline;" onsubmit="return confirm('确定删除？')">
                <input type="hidden" name="action" value="delete">
                <input type="hidden" name="user_id" value="{user.id}">
                <button type="submit" class="btn" style="padding:5px 10px; font-size:12px; background:#e74c3c;">删除</button>
            </form>
            '''
        else:
            delete_btn = '<span style="color:#999;">不可删除</span>'

        user_list_html += f'''
        <tr>
            <td style="padding:10px; border:1px solid #dee2e6;">{user.username}</td>
            <td style="padding:10px; border:1px solid #dee2e6;">{user.employee_id}</td>
            <td style="padding:10px; border:1px solid #dee2e6;">{user.email}</td>
            <td style="padding:10px; border:1px solid #dee2e6;" {role_style}>{role_display}</td>
            <td style="padding:10px; border:1px solid #dee2e6;">{delete_btn}</td>
        </tr>
        '''

    user_list_html += '''
        </tbody>
    </table>

    <script>
        function showCreateForm() {
            document.getElementById('createForm').style.display = 'block';
        }
        function hideCreateForm() {
            document.getElementById('createForm').style.display = 'none';
        }
    </script>
    '''
    return render_base_layout('用户管理', user_list_html, current_user)


@app.route('/admin/team_manage', methods=['GET', 'POST'])
def admin_team_manage():
    """管理员-团队管理（创建/删除团队、指定负责人、添加成员）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'admin':
        flash('无管理员权限！', 'danger')
        return redirect(url_for('index'))

    # 处理团队操作
    if request.method == 'POST':
        action = request.form.get('action')
        try:
            if action == 'create_team':
                # 创建团队
                team_name = request.form.get('team_name')
                leader_id = request.form.get('leader_id')

                if Team.query.filter_by(name=team_name).first():
                    flash('团队名称已存在！', 'danger')
                else:
                    leader_user = db.session.get(User, leader_id)
                    new_team = Team(
                        name=team_name,
                        leader_id=leader_id
                    )
                    db.session.add(new_team)
                    db.session.flush()
                    # 自动将负责人加入团队
                    db.session.add(UserTeam(user_id=leader_id, team_id=new_team.id))
                    flash(f'团队{team_name}创建成功！已将{leader_user.username}设为团队负责人', 'success')

            elif action == 'delete_team':
                # 删除团队
                team_id = request.form.get('team_id')
                team = db.session.get(Team, team_id)
                if team:
                    # 删除团队成员关联
                    UserTeam.query.filter_by(team_id=team_id).delete()
                    # 删除团队
                    db.session.delete(team)
                    flash(f'团队{team.name}删除成功！', 'success')

            elif action == 'add_member':
                # 添加团队成员
                team_id = request.form.get('team_id')
                user_id = request.form.get('user_id')

                if UserTeam.query.filter_by(team_id=team_id, user_id=user_id).first():
                    flash('该用户已在团队中！', 'danger')
                else:
                    db.session.add(UserTeam(team_id=team_id, user_id=user_id))
                    flash('成员添加成功！', 'success')

            elif action == 'remove_member':
                # 移除团队成员
                ut_id = request.form.get('ut_id')
                ut = db.session.get(UserTeam, ut_id)
                if ut:
                    db.session.delete(ut)
                    flash('成员移除成功！', 'success')

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

    # 查询所有团队和用户
    teams = Team.query.all()
    all_users = User.query.filter(User.role != 'admin').all()  # 管理员不加入团队

    # 渲染团队管理页面
    team_html = '''
    <h2>团队管理</h2>

    <!-- 创建团队表单 -->
    <div style="margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h3>创建新团队</h3>
        <form method="POST">
            <input type="hidden" name="action" value="create_team">
            <div class="form-group">
                <label>团队名称 <span class="required">*</span></label>
                <input type="text" name="team_name" required>
            </div>
            <div class="form-group">
                <label>团队负责人 <span class="required">*</span></label>
                <select name="leader_id" required>
                    <option value="">请选择</option>
    '''
    # 填充负责人选项 - 仅允许选择注册为团队负责人的用户
    for user in all_users:
        if user.role == 'team_leader':  # 仅显示注册时就是团队负责人的用户
            team_html += f'<option value="{user.id}">{user.username}（{user.employee_id}）</option>'

    team_html += '''
                </select>
            </div>
            <button type="submit" class="btn">创建团队</button>
        </form>
    </div>

    <!-- 团队列表 -->
    '''
    for team in teams:
        leader = db.session.get(User, team.leader_id)
        # 查询团队成员
        members = UserTeam.query.filter_by(team_id=team.id).all()
        member_list = []
        for ut in members:
            user = db.session.get(User, ut.user_id)
            member_list.append((ut.id, user))

        # 团队卡片
        team_html += f'''
        <div style="margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:20px;">
                <h3>{team.name}</h3>
                <div>
                    <form method="POST" onsubmit="return confirm('确定删除该团队？')" style="display:inline;">
                        <input type="hidden" name="action" value="delete_team">
                        <input type="hidden" name="team_id" value="{team.id}">
                        <button type="submit" class="btn" style="background:#e74c3c;">删除团队</button>
                    </form>
                </div>
            </div>
            <div style="margin-bottom:10px;">
                <strong>团队负责人：</strong>{leader.username}（{leader.employee_id}）
            </div>

            <!-- 添加成员 -->
            <div style="margin-bottom:20px;">
                <form method="POST" style="display:flex; gap:10px; align-items:end;">
                    <input type="hidden" name="action" value="add_member">
                    <input type="hidden" name="team_id" value="{team.id}">
                    <div class="form-group" style="flex:1;">
                        <label>添加团队成员</label>
                        <select name="user_id" required>
                            <option value="">请选择用户</option>
        '''
        # 填充可选用户（排除已加入的）
        for user in all_users:
            is_in_team = any(ut.user_id == user.id for ut in members)
            if not is_in_team:
                team_html += f'<option value="{user.id}">{user.username}（{user.employee_id}）</option>'

        team_html += '''
                        </select>
                    </div>
                    <button type="submit" class="btn">添加</button>
                </form>
            </div>

            <!-- 成员列表 -->
            <div>
                <strong>团队成员：</strong>
                <ul style="margin:10px 0; padding-left:20px;">
        '''
        for ut_id, user in member_list:
            # 移除成员按钮（负责人不可移除）
            remove_btn = ''
            if user.id != team.leader_id:
                remove_btn = f'''
                <form method="POST" style="display:inline; margin-left:10px;">
                    <input type="hidden" name="action" value="remove_member">
                    <input type="hidden" name="ut_id" value="{ut_id}">
                    <button type="submit" class="btn" style="padding:2px 8px; font-size:12px; background:#95a5a6;" onclick="return confirm('确定移除？')">移除</button>
                </form>
                '''
            team_html += f'<li>{user.username}（{user.employee_id}）{remove_btn}</li>'

        team_html += '''
                </ul>
            </div>
        </div>
        '''

    return render_base_layout('团队管理', team_html, current_user)


@app.route('/team/achievements')
def team_achievements():
    """团队负责人-团队成果多维度统计（仅统计公开给本团队的成果）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('无团队负责人权限！', 'danger')
        return redirect(url_for('index'))

    # 获取团队信息和成员ID
    teams = Team.query.filter_by(leader_id=current_user.id).all()
    team_ids = [t.id for t in teams]
    team_user_ids = [ut.user_id for ut in UserTeam.query.filter(UserTeam.team_id.in_(team_ids)).all()]
    if not team_user_ids:
        content = '<div class="alert alert-warning">暂无团队成员数据！</div>'
        return render_base_layout('团队成果统计', content, current_user)

    # 关键：获取当前管理的团队ID字符串列表（用于过滤public_team_ids）
    managed_team_ids_str = [str(t.id) for t in teams]

    # 多维度统计
    stats = {
        'total': {},  # 总数统计
        'by_type': {},  # 按成果类型统计
        'by_year': {},  # 按年份统计
        'by_member': {}  # 按成员统计
    }

    # 成果模型列表
    achievement_models = [
        ('期刊论文', JournalPaper),
        ('会议论文', ConferencePaper),
        ('教材', Textbook),
        ('专著', Monograph),
        ('教研项目', TeachingProject),
        ('专利', Patent),
        ('软著', SoftwareCopyright),
        ('教学成果获奖', TeachingAchievementAward),
        ('教学竞赛获奖', TeachingCompetitionAward),
        ('指导学生获奖', StudentGuidanceAward)
    ]

    # 统计总数和按类型（仅统计公开给本团队的成果）
    total_count = 0
    for name, model in achievement_models:
        # 构建过滤条件：1. 属于团队成员 2. 公开给当前管理的任意团队
        or_conditions = []
        for team_id in managed_team_ids_str:
            # 处理格式：",1,2,3," 避免部分匹配（如1匹配10）
            or_conditions.append(
                func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
            )

        # 执行过滤查询
        query = model.query.filter(
            model.user_id.in_(team_user_ids),
            or_(*or_conditions)
        )
        count = query.count()

        stats['total'][name] = count
        total_count += count
    stats['total']['总计'] = total_count

    # 按年份统计（仅统计公开给本团队的成果）
    year_fields = {
        JournalPaper: 'publish_year',
        ConferencePaper: 'publish_year',
        Textbook: 'publish_date',
        Monograph: 'publish_date',
        TeachingProject: 'start_date',
        Patent: 'apply_date',
        SoftwareCopyright: 'register_date',
        TeachingAchievementAward: 'award_date',
        TeachingCompetitionAward: 'award_date',
        StudentGuidanceAward: 'award_date'
    }
    for name, model in achievement_models:
        # 过滤公开给本团队的成果
        or_conditions = []
        for team_id in managed_team_ids_str:
            or_conditions.append(
                func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
            )

        items = model.query.filter(
            model.user_id.in_(team_user_ids),
            or_(*or_conditions)
        ).all()

        field = year_fields[model]
        for item in items:
            value = getattr(item, field)
            if value:
                year = value.year if isinstance(value, date) else value
                if year not in stats['by_year']:
                    stats['by_year'][year] = {n: 0 for n, _ in achievement_models}
                stats['by_year'][year][name] += 1

    # 按成员统计（仅统计公开给本团队的成果）
    team_users = User.query.filter(User.id.in_(team_user_ids)).all()
    for user in team_users:
        user_count = {}
        for name, model in achievement_models:
            # 过滤该用户公开给本团队的成果
            or_conditions = []
            for team_id in managed_team_ids_str:
                or_conditions.append(
                    func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
                )

            user_count[name] = model.query.filter(
                model.user_id == user.id,
                or_(*or_conditions)
            ).count()

        user_count['总计'] = sum(user_count.values())
        stats['by_member'][user.username] = user_count

    # 渲染统计页面（添加导出按钮）
    stats_html = f'''
    <h2>团队成果统计（负责人：{current_user.username}）</h2>

    <!-- 总数统计 -->
    <div style="margin-bottom:30px;">
        <h3>成果总数</h3>
        <div style="padding:20px; background:#f5f7fa; border-radius:8px;">
            <p>团队总成果数：<strong>{stats['total']['总计']}</strong> 项</p>
            <table style="width:100%; border-collapse:collapse; margin-top:10px;">
                <thead>
                    <tr style="background:#e9ecef;">
                        <th style="padding:10px; border:1px solid #dee2e6;">成果类型</th>
                        <th style="padding:10px; border:1px solid #dee2e6;">数量（项）</th>
                    </tr>
                </thead>
                <tbody>
    '''
    # 核心修改：为每个成果类型添加导出按钮
    for name, count in stats['total'].items():
        if name != '总计':
            # 仅当有成果时显示导出按钮
            export_btn = ''
            if count > 0 and teams:  # 确保有团队ID
                export_btn = f'''
                <a href="/team/export_achievement?team_id={teams[0].id}&type={name}" 
                   class="btn" 
                   style="padding:5px 10px; font-size:12px; background:#27ae60; margin-left:10px;">
                    导出公开成果
                </a>
                '''
            stats_html += f'''
            <tr>
                <td style="padding:10px; border:1px solid #dee2e6;">{name}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">
                    {count}
                    {export_btn}
                </td>
            </tr>
            '''
    stats_html += '''
                </tbody>
            </table>
        </div>
    </div>

    <!-- 按成员统计 -->
    <div style="margin-bottom:30px;">
        <h3>按成员统计（仅统计公开给本团队的成果）</h3>
        <table style="width:100%; border-collapse:collapse;">
            <thead>
                <tr style="background:#e9ecef;">
                    <th style="padding:10px; border:1px solid #dee2e6;">团队成员</th>
    '''
    # 成员统计表头
    for name, _ in achievement_models:
        stats_html += f'<th style="padding:10px; border:1px solid #dee2e6;">{name}</th>'
    stats_html += '<th style="padding:10px; border:1px solid #dee2e6;">总计</th>'
    stats_html += '''
                </tr>
            </thead>
            <tbody>
    '''
    # 成员统计数据
    for username, counts in stats['by_member'].items():
        stats_html += f'<tr><td style="padding:10px; border:1px solid #dee2e6;">{username}</td>'
        for name, _ in achievement_models:
            stats_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{counts[name]}</td>'
        stats_html += f'<td style="padding:10px; border:1px solid #dee2e6;"><strong>{counts["总计"]}</strong></td></tr>'
    stats_html += '''
            </tbody>
        </table>
    </div>

    <!-- 按年份统计 -->
    <div>
        <h3>按年份统计（仅统计公开给本团队的成果）</h3>
        <table style="width:100%; border-collapse:collapse;">
            <thead>
                <tr style="background:#e9ecef;">
                    <th style="padding:10px; border:1px solid #dee2e6;">年份</th>
    '''
    # 年份统计表头
    for name, _ in achievement_models:
        stats_html += f'<th style="padding:10px; border:1px solid #dee2e6;">{name}</th>'
    stats_html += '<th style="padding:10px; border:1px solid #dee2e6;">总计</th>'
    stats_html += '''
                </tr>
            </thead>
            <tbody>
    '''
    # 年份统计数据
    for year in sorted(stats['by_year'].keys(), reverse=True):
        year_data = stats['by_year'][year]
        year_total = sum(year_data.values())
        stats_html += f'<tr><td style="padding:10px; border:1px solid #dee2e6;">{year}</td>'
        for name, _ in achievement_models:
            stats_html += f'<td style="padding:10px; border:1px solid #dee2e6;">{year_data[name]}</td>'
        stats_html += f'<td style="padding:10px; border:1px solid #dee2e6;"><strong>{year_total}</strong></td></tr>'
    stats_html += '''
            </tbody>
        </table>
    </div>
    '''
    return render_base_layout('团队成果统计', stats_html, current_user)


@app.route('/team/export_achievement')
def team_export_achievement():
    """团队负责人导出指定类型的公开成果"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('无团队负责人权限！', 'danger')
        return redirect(url_for('index'))

    # 获取导出参数
    team_id = request.args.get('team_id', type=int)
    achievement_type = request.args.get('type')
    if not team_id or not achievement_type:
        flash('导出参数缺失！', 'danger')
        return redirect(url_for('team_achievements'))

    # 验证团队归属（当前用户是该团队负责人）
    team = db.session.get(Team, team_id)
    if not team or team.leader_id != current_user.id:
        flash('无权限导出该团队成果！', 'danger')
        return redirect(url_for('team_achievements'))

    # 成果类型映射
    type_mapping = {
        '期刊论文': (JournalPaper, 'journal'),
        '会议论文': (ConferencePaper, 'conference'),
        '教材': (Textbook, 'textbook'),
        '专著': (Monograph, 'monograph'),
        '教研项目': (TeachingProject, 'teaching_project'),
        '专利': (Patent, 'patent'),
        '软著': (SoftwareCopyright, 'software_copyright'),
        '教学成果获奖': (TeachingAchievementAward, 'teaching_achievement_award'),
        '教学竞赛获奖': (TeachingCompetitionAward, 'teaching_competition_award'),
        '指导学生获奖': (StudentGuidanceAward, 'student_guidance_award')
    }

    if achievement_type not in type_mapping:
        flash('不支持的成果类型！', 'danger')
        return redirect(url_for('team_achievements'))

    model, export_type = type_mapping[achievement_type]
    team_id_str = str(team_id)

    # 过滤：仅导出公开给该团队的成果
    query = model.query.filter(
        func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id_str, ',')) > 0
    )

    # 字段配置
    fields_config_map = {
        'journal_paper': [
            {'name': 'title', 'label': '论文名称'},
            {'name': 'authors', 'label': '论文作者'},
            {'name': 'journal_name', 'label': '期刊名称'},
            {'name': 'inclusion_status', 'label': '收录情况'},
            {'name': 'year', 'label': '年'},
            {'name': 'volume', 'label': '卷'},
            {'name': 'issue', 'label': '期'},
            {'name': 'page_range', 'label': '起止页码'},
            {'name': 'doi', 'label': 'DOI'},
            {'name': 'publish_year', 'label': '发表年份'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'conference_paper': [
            {'name': 'title', 'label': '论文名称'},
            {'name': 'authors', 'label': '论文作者'},
            {'name': 'conference_name', 'label': '会议名称'},
            {'name': 'conference_time', 'label': '会议时间'},
            {'name': 'conference_place', 'label': '会议地点'},
            {'name': 'page_range', 'label': '起止页码'},
            {'name': 'doi', 'label': 'DOI'},
            {'name': 'publish_year', 'label': '发表年份'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'textbook': [
            {'name': 'title', 'label': '教材名称'},
            {'name': 'textbook_series', 'label': '教材系列'},
            {'name': 'chief_editor', 'label': '主编'},
            {'name': 'associate_editors', 'label': '副主编'},
            {'name': 'editorial_board', 'label': '编委'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'publication_year', 'label': '出版年份'},
            {'name': 'textbook_attachment', 'label': '附件'}
        ],
        'monograph': [
            {'name': 'title', 'label': '专著名称'},
            {'name': 'textbook_series', 'label': '专著系列'},
            {'name': 'chief_editor', 'label': '主编'},
            {'name': 'associate_editors', 'label': '副主编'},
            {'name': 'editorial_board', 'label': '编委'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'publication_year', 'label': '出版年份'},
            {'name': 'monograph_attachment', 'label': '附件'}
        ],
        'teaching_project': [
            {'name': 'title', 'label': '项目名称'},
            {'name': 'project_code', 'label': '项目编号'},
            {'name': 'project_leader', 'label': '项目负责人'},
            {'name': 'project_members', 'label': '项目参与人'},
            {'name': 'approval_department', 'label': '批准部门'},
            {'name': 'approval_date', 'label': '立项时间'},
            {'name': 'funding', 'label': '经费'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'patent': [
            {'name': 'title', 'label': '专利名称'},
            {'name': 'inventors', 'label': '发明人'},
            {'name': 'patent_number', 'label': '专利号'},
            {'name': 'apply_date', 'label': '申请日'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'software_copyright': [
            {'name': 'title', 'label': '软件名称'},
            {'name': 'copyright_owner', 'label': '著作权人'},
            {'name': 'copyright_number', 'label': '登记号'},
            {'name': 'certificate_number', 'label': '证书号'},
            {'name': 'register_date', 'label': '登记日期'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'teaching_achievement_award': [
            {'name': 'title', 'label': '成果名称'},
            {'name': 'achievement_type', 'label': '教学成果奖类型', 'relation': 'achievement_type',
             'relation_field': 'type_name'},
            {'name': 'achievement_level', 'label': '成果等级', 'relation': 'achievement_level',
             'relation_field': 'level_name'},
            {'name': 'main_contributors', 'label': '主要完成人'},
            {'name': 'completing_units', 'label': '成果完成单位'},
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'award_rank', 'label': '获奖等级', 'relation': 'award_rank', 'relation_field': 'rank_name'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'awarding_unit', 'label': '颁奖单位'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'teaching_competition_award': [
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'title', 'label': '竞赛名称'},
            {'name': 'award_rank', 'label': '获奖等级', 'relation': 'award_rank', 'relation_field': 'rank_name'},
            {'name': 'winners', 'label': '获奖人'},
            {'name': 'winner_unit', 'label': '获奖人所在单位'},
            {'name': 'competition_level', 'label': '竞赛等级', 'relation': 'competition_level', 'relation_field': 'level_name'},
            {'name': 'competition_name', 'label': '竞赛主办方'},
            {'name': 'award_date', 'label': '获奖日期'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'attachment', 'label': '附件'}
        ],
        'student_guidance_award': [
            {'name': 'award_year', 'label': '获奖年度'},
            {'name': 'title', 'label': '竞赛名称'},
            {'name': 'award_rank', 'label': '获奖等级', 'relation': 'award_rank', 'relation_field': 'rank_name'},
            {'name': 'student_name', 'label': '获奖学生'},
            {'name': 'project_name', 'label': '获奖项目名称'},
            {'name': 'teacher_name', 'label': '指导教师'},
            {'name': 'student_unit', 'label': '获奖学生所在单位'},
            {'name': 'competition_level', 'label': '竞赛等级', 'relation': 'competition_level', 'relation_field': 'level_name'},
            {'name': 'organizer', 'label': '竞赛主办方'},
            {'name': 'certificate_number', 'label': '证书编号'},
            {'name': 'attachment', 'label': '附件'}
        ]
    }

    fields_config = fields_config_map.get(export_type, [])
    items = query.all()

    # 核心修改：所有类型统一导出为Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{achievement_type}公开成果'

    # 表头
    headers = [f['label'] for f in fields_config]
    ws.append(headers)

    # 数据行
    for item in items:
        row = []
        for field in fields_config:
            field_name = field['name']
            value = getattr(item, field_name, '')

            # 处理关联表字段
            if field.get('relation'):
                relation_obj = getattr(item, field['relation'], None)
                if relation_obj:
                    value = getattr(relation_obj, field.get('relation_field', 'name'), '')
                else:
                    value = ''

            if value is None:
                value = ''
            elif isinstance(value, (date, datetime)):
                value = value.strftime('%Y-%m-%d') if value else ''
            elif field_name == 'attachment' and value:
                value = os.path.basename(value) if value else ''
            row.append(value)
        ws.append(row)

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'{achievement_type}_公开成果_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/team/manage_members', methods=['GET', 'POST'])
def manage_members():
    """团队负责人-团队成员管理（添加/移除）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('无团队负责人权限！', 'danger')
        return redirect(url_for('index'))

    # 处理成员操作
    if request.method == 'POST':
        try:
            action = request.form.get('action')
            if action == 'add_member':
                # 添加成员
                team_id = request.form.get('team_id')
                user_id = request.form.get('user_id')

                if UserTeam.query.filter_by(team_id=team_id, user_id=user_id).first():
                    flash('该用户已在团队中！', 'danger')
                else:
                    db.session.add(UserTeam(team_id=team_id, user_id=user_id))
                    flash('成员添加成功！', 'success')

            elif action == 'remove_member':
                # 移除成员
                ut_id = request.form.get('ut_id')
                ut = db.session.get(UserTeam, ut_id)
                if ut:
                    db.session.delete(ut)
                    flash('成员移除成功！', 'success')

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

    # 获取当前用户管理的团队
    teams = Team.query.filter_by(leader_id=current_user.id).all()
    if not teams:
        content = '<div class="alert alert-warning">您尚未管理任何团队！</div>'
        return render_base_layout('团队成员管理', content, current_user)

    # 可添加的用户（非管理员、未加入当前团队）
    all_users = User.query.filter(User.role != 'admin').all()

    # 渲染成员管理页面
    member_html = '''
    <h2>团队成员管理</h2>
    '''
    for team in teams:
        # 查询团队现有成员
        members = UserTeam.query.filter_by(team_id=team.id).all()
        # 可添加的用户（排除已加入的）
        available_users = []
        for user in all_users:
            is_in_team = any(ut.user_id == user.id for ut in members)
            if not is_in_team:
                available_users.append(user)

        # 团队卡片
        member_html += f'''
        <div style="margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
            <h3>{team.name}</h3>

            <!-- 添加成员 -->
            <div style="margin-bottom:20px;">
                <form method="POST" style="display:flex; gap:10px; align-items:end;">
                    <input type="hidden" name="action" value="add_member">
                    <input type="hidden" name="team_id" value="{team.id}">
                    <div class="form-group" style="flex:1;">
                        <label>添加团队成员</label>
                        <select name="user_id" required>
                            <option value="">请选择用户</option>
        '''
        # 填充可选用户
        for user in available_users:
            member_html += f'<option value="{user.id}">{user.username}（{user.employee_id}）</option>'
        member_html += '''
                        </select>
                    </div>
                    <button type="submit" class="btn">添加</button>
                </form>
            </div>

            <!-- 现有成员 -->
            <div>
                <strong>当前成员：</strong>
                <ul style="margin:10px 0; padding-left:20px;">
        '''
        # 成员列表
        for ut in members:
            user = db.session.get(User, ut.user_id)
            # 负责人不可移除
            if user.id == team.leader_id:
                member_html += f'<li>{user.username}（{user.employee_id}）<span style="color:#999;">（团队负责人）</span></li>'
            else:
                member_html += f'''
                <li>
                    {user.username}（{user.employee_id}）
                    <form method="POST" style="display:inline; margin-left:10px;">
                        <input type="hidden" name="action" value="remove_member">
                        <input type="hidden" name="ut_id" value="{ut.id}">
                        <button type="submit" class="btn" style="padding:2px 8px; font-size:12px;" onclick="return confirm('确定移除？')">移除</button>
                    </form>
                </li>
                '''
        member_html += '''
                </ul>
            </div>
        </div>
        '''

    return render_base_layout('团队成员管理', member_html, current_user)


@app.route('/download')
def download_file():
    """通用文件下载"""
    file_path = request.args.get('path')
    if not file_path or not os.path.exists(file_path):
        flash('文件不存在！', 'danger')
        return redirect(url_for('index'))
    return send_file(file_path, as_attachment=True)


# 1. 期刊论文管理
@app.route('/achievement/journal_paper', methods=['GET', 'POST'])
def journal_paper_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '论文名称', 'type': 'text', 'required': True},
        {'name': 'authors', 'label': '论文作者', 'type': 'text', 'required': True},
        {'name': 'corresponding_authors', 'label': '通讯作者', 'type': 'text'},
        {'name': 'journal_name', 'label': '期刊名称', 'type': 'text', 'required': True},
        {'name': 'inclusion_type_ids', 'label': '收录情况', 'type': 'select_multiple', 'options': []},
        {'name': 'year', 'label': '年', 'type': 'integer'},
        {'name': 'volume', 'label': '卷', 'type': 'text'},
        {'name': 'issue', 'label': '期', 'type': 'text'},
        {'name': 'page_range', 'label': '起止页码', 'type': 'text'},
        {'name': 'doi', 'label': 'DOI', 'type': 'text'},
        {'name': 'publish_year', 'label': '发表年份', 'type': 'integer'},
        {'name': 'publish_date', 'label': '发表日期', 'type': 'date'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '论文附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')

    if request.method == 'POST':
        return handle_achievement_submit(JournalPaper, fields_config)

    if action == 'add':
        # 获取收录类型选项
        inclusion_types = InclusionType.query.filter_by(is_active=True).order_by(InclusionType.sort_order).all()
        fields_config[4]['options'] = [t.type_name for t in inclusion_types]
        return render_achievement_form(JournalPaper, '新增期刊论文', fields_config)
    elif action == 'edit':
        # 获取收录类型选项
        inclusion_types = InclusionType.query.filter_by(is_active=True).order_by(InclusionType.sort_order).all()
        fields_config[4]['options'] = [t.type_name for t in inclusion_types]
        item_id = request.args.get('id')
        return render_achievement_form(JournalPaper, '修改期刊论文', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(JournalPaper, item_id)
    # 核心修改：传递时间筛选参数
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(JournalPaper, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(JournalPaper, '期刊论文', fields_config, current_user)
    else:
        return render_achievement_list(JournalPaper, '期刊论文管理', fields_config, current_user)

# 2. 会议论文管理 - 修改导出逻辑
@app.route('/achievement/conference_paper', methods=['GET', 'POST'])
def conference_paper_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '论文名称', 'type': 'text', 'required': True},
        {'name': 'authors', 'label': '论文作者', 'type': 'text', 'required': True},
        {'name': 'corresponding_authors', 'label': '通讯作者', 'type': 'text'},
        {'name': 'conference_name', 'label': '会议名称', 'type': 'text', 'required': True},
        {'name': 'conference_start_date', 'label': '会议开始日期', 'type': 'date'},
        {'name': 'conference_end_date', 'label': '会议结束日期', 'type': 'date'},
        {'name': 'conference_place', 'label': '会议地点', 'type': 'text'},
        {'name': 'page_range', 'label': '起止页码', 'type': 'text'},
        {'name': 'doi', 'label': 'DOI', 'type': 'text'},
        {'name': 'publish_year', 'label': '发表年份', 'type': 'integer'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '论文附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(ConferencePaper, fields_config)

    if action == 'add':
        return render_achievement_form(ConferencePaper, '新增会议论文', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(ConferencePaper, '修改会议论文', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(ConferencePaper, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(ConferencePaper, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(ConferencePaper, '会议论文', fields_config, current_user)
    else:
        return render_achievement_list(ConferencePaper, '会议论文管理', fields_config, current_user)

# 3. 教材管理（复用通用函数）
@app.route('/achievement/textbook', methods=['GET', 'POST'])
def textbook_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 核心修改：更新字段配置
    fields_config = [
        {'name': 'title', 'label': '教材名称', 'type': 'text', 'required': True},
        {'name': 'textbook_series', 'label': '教材系列', 'type': 'text'},
        {'name': 'chief_editor', 'label': '主编', 'type': 'text'},
        {'name': 'associate_editors', 'label': '副主编', 'type': 'text'},
        {'name': 'editorial_board', 'label': '编委', 'type': 'text'},
        {'name': 'publisher', 'label': '出版社', 'type': 'text'},
        {'name': 'isbn', 'label': 'ISBN', 'type': 'text'},
        {'name': 'cip_number', 'label': 'CIP 核字号', 'type': 'text'},
        {'name': 'publication_year', 'label': '出版年份', 'type': 'integer'},
        {'name': 'publication_month', 'label': '出版月份', 'type': 'integer'},
        {'name': 'publish_date', 'label': '出版日期', 'type': 'date'},
        {'name': 'edition', 'label': '版次', 'type': 'text'},
        {'name': 'word_count', 'label': '字数', 'type': 'text'},
        {'name': 'price', 'label': '定价', 'type': 'text'},
        # 教材级别下拉框（从数据库读取）
        {'name': 'textbook_level_id', 'label': '教材级别', 'type': 'select', 'options': []},
        # 教材类型下拉框
        {'name': 'textbook_type', 'label': '教材类型', 'type': 'select', 'options': [
            '纸质教材', '数字教材'
        ]},
        {'name': 'applicable_majors', 'label': '适用专业', 'type': 'text'},
        {'name': 'remarks', 'label': '备注', 'type': 'text'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        # 附件字段名更新
        {'name': 'textbook_attachment', 'label': '教材附件', 'type': 'file'}
    ]

    # 处理表单提交（适配新字段）
    if request.method == 'POST':
        item_id = request.form.get('id')
        item = db.session.get(Textbook, item_id) if item_id else None

        if item and item.user_id != current_user.id:
            flash('无权限修改该成果！', 'danger')
            return redirect(url_for('textbook_manage'))

        if not item:
            item = Textbook()
            item.user_id = current_user.id
            item.create_time = datetime.now()

        # 处理普通字段（包含新增字段）
        for field in fields_config:
            field_name = field['name']
            field_type = field.get('type', 'text')
            if field_type == 'file':
                continue

            value = request.form.get(field_name, '').strip()

            # 日期字段处理
            if field_type == 'date':
                if value == '':
                    value = None
                else:
                    try:
                        value = datetime.strptime(value, '%Y-%m-%d').date()
                    except ValueError:
                        flash(f'{field["label"]}格式错误（需为 YYYY-MM-DD）！', 'danger')
                        return redirect(url_for('textbook_manage', action='add'))
            # 整数字段（出版年/月）
            elif field_type == 'integer':
                if value == '':
                    value = None
                else:
                    try:
                        value = int(value)
                    except ValueError:
                        flash(f'{field["label"]}必须为数字！', 'danger')
                        return redirect(url_for('textbook_manage', action='add'))
            # 普通文本字段
            elif value == '':
                value = None

            setattr(item, field_name, value)

        # 处理文件上传（字段名更新为 textbook_attachment）
        file = request.files.get('textbook_attachment')
        if file and file.filename:
            old_path = getattr(item, 'textbook_attachment', '')
            if old_path and os.path.exists(old_path):
                os.remove(old_path)
            sub_folder = 'textbook'
            new_path = handle_file_upload(file, sub_folder)
            setattr(item, 'textbook_attachment', new_path)

        item.update_time = datetime.now()

        # 强制提交数据库
        try:
            if not item_id:
                db.session.add(item)
            db.session.commit()
            flash(f'{"修改" if item_id else "新增"}教材成功！', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

        return redirect(url_for('textbook_manage'))

    # 其余逻辑（action 分支）保持不变，仅渲染和列表展示会自动适配新字段
    action = request.args.get('action', 'list')
    if action == 'add':
        # 获取教材级别选项
        levels = TextbookLevel.query.filter_by(is_active=True).order_by(TextbookLevel.sort_order).all()
        fields_config[14]['options'] = [l.level_name for l in levels]
        return render_achievement_form(Textbook, '新增教材', fields_config)
    elif action == 'edit':
        # 获取教材级别选项
        levels = TextbookLevel.query.filter_by(is_active=True).order_by(TextbookLevel.sort_order).all()
        fields_config[14]['options'] = [l.level_name for l in levels]
        item_id = request.args.get('id')
        return render_achievement_form(Textbook, '修改教材', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(Textbook, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(Textbook, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(Textbook, '教材', fields_config, current_user)
    else:
        return render_achievement_list(Textbook, '教材管理', fields_config, current_user)



# 4. 专著管理（复用通用函数）
@app.route('/achievement/monograph', methods=['GET', 'POST'])
def monograph_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '专著名称', 'type': 'text', 'required': True},
        {'name': 'textbook_series', 'label': '专著系列', 'type': 'text'},
        {'name': 'chief_editor', 'label': '主编', 'type': 'text'},
        {'name': 'associate_editors', 'label': '副主编', 'type': 'text'},
        {'name': 'editorial_board', 'label': '编委', 'type': 'text'},
        {'name': 'publisher', 'label': '出版社', 'type': 'text'},
        {'name': 'isbn', 'label': 'ISBN', 'type': 'text'},
        {'name': 'cip_number', 'label': 'CIP 核字号', 'type': 'text'},
        {'name': 'publication_year', 'label': '出版年份', 'type': 'integer'},
        {'name': 'publication_month', 'label': '出版月份', 'type': 'integer'},
        {'name': 'publish_date', 'label': '出版日期', 'type': 'date'},
        {'name': 'edition', 'label': '版次', 'type': 'text'},
        {'name': 'word_count', 'label': '字数', 'type': 'text'},
        {'name': 'price', 'label': '定价', 'type': 'text'},
        # 专著类型下拉框
        {'name': 'monograph_type', 'label': '专著类型', 'type': 'select', 'options': [
            '学术专著', '技术专著', '科普著作', '其它'
        ]},
        {'name': 'applicable_majors', 'label': '适用专业', 'type': 'text'},
        {'name': 'remarks', 'label': '备注', 'type': 'text'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        # 附件字段名更新
        {'name': 'monograph_attachment', 'label': '专著附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        item_id = request.form.get('id')
        item = db.session.get(Monograph, item_id) if item_id else None

        if item and item.user_id != current_user.id:
            flash('无权限修改该成果！', 'danger')
            return redirect(url_for('monograph_manage'))

        if not item:
            item = Monograph()
            item.user_id = current_user.id
            item.create_time = datetime.now()

        # 处理普通字段（包含新增字段）
        for field in fields_config:
            field_name = field['name']
            field_type = field.get('type', 'text')
            if field_type == 'file':
                continue

            value = request.form.get(field_name, '').strip()

            # 日期字段处理
            if field_type == 'date':
                if value == '':
                    value = None
                else:
                    try:
                        value = datetime.strptime(value, '%Y-%m-%d').date()
                    except ValueError:
                        flash(f'{field["label"]}格式错误（需为 YYYY-MM-DD）！', 'danger')
                        return redirect(url_for('monograph_manage', action='add'))
            # 整数字段（出版年/月）
            elif field_type == 'integer':
                if value == '':
                    value = None
                else:
                    try:
                        value = int(value)
                    except ValueError:
                        flash(f'{field["label"]}必须为数字！', 'danger')
                        return redirect(url_for('monograph_manage', action='add'))
            # 普通文本字段
            elif value == '':
                value = None

            setattr(item, field_name, value)

        # 处理文件上传（字段名更新为 monograph_attachment）
        file = request.files.get('monograph_attachment')
        if file and file.filename:
            old_path = getattr(item, 'monograph_attachment', '')
            if old_path and os.path.exists(old_path):
                os.remove(old_path)
            sub_folder = 'monograph'
            new_path = handle_file_upload(file, sub_folder)
            setattr(item, 'monograph_attachment', new_path)

        item.update_time = datetime.now()

        # 强制提交数据库
        try:
            if not item_id:
                db.session.add(item)
            db.session.commit()
            flash(f'{"修改" if item_id else "新增"}专著成功！', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

        return redirect(url_for('monograph_manage'))

    # 其余逻辑（action 分支）保持不变，仅渲染和列表展示会自动适配新字段
    action = request.args.get('action', 'list')
    if action == 'add':
        return render_achievement_form(Monograph, '新增专著', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(Monograph, '修改专著', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(Monograph, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(Monograph, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(Monograph, '专著', fields_config, current_user)
    else:
        return render_achievement_list(Monograph, '专著管理', fields_config, current_user)

# 5. 教研教改和课程建设项目管理（复用通用函数）
@app.route('/achievement/teaching_project', methods=['GET', 'POST'])
def teaching_project_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 获取字典表选项
    project_types = [t.type_name for t in ProjectType.query.order_by(ProjectType.sort_order).all()]
    project_levels = [l.level_name for l in ProjectLevel.query.order_by(ProjectLevel.sort_order).all()]
    project_categories = [c.category_name for c in ProjectCategory.query.order_by(ProjectCategory.sort_order).all()]
    project_statuses = [s.status_name for s in ProjectStatus.query.order_by(ProjectStatus.sort_order).all()]
    
    fields_config = [
        {'name': 'title', 'label': '项目名称', 'type': 'text', 'required': True},
        {'name': 'project_code', 'label': '项目编号', 'type': 'text'},
        {'name': 'project_type_id', 'label': '项目类型', 'type': 'select', 'options': project_types},
        {'name': 'project_leader', 'label': '项目负责人', 'type': 'text'},
        {'name': 'project_members', 'label': '项目参与人', 'type': 'text', 'placeholder': '多人请用顿号分隔'},
        {'name': 'approval_department', 'label': '项目批准部门', 'type': 'text'},
        {'name': 'approval_date', 'label': '项目立项时间', 'type': 'month'},
        {'name': 'project_level_id', 'label': '项目级别', 'type': 'select', 'options': project_levels},
        {'name': 'project_category_id', 'label': '项目类别', 'type': 'select', 'options': project_categories},
        {'name': 'funding', 'label': '项目经费（元）', 'type': 'number', 'step': '0.01'},
        {'name': 'start_date', 'label': '项目开始时间', 'type': 'date'},
        {'name': 'end_date', 'label': '项目结束时间', 'type': 'date'},
        {'name': 'project_status_id', 'label': '项目状态', 'type': 'select', 'options': project_statuses},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '项目附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(TeachingProject, fields_config)

    if action == 'add':
        return render_achievement_form(TeachingProject, '新增教研教改和课程建设项目', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(TeachingProject, '修改教研教改和课程建设项目', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(TeachingProject, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(TeachingProject, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(TeachingProject, '教研教改和课程建设项目', fields_config, current_user)
    else:
        return render_achievement_list(TeachingProject, '教研教改和课程建设项目管理', fields_config, current_user)

# 6. 专利管理（复用通用函数）

@app.route('/achievement/patent', methods=['GET', 'POST'])
def patent_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '专利名称', 'type': 'text', 'required': True},
        {'name': 'patent_type_id', 'label': '专利类型', 'type': 'select',
         'options': [(t.id, t.type_name) for t in PatentType.query.order_by(PatentType.sort_order).all()], 'required': True},
        {'name': 'patentee', 'label': '专利权人', 'type': 'text'},
        {'name': 'address', 'label': '地址', 'type': 'text'},
        {'name': 'inventors', 'label': '发明人', 'type': 'text'},
        {'name': 'patent_status_id', 'label': '专利状态', 'type': 'select',
         'options': [(s.id, s.status_name) for s in PatentStatus.query.order_by(PatentStatus.sort_order).all()]},
        {'name': 'patent_number', 'label': '专利号', 'type': 'text'},
        {'name': 'grant_announcement_number', 'label': '授权公告号', 'type': 'text'},
        {'name': 'apply_date', 'label': '专利申请日', 'type': 'date'},
        {'name': 'grant_announcement_date', 'label': '授权公告日', 'type': 'date'},
        {'name': 'applicant_at_apply_date', 'label': '申请日时申请人', 'type': 'text'},
        {'name': 'inventor_at_apply_date', 'label': '申请日时发明人', 'type': 'text'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(Patent, fields_config)

    if action == 'add':
        return render_achievement_form(Patent, '新增专利', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(Patent, '修改专利', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(Patent, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(Patent, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(Patent, '专利', fields_config, current_user)
    else:
        return render_achievement_list(Patent, '专利管理', fields_config, current_user)


# 7. 软件著作管理（复用通用函数）
@app.route('/achievement/software_copyright', methods=['GET', 'POST'])
def software_copyright_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '软件名称', 'type': 'text', 'required': True},
        {'name': 'copyright_owner', 'label': '著作权人', 'type': 'text'},
        {'name': 'completion_date', 'label': '开发完成日期', 'type': 'date'},
        {'name': 'first_publication_date', 'label': '首次发表日期', 'type': 'date'},
        {'name': 'right_acquisition_method', 'label': '权利取得方式', 'type': 'select', 'options': ['原始取得', '受让取得', '继承取得', '其他']},
        {'name': 'right_scope', 'label': '权利范围', 'type': 'select', 'options': ['全部权利', '部分权利']},
        {'name': 'copyright_number', 'label': '登记号', 'type': 'text'},
        {'name': 'certificate_number', 'label': '证书号', 'type': 'text'},
        {'name': 'register_date', 'label': '登记日期', 'type': 'date'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(SoftwareCopyright, fields_config)

    if action == 'add':
        return render_achievement_form(SoftwareCopyright, '新增软件著作', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(SoftwareCopyright, '修改软件著作', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(SoftwareCopyright, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(SoftwareCopyright, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(SoftwareCopyright, '软件著作', fields_config, current_user)
    else:
        return render_achievement_list(SoftwareCopyright, '软件著作管理', fields_config, current_user)


# 8. 教学成果获奖管理（复用通用函数）
@app.route('/achievement/teaching_achievement_award', methods=['GET', 'POST'])
def teaching_achievement_award_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '成果名称', 'type': 'text', 'required': True},
        {'name': 'achievement_type_id', 'label': '教学成果奖类型', 'type': 'select', 'options': []},
        {'name': 'achievement_level_id', 'label': '成果等级', 'type': 'select', 'options': []},
        {'name': 'main_contributors', 'label': '主要完成人', 'type': 'text', 'placeholder': '多人用分号分隔'},
        {'name': 'completing_units', 'label': '成果完成单位', 'type': 'text', 'placeholder': '多个用分号分隔'},
        {'name': 'award_year', 'label': '获奖年度', 'type': 'integer'},
        {'name': 'award_rank_id', 'label': '获奖等级', 'type': 'select', 'options': []},
        {'name': 'certificate_number', 'label': '证书编号', 'type': 'text'},
        {'name': 'awarding_unit', 'label': '颁奖单位', 'type': 'text'},
        {'name': 'award_date', 'label': '获奖日期', 'type': 'date'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(TeachingAchievementAward, fields_config)

    if action == 'add':
        achievement_types = TeachingAchievementType.query.filter_by(is_active=True).order_by(TeachingAchievementType.sort_order).all()
        achievement_levels = AchievementLevel.query.filter_by(is_active=True).order_by(AchievementLevel.sort_order).all()
        award_ranks = AwardRank.query.filter_by(is_active=True).order_by(AwardRank.sort_order).all()
        fields_config[1]['options'] = [(t.id, t.type_name) for t in achievement_types]
        fields_config[2]['options'] = [(l.id, l.level_name) for l in achievement_levels]
        fields_config[6]['options'] = [(r.id, r.rank_name) for r in award_ranks]
        return render_achievement_form(TeachingAchievementAward, '新增教学成果获奖', fields_config)
    elif action == 'edit':
        achievement_types = TeachingAchievementType.query.filter_by(is_active=True).order_by(TeachingAchievementType.sort_order).all()
        achievement_levels = AchievementLevel.query.filter_by(is_active=True).order_by(AchievementLevel.sort_order).all()
        award_ranks = AwardRank.query.filter_by(is_active=True).order_by(AwardRank.sort_order).all()
        fields_config[1]['options'] = [(t.id, t.type_name) for t in achievement_types]
        fields_config[2]['options'] = [(l.id, l.level_name) for l in achievement_levels]
        fields_config[6]['options'] = [(r.id, r.rank_name) for r in award_ranks]
        item_id = request.args.get('id')
        return render_achievement_form(TeachingAchievementAward, '修改教学成果获奖', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(TeachingAchievementAward, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(TeachingAchievementAward, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(TeachingAchievementAward, '教学成果获奖', fields_config, current_user)
    else:
        return render_achievement_list(TeachingAchievementAward, '教学成果获奖管理', fields_config, current_user)


# 9. 教学竞赛获奖管理（复用通用函数）
@app.route('/achievement/teaching_competition_award', methods=['GET', 'POST'])
def teaching_competition_award_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '竞赛名称', 'type': 'text', 'required': True},
        {'name': 'award_year', 'label': '获奖年度', 'type': 'text'},
        {'name': 'competition_level_id', 'label': '竞赛等级', 'type': 'select',
         'options_from_model': 'AchievementLevel'},
        {'name': 'award_rank_id', 'label': '获奖等级', 'type': 'select', 'options_from_model': 'AwardRank'},
        {'name': 'winners', 'label': '获奖人', 'type': 'text'},
        {'name': 'winner_unit', 'label': '获奖人所在单位', 'type': 'text'},
        {'name': 'competition_name', 'label': '竞赛主办方', 'type': 'text'},
        {'name': 'award_date', 'label': '获奖日期', 'type': 'date'},
        {'name': 'certificate_number', 'label': '证书编号', 'type': 'text'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(TeachingCompetitionAward, fields_config)

    if action == 'add':
        return render_achievement_form(TeachingCompetitionAward, '新增教学竞赛获奖', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(TeachingCompetitionAward, '修改教学竞赛获奖', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(TeachingCompetitionAward, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(TeachingCompetitionAward, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(TeachingCompetitionAward, '教学竞赛获奖', fields_config, current_user)
    else:
        return render_achievement_list(TeachingCompetitionAward, '教学竞赛获奖管理', fields_config, current_user)


# 10. 指导学生获奖管理（复用通用函数）
@app.route('/achievement/student_guidance_award', methods=['GET', 'POST'])
def student_guidance_award_manage():
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    fields_config = [
        {'name': 'title', 'label': '获奖名称', 'type': 'text', 'required': True},
        {'name': 'award_year', 'label': '获奖年度', 'type': 'text'},
        {'name': 'competition_name', 'label': '竞赛名称', 'type': 'text'},
        {'name': 'competition_level_id', 'label': '竞赛等级', 'type': 'select',
         'options_from_model': 'AchievementLevel'},
        {'name': 'award_rank_id', 'label': '获奖等级', 'type': 'select', 'options_from_model': 'AwardRank'},
        {'name': 'student_name', 'label': '获奖学生', 'type': 'text'},
        {'name': 'project_name', 'label': '获奖项目名称', 'type': 'text'},
        {'name': 'teacher_name', 'label': '指导教师', 'type': 'text'},
        {'name': 'student_unit', 'label': '获奖学生所在单位', 'type': 'text'},
        {'name': 'organizer', 'label': '竞赛主办方', 'type': 'text'},
        {'name': 'certificate_number', 'label': '证书编号', 'type': 'text'},
        {'name': 'award_date', 'label': '获奖日期', 'type': 'date'},
        {'name': 'public_team_ids', 'label': '公开团队', 'type': 'select_multiple'},
        {'name': 'attachment', 'label': '附件', 'type': 'file'}
    ]

    action = request.args.get('action', 'list')
    if request.method == 'POST':
        return handle_achievement_submit(StudentGuidanceAward, fields_config)

    if action == 'add':
        return render_achievement_form(StudentGuidanceAward, '新增指导学生获奖', fields_config)
    elif action == 'edit':
        item_id = request.args.get('id')
        return render_achievement_form(StudentGuidanceAward, '修改指导学生获奖', fields_config, item_id)
    elif action == 'delete':
        item_id = request.args.get('id')
        return handle_achievement_delete(StudentGuidanceAward, item_id)
    elif action == 'export':
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        return export_achievement_excel(StudentGuidanceAward, fields_config, current_user, start_date, end_date)
    elif action == 'stats':
        return render_achievement_stats(StudentGuidanceAward, '指导学生获奖', fields_config, current_user)
    else:
        return render_achievement_list(StudentGuidanceAward, '指导学生获奖管理', fields_config, current_user)


@app.route('/team/member_achievements')
def member_achievements():
    """团队负责人-查看成员具体成果详情"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('无团队负责人权限！', 'danger')
        return redirect(url_for('index'))

    # 获取筛选参数
    member_id = request.args.get('member_id')
    achievement_type = request.args.get('type', 'all')

    # 获取当前用户管理的团队ID
    managed_teams = Team.query.filter_by(leader_id=current_user.id).all()
    managed_team_ids = [str(t.id) for t in managed_teams]

    # 获取团队信息
    teams = Team.query.filter_by(leader_id=current_user.id).all()
    team_ids = [t.id for t in teams]
    team_user_ids = [ut.user_id for ut in UserTeam.query.filter(UserTeam.team_id.in_(team_ids)).all()]

    if not team_user_ids:
        content = '<div class="alert alert-warning">暂无团队成员数据！</div>'
        return render_base_layout('团队成员成果详情', content, current_user)

    # 获取团队成员列表
    team_members = User.query.filter(User.id.in_(team_user_ids)).all()

    # 成果类型映射
    type_mapping = {
        'all': '所有成果',
        'journal_paper': '期刊论文',
        'conference_paper': '会议论文',
        'textbook': '教材',
        'monograph': '专著',
        'teaching_project': '教研项目',
        'patent': '专利',
        'software_copyright': '软件著作',
        'teaching_achievement_award': '教学成果获奖',
        'teaching_competition_award': '教学竞赛获奖',
        'student_guidance_award': '指导学生获奖'
    }

    model_mapping = {
        'journal_paper': JournalPaper,
        'conference_paper': ConferencePaper,
        'textbook': Textbook,
        'monograph': Monograph,
        'teaching_project': TeachingProject,
        'patent': Patent,
        'software_copyright': SoftwareCopyright,
        'teaching_achievement_award': TeachingAchievementAward,
        'teaching_competition_award': TeachingCompetitionAward,
        'student_guidance_award': StudentGuidanceAward
    }

    # 生成筛选表单
    filter_html = f'''
    <div style="margin-bottom:20px; padding:20px; background:#f5f7fa; border-radius:8px;">
        <form method="GET">
            <div class="form-row" style="display:flex; gap:20px; margin-bottom:10px;">
                <div class="form-group" style="flex:1;">
                    <label>选择团队成员</label>
                    <select name="member_id" required onchange="this.form.submit()">
                        <option value="">全部成员</option>
    '''
    for member in team_members:
        selected = 'selected' if str(member.id) == member_id else ''
        filter_html += f'<option value="{member.id}" {selected}>{member.username}（{member.employee_id}）</option>'

    filter_html += f'''
                    </select>
                </div>
                <div class="form-group" style="flex:1;">
                    <label>成果类型</label>
                    <select name="type" onchange="this.form.submit()">
    '''
    for type_key, type_name in type_mapping.items():
        selected = 'selected' if type_key == achievement_type else ''
        filter_html += f'<option value="{type_key}" {selected}>{type_name}</option>'

    filter_html += '''
                    </select>
                </div>
            </div>
        </form>
    </div>
    '''

    # 查询成果数据（核心：仅显示公开给当前团队的成果）
    achievements = []

    def filter_public_achievements(query):
        """过滤出公开给当前团队的成果"""
        or_conditions = []
        for team_id in managed_team_ids:
            or_conditions.append(
                func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id, ',')) > 0
            )
        return query.filter(or_(*or_conditions))

    if member_id and achievement_type != 'all' and achievement_type in model_mapping:
        # 筛选指定成员的指定类型成果（仅公开给当前团队的）
        model = model_mapping[achievement_type]
        query = model.query.filter_by(user_id=member_id)
        query = filter_public_achievements(query)
        achievements = query.order_by(model.update_time.desc()).all()

    elif member_id and achievement_type == 'all':
        # 筛选指定成员的所有成果（仅公开给当前团队的）
        for model in model_mapping.values():
            query = model.query.filter_by(user_id=member_id)
            query = filter_public_achievements(query)
            items = query.order_by(model.update_time.desc()).all()
            for item in items:
                item.type_name = [k for k, v in model_mapping.items() if v == model][0]
                achievements.append(item)

    elif achievement_type != 'all' and achievement_type in model_mapping:
        # 筛选所有成员的指定类型成果（仅公开给当前团队的）
        model = model_mapping[achievement_type]
        query = model.query.filter(model.user_id.in_(team_user_ids))
        query = filter_public_achievements(query)
        achievements = query.order_by(model.update_time.desc()).all()

    else:
        # 所有成果（仅公开给当前团队的）
        for model in model_mapping.values():
            query = model.query.filter(model.user_id.in_(team_user_ids))
            query = filter_public_achievements(query)
            items = query.order_by(model.update_time.desc()).all()
            for item in items:
                item.type_name = [k for k, v in model_mapping.items() if v == model][0]
                achievements.append(item)

    # 生成成果列表
    list_html = '''
    <table style="width:100%; border-collapse:collapse;">
        <thead>
            <tr style="background:#e9ecef;">
                <th style="padding:10px; border:1px solid #dee2e6;">成果类型</th>
                <th style="padding:10px; border:1px solid #dee2e6;">成果名称</th>
                <th style="padding:10px; border:1px solid #dee2e6;">所属成员</th>
                <th style="padding:10px; border:1px solid #dee2e6;">创建时间</th>
                <th style="padding:10px; border:1px solid #dee2e6;">操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    if not achievements:
        list_html += '''
        <tr>
            <td colspan="5" style="padding:20px; text-align:center; border:1px solid #dee2e6;">暂无成果数据（仅显示成员公开给本团队的成果）</td>
        </tr>
        '''
    else:
        for item in achievements:
            # 获取成果类型名称
            if hasattr(item, 'type_name'):
                type_name = type_mapping.get(item.type_name, '未知类型')
            else:
                type_name = type_mapping.get(achievement_type, '未知类型')

            # 获取所属成员
            member = User.query.get(item.user_id)
            member_name = f'{member.username}（{member.employee_id}）' if member else '未知成员'

            # 创建时间
            create_time = item.create_time.strftime('%Y-%m-%d %H:%M') if hasattr(item, 'create_time') else ''

            # 查看详情链接
            detail_link = f'/achievement/{achievement_type if achievement_type != "all" else item.type_name}?action=edit&id={item.id}'

            list_html += f'''
            <tr>
                <td style="padding:10px; border:1px solid #dee2e6;">{type_name}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">{getattr(item, 'title', '无名称')}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">{member_name}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">{create_time}</td>
                <td style="padding:10px; border:1px solid #dee2e6;">
                    <a href="{detail_link}" class="btn" style="padding:5px 10px; font-size:12px;">查看详情</a>
                </td>
            </tr>
            '''

    list_html += '''
        </tbody>
    </table>
    '''

    content = filter_html + list_html
    return render_base_layout('团队成员成果详情', content, current_user)


# ---------------------- 多维度统计 + 图表展示 ----------------------
@app.route('/stats/dashboard')
def stats_dashboard():
    """多维度统计仪表盘（个人/团队）- 移除年度成果趋势"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 确定统计范围
    if current_user.role == 'team_leader':
        # 团队负责人：统计团队数据
        teams = Team.query.filter_by(leader_id=current_user.id).all()
        team_ids = [t.id for t in teams]
        user_ids = [ut.user_id for ut in UserTeam.query.filter(UserTeam.team_id.in_(team_ids)).all()] + [
            current_user.id]
        stats_scope = '团队'
    else:
        # 普通教师：统计个人数据
        user_ids = [current_user.id]
        stats_scope = '个人'

    def count_user_achievements(model_class):
        """统计用户参与的成果数量"""
        # 统计直接拥有的成果数量（user_id 在 user_ids 中）
        count = model_class.query.filter(model_class.user_id.in_(user_ids)).count()
        return count

    # 成果类型列表（移除作者关联表）
    achievement_types = [
        ('期刊论文', JournalPaper),
        ('会议论文', ConferencePaper),
        ('教材', Textbook),
        ('专著', Monograph),
        ('教研项目', TeachingProject),
        ('专利', Patent),
        ('软著', SoftwareCopyright),
        ('教学成果获奖', TeachingAchievementAward),
        ('教学竞赛获奖', TeachingCompetitionAward),
        ('指导学生获奖', StudentGuidanceAward)
    ]

    type_stats = []
    total_count = 0

    for name, model in achievement_types:
        count = count_user_achievements(model)
        type_stats.append({'name': name, 'value': count})
        total_count += count


    # 准备饼图数据
    type_labels = [item['name'] for item in type_stats]
    type_values = [item['value'] for item in type_stats]

    # 渲染统计仪表盘（仅保留成果类型分布饼图）
    content = f'''
    <h2>{stats_scope}成果统计仪表盘</h2>
    <div style="margin-bottom:30px; font-size:18px;">
        成果总数：<strong style="color:#3498db; font-size:24px;">{total_count}</strong> 项
    </div>

    <!-- 成果类型分布（饼图） -->
    <div style="margin-bottom:40px; height:400px;">
        <h3 style="margin-bottom:10px;">成果类型分布</h3>
        <canvas id="typeChart"></canvas>
    </div>

    <!-- 引入Chart.js -->
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script>
        // 饼图 - 成果类型分布
        const typeCtx = document.getElementById('typeChart').getContext('2d');
        new Chart(typeCtx, {{
            type: 'pie',
            data: {{
                labels: {json.dumps(type_labels)},
                datasets: [{{
                    label: '成果数量',
                    data: {json.dumps(type_values)},
                    backgroundColor: [
                        '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF',
                        '#C9CBCF', '#FF9F40', '#FFCD56', '#45B7D1', '#66AA00'
                    ],
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{
                        position: 'right',
                    }}
                }}
            }}
        }});
    </script>
    '''

    return render_base_layout(f'{stats_scope}成果统计仪表盘', content, current_user)

    # ---------------------- 团队列表与创建功能 ----------------------


@app.route('/team/list', methods=['GET', 'POST'])
def team_list():
    """团队列表（查看所有团队）+ 创建团队按钮"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 处理创建团队请求（仅管理员/团队负责人可创建）
    if request.method == 'POST':
        # 权限控制：仅管理员或团队负责人可创建团队
        if current_user.role not in ['admin', 'team_leader']:
            flash('无创建团队权限！', 'danger')
            return redirect(url_for('team_list'))

        team_name = request.form.get('team_name', '').strip()
        if not team_name:
            flash('团队名称不能为空！', 'danger')
            return redirect(url_for('team_list'))

        # 检查团队名称是否重复
        if Team.query.filter_by(name=team_name).first():
            flash('团队名称已存在！', 'danger')
            return redirect(url_for('team_list'))

        # 创建团队（负责人为当前用户）
        try:
            new_team = Team(
                name=team_name,
                leader_id=current_user.id
            )
            db.session.add(new_team)
            # 先提交获取team_id
            db.session.flush()  # 关键：先刷新会话，生成new_team.id但不提交事务
            # 再创建用户-团队关联
            db.session.add(UserTeam(user_id=current_user.id, team_id=new_team.id))
            db.session.commit()
            flash(f'团队「{team_name}」创建成功！', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'创建团队失败：{str(e)}', 'danger')

        return redirect(url_for('team_list'))

    # 根据用户角色筛选团队列表
    if current_user.role == 'admin':
        # 管理员：查看所有团队
        teams = Team.query.order_by(Team.create_time.desc()).all()
    elif current_user.role == 'team_leader':
        # 团队负责人：查看自己管理的团队
        teams = Team.query.filter_by(leader_id=current_user.id).order_by(Team.create_time.desc()).all()
    else:
        # 普通教师：查看自己加入的团队
        user_teams = UserTeam.query.filter_by(user_id=current_user.id).all()
        team_ids = [ut.team_id for ut in user_teams]
        teams = Team.query.filter(Team.id.in_(team_ids)).order_by(Team.create_time.desc()).all()

    # 构建团队列表HTML
    team_list_html = f'''
        <h2>团队管理</h2>

        <!-- 创建团队按钮 + 表单 -->
        <div style="margin-bottom:30px;">
            <button onclick="toggleCreateForm()" class="btn" style="background:#27ae60;">📝 创建新团队</button>

            <!-- 创建团队表单（默认隐藏） -->
            <div id="createTeamForm" style="display:none; margin-top:20px; padding:20px; border:1px solid #eee; border-radius:8px;">
                <h3 style="margin-bottom:20px;">创建新团队</h3>
                <form method="POST">
                    <div class="form-group">
                        <label>团队名称 <span class="required" style="color:red;">*</span></label>
                        <input type="text" name="team_name" required placeholder="请输入团队名称">
                    </div>
                    <button type="submit" class="btn">确认创建</button>
                    <button type="button" onclick="toggleCreateForm()" class="btn" style="background:#95a5a6; margin-left:10px;">取消</button>
                </form>
            </div>
        </div>

        <!-- 团队列表 -->
        <div style="margin-top:20px;">
            <h3>{"所有团队" if current_user.role == 'admin' else "我的团队"}</h3>
            {f'<div class="alert alert-info">暂无团队数据</div>' if not teams else ''}

            <div style="display:grid; grid-template-columns: repeat(auto-fill, minmax(300px, 1fr)); gap:20px; margin-top:20px;">
        '''

    # 渲染每个团队卡片
    for team in teams:
        leader = User.query.get(team.leader_id)
        leader_name = leader.username if leader else '未知'

        # 获取团队成员数量
        member_count = UserTeam.query.filter_by(team_id=team.id).count()

        # 团队操作按钮
        action_buttons = ''
        if current_user.id == team.leader_id or current_user.role == 'admin':
            action_buttons = f'''
                <a href="/team/manage_members?team_id={team.id}" class="btn" style="padding:5px 10px; font-size:12px; margin-right:5px;">管理成员</a>
                <a href="/team/achievements?team_id={team.id}" class="btn" style="padding:5px 10px; font-size:12px; margin-right:5px;">成果统计</a>
                <a href="/team/member_achievements?team_id={team.id}" class="btn" style="padding:5px 10px; font-size:12px;">成员成果</a>
                '''
        elif current_user.role == 'teacher':
            action_buttons = '<span style="color:#7f8c8d;">普通成员（仅查看）</span>'

        # 团队卡片
        team_list_html += f'''
            <div style="border:1px solid #eee; border-radius:8px; padding:20px; background:white; box-shadow:0 2px 5px rgba(0,0,0,0.05);">
                <h4 style="margin-bottom:10px; color:#2c3e50;">{team.name}</h4>
                <p><strong>负责人：</strong>{leader_name}</p>
                <p><strong>创建时间：</strong>{team.create_time.strftime('%Y-%m-%d')}</p>
                <p><strong>成员数量：</strong>{member_count} 人</p>
                <div style="margin-top:15px;">{action_buttons}</div>
            </div>
            '''

    team_list_html += '''
            </div>
        </div>

        <script>
            // 显示/隐藏创建团队表单
            function toggleCreateForm() {
                const form = document.getElementById('createTeamForm');
                form.style.display = form.style.display === 'none' ? 'block' : 'none';
            }
        </script>
        '''

    return render_base_layout('团队列表', team_list_html, current_user)

# ========== 期刊论文智能导入路由 ==========
@app.route('/achievement/journal_paper/import', methods=['GET', 'POST'])
def journal_paper_import():
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('index'))

    zhipu_api_key = get_zhipu_api_key(current_user)
    if not zhipu_api_key:
        content = '''
        <div class="alert alert-danger">
            未配置智谱 AI API Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型 API 配置</a> 配置智谱 AI API Key。
        </div>
        <a href="/achievement/journal_paper" class="btn">返回列表</a>
        '''
        return render_base_layout('期刊论文智能导入', content, current_user)

    if request.method == 'POST':
        keyword = request.form.get('keyword', '').strip()
        max_papers = request.form.get('max_papers', 3, type=int)
        driver_path = request.form.get('driver_path',
                                       r'C:\Users\mtlxzmd\OneDrive\桌面\新建文件夹\毕设\msedgedriver.exe')

        if not keyword:
            flash('搜索关键词不能为空！', 'danger')
            return redirect('/achievement/journal_paper/import')

        flash('开始爬取知网数据，请稍候...', 'success')
        papers = crawl_cnki_journal(keyword, max_papers, driver_path)

        if not papers:
            content = '''
            <div class="alert alert-warning">未爬取到任何期刊论文数据！</div>
            <a href="/achievement/journal_paper/import" class="btn">重新导入</a>
            <a href="/achievement/journal_paper" class="btn">返回列表</a>
            '''
            return render_base_layout('期刊论文智能导入', content, current_user)

        success_count = 0
        for paper in papers:
            try:
                ai_result = ai_analyze_journal_full(paper['引用格式'], zhipu_api_key)

                publish_date = None
                if paper.get('发表日期'):
                    try:
                        publish_date = datetime.strptime(paper['发表日期'], '%Y-%m-%d').date()
                    except:
                        pass

                journal_paper = JournalPaper(
                    user_id=current_user.id,
                    title=paper['论文名称'],
                    authors=paper['论文作者'],
                    corresponding_authors=paper.get('通讯作者', ''),
                    journal_name=paper['期刊名称'],
                    inclusion_status=paper.get('论文收录情况', ''),
                    year=paper.get('年') or ai_result.get('年'),
                    volume=paper.get('卷') or ai_result.get('卷'),
                    issue=paper.get('期') or ai_result.get('期'),
                    page_range=paper.get('起止页码') or ai_result.get('起止页码'),
                    doi=paper.get('DOI') or ai_result.get('DOI'),
                    publish_year=paper.get('发表年份') or (ai_result.get('年') if ai_result.get('年') else None),
                    publish_date=publish_date,
                    create_time=datetime.now(),
                    update_time=datetime.now()
                )
                db.session.add(journal_paper)
                db.session.flush()

                auto_link_contributors(journal_paper, 'journal_paper', paper['论文作者'], current_user.id)

                success_count += 1
            except Exception as e:
                print(f"导入期刊论文失败：{e}")
                continue

        db.session.commit()
        content = f'''
        <div class="alert alert-success">
            智能导入完成！<br>
            共爬取 {len(papers)} 篇论文，成功导入 {success_count} 篇。
        </div>
        <a href="/achievement/journal_paper" class="btn">查看论文列表</a>
        <a href="/achievement/journal_paper/import" class="btn">继续导入</a>
        '''
        return render_base_layout('期刊论文智能导入', content, current_user)

    # GET请求：显示导入表单
    form_html = '''
    <h2>期刊论文智能导入（知网爬取）</h2>
    <form method="POST">
        <div class="form-group">
            <label>搜索作者 <span style="color:red;">*</span></label>
            <input type="text" name="keyword" placeholder="作者名" required>
        </div>
        <div class="form-group">
            <label>最大导入数量</label>
            <input type="number" name="max_papers" value="3" min="1" max="10">
        </div>
        <div class="form-group">
            <name="driver_path" value="C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedgedriver.exe" style="width:100%;">
        </div>
        <div class="form-group">
            <button type="submit" class="btn" style="background:#27ae60;">开始智能导入</button>
            <a href="/achievement/journal_paper" class="btn" style="background:#95a5a6; margin-left:10px;">取消</a>
        </div>
    </form>
    '''
    return render_base_layout('期刊论文智能导入', form_html, current_user)

# ========== 会议论文智能导入路由 ==========
@app.route('/achievement/conference_paper/import', methods=['GET', 'POST'])
def conference_paper_import():
    current_user = get_current_user()
    if not current_user or current_user.role == 'admin':
        return redirect(url_for('login'))

    zhipu_api_key = get_zhipu_api_key(current_user)
    if not zhipu_api_key:
        content = '''
        <div class="alert alert-danger">
            未配置智谱AI API Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型API配置</a> 配置智谱AI API Key。
        </div>
        <a href="/achievement/conference_paper" class="btn">返回列表</a>
        '''
        return render_base_layout('会议论文智能导入', content, current_user)

    if request.method == 'POST':
        keyword = request.form.get('keyword', '').strip()
        max_papers = request.form.get('max_papers', 3, type=int)
        driver_path = request.form.get('driver_path',
                                       r'C:\Users\mtlxzmd\OneDrive\桌面\新建文件夹\毕设\msedgedriver.exe')

        if not keyword:
            flash('搜索关键词不能为空！', 'danger')
            return redirect('/achievement/conference_paper/import')

        flash('开始爬取知网数据，请稍候...', 'success')
        papers = crawl_cnki_conference(keyword, max_papers, driver_path)

        if not papers:
            content = '''
            <div class="alert alert-warning">未爬取到任何会议论文数据！</div>
            <a href="/achievement/conference_paper/import" class="btn">重新导入</a>
            <a href="/achievement/conference_paper" class="btn">返回列表</a>
            '''
            return render_base_layout('会议论文智能导入', content, current_user)

        success_count = 0
        for paper in papers:
            try:
                ai_result = ai_analyze_citation(paper['引用格式'], zhipu_api_key)

                conference_start_date = None
                conference_end_date = None
                conference_time_str = None

                if paper.get('会议时间'):
                    time_text = paper['会议时间'].strip()
                    conference_time_str = time_text

                    import re
                    date_range_pattern = r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*[-–—]\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})'
                    match = re.search(date_range_pattern, time_text)

                    if match:
                        start_year, start_month, start_day = match.group(1), match.group(2), match.group(3)
                        end_year, end_month, end_day = match.group(4), match.group(5), match.group(6)
                        try:
                            conference_start_date = datetime(int(start_year), int(start_month), int(start_day)).date()
                            conference_end_date = datetime(int(end_year), int(end_month), int(end_day)).date()
                        except:
                            pass
                    else:
                        single_date_pattern = r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})'
                        match = re.search(single_date_pattern, time_text)
                        if match:
                            try:
                                year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
                                conference_start_date = datetime(year, month, day).date()
                                conference_end_date = conference_start_date
                            except:
                                pass

                conference_paper = ConferencePaper(
                    user_id=current_user.id,
                    title=paper['论文名称'],
                    authors=paper['论文作者'],
                    corresponding_authors=paper.get('通讯作者', ''),
                    conference_name=paper['会议名称'],
                    conference_time=conference_time_str,
                    conference_start_date=conference_start_date,
                    conference_end_date=conference_end_date,
                    conference_place=paper.get('会议地点'),
                    page_range=paper.get('起止页码') or ai_result.get('起止页码'),
                    doi=paper.get('DOI') or ai_result.get('DOI'),
                    publish_year=paper.get('发表年份') or (ai_result.get('年') if ai_result.get('年') else None),
                    create_time=datetime.now(),
                    update_time=datetime.now()
                )
                db.session.add(conference_paper)
                db.session.flush()

                auto_link_contributors(conference_paper, 'conference_paper', paper['论文作者'], current_user.id)

                success_count += 1
            except Exception as e:
                print(f"导入会议论文失败：{e}")
                continue

        db.session.commit()
        content = f'''
        <div class="alert alert-success">
            智能导入完成！<br>
            共爬取 {len(papers)} 篇论文，成功导入 {success_count} 篇。
        </div>
        <a href="/achievement/conference_paper" class="btn">查看论文列表</a>
        <a href="/achievement/conference_paper/import" class="btn">继续导入</a>
        '''
        return render_base_layout('会议论文智能导入', content, current_user)

    # GET请求：显示导入表单
    form_html = '''
    <h2>会议论文智能导入（知网爬取）</h2>
    <form method="POST">
        <div class="form-group">
            <label>搜索作者 <span style="color:red;">*</span></label>
            <input type="text" name="keyword" placeholder="作者名" required>
        </div>
        <div class="form-group">
            <label>最大导入数量</label>
            <input type="number" name="max_papers" value="3" min="1" max="10">
        </div>
        <div class="form-group">
            <name="driver_path" value="C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedgedriver.exe" style="width:100%;">
        </div>
        <div class="form-group">
            <button type="submit" class="btn" style="background:#27ae60;">开始智能导入</button>
            <a href="/achievement/conference_paper" class="btn" style="background:#95a5a6; margin-left:10px;">取消</a>
        </div>
    </form>
    '''
    return render_base_layout('会议论文智能导入', form_html, current_user)


@app.route('/achievement/ocr_import', methods=['GET', 'POST'])
def ocr_import():
    """OCR图片识别导入成果（新增PDF支持）"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 检查百度API配置（OCR必需）
    api_config = current_user.get_api_config()
    if not api_config.get('baidu', {}).get('api_key') or not api_config.get('baidu', {}).get('secret_key'):
        content = '''
        <div class="alert alert-danger">
            未配置百度API Key/Secret Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型API配置</a> 配置百度API。
        </div>
        <a href="/" class="btn">返回首页</a>
        '''
        return render_base_layout('OCR智能导入', content, current_user)

    # 检查智谱API配置（AI分析可选）
    zhipu_configured = bool(get_zhipu_api_key(current_user))

    # 初始化success变量（关键修复：提前定义）
    success = False
    ocr_text = ""
    temp_images = []
    achievement_info = {}

    if request.method == 'POST':
        # 处理文件上传
        if 'image_file' not in request.files:
            flash('请上传图片/PDF文件！', 'danger')
            return redirect(request.url)

        file = request.files['image_file']
        if file.filename == '':
            flash('请选择图片/PDF文件！', 'danger')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            # 保存上传文件到临时目录
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_temp')
            if not os.path.exists(temp_path):
                os.makedirs(temp_path)

            # 先从原始文件名提取扩展名，避免 secure_filename 过滤掉中文字符后丢失信息
            original_filename = file.filename

            # 检查是否有扩展名
            if '.' not in original_filename:
                flash('文件格式错误，请重新上传！', 'danger')
                return redirect(request.url)

            # 提取扩展名
            file_ext = original_filename.rsplit('.', 1)[1].lower()

            # 生成安全文件名（使用时间戳 + 原扩展名）
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"{timestamp}.{file_ext}"

            upload_path = os.path.join(temp_path, filename)
            file.save(upload_path)

            try:
                # ========== 新增PDF处理逻辑 ==========
                if file_ext == 'pdf':
                    # PDF文件：先转图片，再逐页OCR
                    flash('检测到PDF文件，正在转换为图片...', 'info')
                    try:
                        temp_images = pdf_to_images(upload_path, temp_path)
                    except Exception as e:
                        # 专门处理poppler错误
                        flash(f'PDF转换失败：{str(e)}<br>请安装poppler并配置路径！', 'danger')
                        # 清理临时文件
                        if os.path.exists(upload_path):
                            os.remove(upload_path)
                        content = f'''
                        <a href="/achievement/ocr_import" class="btn">重新上传</a>
                        '''
                        return render_base_layout('OCR智能导入', content, current_user)

                    # 逐页识别PDF转换后的图片
                    for idx, img_path in enumerate(temp_images):
                        page_text, err = baidu_ocr_recognize(img_path, current_user)
                        if err:
                            flash(f'第{idx + 1}页识别失败：{err}', 'warning')
                            continue
                        ocr_text += f"\n=== 第{idx + 1}页 ===\n{page_text}"
                else:
                    # 图片文件：直接OCR识别
                    ocr_text, err = baidu_ocr_recognize(upload_path, current_user)
                    if err:
                        flash(f'OCR识别失败：{err}', 'danger')
                        # 清理临时文件
                        if os.path.exists(upload_path):
                            os.remove(upload_path)
                        return redirect(request.url)

                if not ocr_text.strip():
                    flash('未识别到任何文本！', 'warning')
                    # 清理临时文件
                    if os.path.exists(upload_path):
                        os.remove(upload_path)
                    return redirect(request.url)

                # ========== 原有AI分析逻辑保持不变 ==========
                # AI分析（如果配置了智谱API）
                ai_info = {}
                if zhipu_configured:
                    flash('正在使用AI分析识别结果，请稍候...', 'info')
                    ai_info = ai_analyze_achievement_text(ocr_text, get_zhipu_api_key(current_user))
                    achievement_info = {
                        'type_name': ai_info.get('type_name'),
                        'title': ai_info.get('title'),
                        'raw_text': ocr_text,
                        'confidence': ai_info.get('confidence'),
                        'ai_data': ai_info
                    }
                else:
                    # 兜底：基础解析
                    achievement_info = extract_achievement_info(ocr_text)
                    achievement_info['raw_text'] = ocr_text

                # 创建成果记录
                success, msg, type_name, achievement_id = create_achievement_from_ocr(achievement_info, current_user)
                flash(msg, 'success' if success else 'danger')


            except Exception as e:
                # 捕获所有异常并友好提示
                success = False
                msg = f'处理失败：{str(e)}'
                logger.error(f"OCR导入处理异常：{str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                flash(msg, 'danger')
            finally:
                # ========== 清理临时文件 ==========
                # 删除上传的源文件
                if os.path.exists(upload_path):
                    os.remove(upload_path)

                # 删除 PDF 转换的临时图片
                for img_path in temp_images:
                    if os.path.exists(img_path):
                        os.remove(img_path)
                # 反馈结果
                if success:
                    # 映射成果类型到对应路由
                    type_route_mapping = {
                        '期刊论文': '/achievement/journal_paper',
                        '会议论文': '/achievement/conference_paper',
                        '教材': '/achievement/textbook',
                        '专著': '/achievement/monograph',
                        '发明专利': '/achievement/patent',
                        '实用新型专利': '/achievement/patent',
                        '软著': '/achievement/software_copyright',
                        '教学成果获奖': '/achievement/teaching_achievement_award',
                        '教学竞赛获奖': '/achievement/teaching_competition_award',
                        '指导学生获奖': '/achievement/student_guidance_award',
                        '教研教改和课程建设项目': '/achievement/teaching_project'
                    }
                    target_route = type_route_mapping.get(type_name, '/')

                    # 显示 AI 分析详细信息
                    ai_feedback = ""
                    if zhipu_configured and ai_info:

                        # 生成全量字段展示
                        field_html = ""

                        # 根据成果类型筛选展示字段
                        type_fields = {
                            '期刊论文': [
                                ('论文名称', 'title'),
                                ('作者', 'authors'),
                                ('通讯作者', 'corresponding_authors'),
                                ('期刊名称', 'journal_name'),
                                ('收录情况', 'inclusion_status'),
                                ('年', 'year'),
                                ('卷', 'volume'),
                                ('期', 'issue'),
                                ('起止页码', 'page_range'),
                                ('DOI', 'doi'),
                                ('发表年份', 'publish_year'),
                                ('发表日期', 'publish_date'),
                            ],
                            '会议论文': [
                                ('论文名称', 'title'),
                                ('作者', 'authors'),
                                ('通讯作者', 'corresponding_authors'),
                                ('会议名称', 'conference_name'),
                                ('会议时间', 'conference_time'),
                                ('会议地点', 'conference_place'),
                                ('起止页码', 'page_range'),
                                ('DOI', 'doi'),
                                ('发表年份', 'publish_year'),
                            ],
                            '教材': [
                                ('教材名称', 'title'),
                                ('教材系列', 'textbook_series'),
                                ('主编', 'chief_editor'),
                                ('副主编', 'associate_editors'),
                                ('编委', 'editorial_board'),
                                ('出版社', 'publisher'),
                                ('ISBN', 'isbn'),
                                ('CIP核字号', 'cip_number'),
                                ('出版年份', 'publication_year'),
                                ('出版月份', 'publication_month'),
                                ('版次', 'edition'),
                                ('字数', 'word_count'),
                                ('定价', 'price'),
                                ('教材级别', 'textbook_level'),
                                ('教材类型', 'textbook_type'),
                                ('适用专业', 'applicable_majors'),
                                ('备注', 'remarks'),
                            ],
                            '专利': [
                                ('专利名称', 'title'),
                                ('专利类型', 'patent_type'),
                                ('专利号', 'patent_number'),
                                ('申请日期', 'apply_date'),
                                ('授权日期', 'grant_date'),
                                ('状态', 'status'),
                            ],
                            '软著': [
                                ('软件名称', 'title'),
                                ('著作权人', 'copyright_owner'),
                                ('开发完成日期', 'completion_date'),
                                ('首次发表日期', 'first_publication_date'),
                                ('权利取得方式', 'right_acquisition_method'),
                                ('权利范围', 'right_scope'),
                                ('登记号', 'copyright_number'),
                                ('证书号', 'certificate_number'),
                                ('登记日期', 'register_date'),
                            ],
                            '获奖类': [
                                ('获奖名称', 'title'),
                                ('获奖等级', 'award_level'),
                                ('获奖等次', 'award_rank'),
                                ('获奖日期', 'award_date'),
                                ('主办方/竞赛名称', 'competition_name'),
                                ('获奖学生', 'student_name'),
                            ]
                        }
                        # 匹配当前成果类型的展示字段
                        current_fields = []
                        if ai_info['type_name'] in ['发明专利', '实用新型专利']:
                            current_fields = type_fields['专利']
                        elif '获奖' in ai_info['type_name']:
                            current_fields = type_fields['获奖类']
                        else:
                            current_fields = type_fields.get(ai_info['type_name'], [])

                        # 生成字段展示 HTML
                        for label, key in current_fields:
                            value = ai_info.get(key, '未识别')
                            if value:
                                field_html += f"<p><strong>{label}：</strong>{value}</p>"

                        ai_feedback = f'''
                                    <div class="alert alert-info">
                                        <h4>AI 智能提取的全量字段（置信度：{ai_info.get('confidence', 0):.2f}）</h4>
                                        {field_html if field_html else "<p>未提取到详细字段</p>"}
                                    </div>
                                    '''
                    content = f'''
                                <div class="alert alert-success">
                                    <h4>操作成功！</h4>
                                    <p>{msg}</p>
                                </div>
                                {ai_feedback}
                                <a href="{target_route}" class="btn">查看成果列表</a>
                                <a href="/achievement/ocr_import" class="btn">继续识别</a>
                                '''
                else:
                    # 确保 msg 变量一定存在
                    error_message = locals().get('msg', '处理过程中发生未知错误，请查看日志获取详细信息')
                    content = f'''
                                <div class="alert alert-danger">
                                    <h4>操作失败</h4>
                                    <p>{error_message}</p>
                                </div>
                                <div class="alert alert-info">
                                    <h4>OCR识别结果（仅展示，不存储）</h4>
                                    <pre style="margin:10px 0; padding:10px; background:#f5f7fa; border-radius:4px; max-height:400px; overflow:auto;">{ocr_text if ocr_text else '无识别内容'}</pre>
                                    {f"<p>识别成果类型：<strong>{achievement_info.get('type_name', '未知')}</strong>（置信度：{achievement_info.get('confidence', 0)}）</p>" if achievement_info else ""}
                                    {f"<p>识别标题：<strong>{achievement_info.get('title', '未识别')}</strong></p>" if achievement_info else ""}
                                </div>
                                <a href="/achievement/ocr_import" class="btn">重新识别</a>
                                <a href="/" class="btn">返回首页</a>

                                '''

                return render_base_layout('OCR智能导入', content, current_user)
    # GET请求：显示上传表单（更新提示文本）
    ai_tip = ""
    if not zhipu_configured:
        ai_tip = '''
        <div class="alert alert-warning">
            未配置智谱AI API Key，将使用基础OCR识别（无AI智能分析）<br>
            配置地址：<a href="/user/api_config">个人设置 > 大模型API配置</a>
        </div>
        '''

    form_html = f'''
        <h2>OCR智能导入成果（支持图片/PDF）</h2>
        {ai_tip}
        <form method="POST" enctype="multipart/form-data">
            <div class="form-group">
                <label>上传成果图片/PDF <span style="color:red;">*</span></label>
                <input type="file" name="image_file" accept="image/*,.pdf" required>
                <!-- 新增：文件大小提示 -->
                <p style="margin-top:5px; color:#666;">
                    支持格式：JPG/PNG/GIF/PDF，PDF文件会自动转换为图片逐页识别<br>
                    <strong>文件大小限制：100MB</strong>，过大的PDF建议先拆分后上传
                </p>
            </div>
            <button type="submit" class="btn" style="background:#27ae60;">开始识别并导入</button>
            <a href="/" class="btn" style="background:#95a5a6; margin-left:10px;">取消</a>
        </form>
        '''
    return render_base_layout('OCR智能导入', form_html, current_user)


@app.route('/achievement/voice_export', methods=['GET', 'POST'])
def voice_export():
    """语音导出成果（支持手动修改识别文字）"""
    current_user = get_current_user()
    if not current_user:
        return redirect(url_for('login'))

    # 检查百度API配置（语音识别必需）
    api_config = current_user.get_api_config()
    if not api_config.get('baidu', {}).get('api_key') or not api_config.get('baidu', {}).get('secret_key'):
        content = '''
        <div class="alert alert-danger">
            未配置百度API Key/Secret Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型API配置</a> 配置百度语音识别API。
        </div>
        <a href="/" class="btn">返回首页</a>
        '''
        return render_base_layout('语音导出成果', content, current_user)

    # 处理音频数据或手动修改后的文字提交
    if request.method == 'POST':
        try:
            # 情况1：上传录音文件（语音识别）
            if 'audio_blob' in request.files:
                audio_data = request.files['audio_blob'].read()
                # 音频转文字
                voice_text, err = audio_to_text(audio_data, current_user)
                if err:
                    return json.dumps({'status': 'error', 'msg': f'语音识别失败：{err}'})

                return json.dumps({
                    'status': 'success',
                    'voice_text': voice_text,
                    'export_url': '',
                    'msg': '语音识别完成，可手动修改后导出'
                })

            # 情况2：手动修改文字后提交导出
            elif 'voice_text' in request.form:
                voice_text = request.form.get('voice_text', '').strip()
                if not voice_text:
                    return json.dumps({'status': 'error', 'msg': '请输入导出指令'})

                # 解析语音指令
                cmd = parse_voice_command(voice_text)

                # 根据指令生成导出链接（核心修改：添加时间参数）
                export_url = ''
                export_msg = ''
                type_route_mapping = {
                    '期刊论文': '/achievement/journal_paper',
                    '会议论文': '/achievement/conference_paper',
                    '教材': '/achievement/textbook',
                    '专著': '/achievement/monograph',
                    '专利': '/achievement/patent',
                    '软著': '/achievement/software_copyright',
                    '教学成果获奖': '/achievement/teaching_achievement_award',
                    '教学竞赛获奖': '/achievement/teaching_competition_award',
                    '指导学生获奖': '/achievement/student_guidance_award'
                }

                if cmd['action'] == 'export' and cmd['type_name']:
                    base_url = type_route_mapping.get(cmd['type_name'], '')
                    if base_url:
                        # 拼接时间参数
                        export_url = f"{base_url}?action=export"
                        if cmd['start_date']:
                            export_url += f"&start_date={cmd['start_date']}"
                        if cmd['end_date']:
                            export_url += f"&end_date={cmd['end_date']}"
                        export_msg = f'已识别指令：导出{cmd["start_date"] if cmd["start_date"] else ""}{cmd["end_date"] if cmd["end_date"] else ""}的{cmd["type_name"]}'
                    else:
                        export_msg = f'暂不支持导出{cmd["type_name"]}类型成果'
                else:
                    export_msg = f'未识别有效导出指令，识别文本：{voice_text}'

                return json.dumps({
                    'status': 'success',
                    'voice_text': voice_text,
                    'export_url': export_url,
                    'msg': export_msg
                })

        except Exception as e:
            return json.dumps({'status': 'error', 'msg': f'处理失败：{str(e)}'})

    # 渲染语音导出页面（增加手动修改功能）
    # 原有HTML代码保持不变...
    form_html = '''
    <h2>语音导出成果</h2>
    <div class="alert alert-info">
        支持语音指令示例：<br>
        - 导出2024年的期刊论文<br>
        - 导出团队的教学竞赛获奖<br>
        - 导出我的所有教材
    </div>

    <!-- 录音区域 -->
    <div style="margin:20px 0; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h4>网页录音导出</h4>
        <button id="recordBtn" class="btn" style="background:#27ae60;">开始录音</button>
        <button id="stopBtn" class="btn" style="background:#e74c3c; display:none;">停止录音</button>
        <div id="recordStatus" style="margin-top:10px; color:#666;"></div>

        <!-- 识别结果展示 + 手动修改 -->
        <div id="resultArea" style="margin-top:20px; display:none;">
            <div class="alert alert-info">
                <h5>语音识别结果（可手动修改）：</h5>
                <textarea id="voiceTextInput" style="width:100%; height:100px; margin:10px 0; padding:10px;" placeholder="请输入导出指令..."></textarea>
                <button id="submitTextBtn" class="btn">确认导出</button>
            </div>
            <div id="exportArea"></div>
        </div>
    </div>

    <script>
        let recorder = null;
        let audioBlob = null;

        // 开始录音
        document.getElementById('recordBtn').addEventListener('click', async () => {
            try {
                // 获取麦克风权限
                const stream = await navigator.mediaDevices.getUserMedia({ audio: true });
                recorder = new MediaRecorder(stream);
                const chunks = [];

                // 收集录音数据
                recorder.ondataavailable = (e) => chunks.push(e.data);

                // 录音停止后处理
                recorder.onstop = async () => {
                    audioBlob = new Blob(chunks, { type: 'audio/webm' });
                    document.getElementById('recordStatus').textContent = '录音完成，正在识别...';

                    // 创建FormData并提交录音数据
                    const formData = new FormData();
                    formData.append('audio_blob', audioBlob, 'record.webm');

                    // 发送请求识别语音
                    const response = await fetch('/achievement/voice_export', {
                        method: 'POST',
                        body: formData
                    });

                    const result = await response.json();
                    document.getElementById('recordStatus').textContent = '';

                    // 展示结果（可修改）
                    document.getElementById('resultArea').style.display = 'block';
                    if (result.status === 'success') {
                        // 填充识别结果到文本框
                        document.getElementById('voiceTextInput').value = result.voice_text;
                        document.getElementById('exportArea').innerHTML = 
                            `<div class="alert alert-success">${result.msg}</div>`;
                    } else {
                        document.getElementById('exportArea').innerHTML = 
                            `<div class="alert alert-danger">${result.msg}</div>`;
                    }

                    // 停止所有音轨
                    stream.getTracks().forEach(track => track.stop());
                };

                // 开始录音
                recorder.start();
                document.getElementById('recordBtn').style.display = 'none';
                document.getElementById('stopBtn').style.display = 'inline-block';
                document.getElementById('recordStatus').textContent = '正在录音...（点击停止按钮结束）';

            } catch (err) {
                document.getElementById('recordStatus').textContent = `录音权限获取失败：${err.message}`;
            }
        });

        // 停止录音
        document.getElementById('stopBtn').addEventListener('click', () => {
            if (recorder && recorder.state === 'recording') {
                recorder.stop();
                document.getElementById('recordBtn').style.display = 'inline-block';
                document.getElementById('stopBtn').style.display = 'none';
            }
        });

        // 手动修改文字后提交导出
        document.getElementById('submitTextBtn').addEventListener('click', async () => {
            const voiceText = document.getElementById('voiceTextInput').value.trim();
            if (!voiceText) {
                alert('请输入导出指令');
                return;
            }

            document.getElementById('exportArea').innerHTML = '<div class="alert alert-info">正在解析导出指令...</div>';

            // 提交手动修改后的文字
            const formData = new FormData();
            formData.append('voice_text', voiceText);

            const response = await fetch('/achievement/voice_export', {
                method: 'POST',
                body: formData
            });

            const result = await response.json();
            if (result.status === 'success') {
                // 显示导出链接
                let exportHtml = '';
                if (result.export_url) {
                    exportHtml = `
                    <div class="alert alert-success">
                        <p>${result.msg}</p>
                        <a href="${result.export_url}" class="btn">点击导出成果</a>
                    </div>
                    `;
                } else {
                    exportHtml = `<div class="alert alert-warning">${result.msg}</div>`;
                }
                document.getElementById('exportArea').innerHTML = exportHtml;
            } else {
                document.getElementById('exportArea').innerHTML = 
                    `<div class="alert alert-danger">${result.msg}</div>`;
            }
        });
    </script>
    '''
    return render_base_layout('语音导出成果', form_html, current_user)

@app.route('/team/voice_export', methods=['GET', 'POST'])
def team_voice_export():
    """团队负责人专属：语音导出团队公开成果（支持指定老师+成果类型）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('仅团队负责人可使用此功能！', 'danger')
        return redirect(url_for('index'))

    api_config = current_user.get_api_config()
    if not api_config.get('baidu', {}).get('api_key') or not api_config.get('baidu', {}).get('secret_key'):
        content = '''
        <div class="alert alert-danger">
            未配置百度API Key/Secret Key！<br>
            请先前往 <a href="/user/api_config">个人设置 > 大模型API配置</a> 配置百度语音识别API。
        </div>
        <a href="/team/list" class="btn">返回团队列表</a>
        '''
        return render_base_layout('团队语音导出成果', content, current_user)

    managed_teams = Team.query.filter_by(leader_id=current_user.id).all()
    if not managed_teams:
        content = '''
        <div class="alert alert-warning">
            您尚未管理任何团队！
        </div>
        <a href="/team/list" class="btn">创建团队</a>
        '''
        return render_base_layout('团队语音导出成果', content, current_user)
    current_team = managed_teams[0]
    team_id_str = str(current_team.id)

    if request.method == 'POST':
        try:
            if 'audio_blob' in request.files:
                audio_data = request.files['audio_blob'].read()
                voice_text, err = audio_to_text(audio_data, current_user)
                if err:
                    return json.dumps({'status': 'error', 'msg': f'语音识别失败：{err}'})

                return json.dumps({
                    'status': 'success',
                    'voice_text': voice_text,
                    'export_url': '',
                    'msg': '语音识别完成，可手动修改后导出'
                })

            elif 'voice_text' in request.form:
                voice_text = request.form.get('voice_text', '').strip()
                if not voice_text:
                    return json.dumps({'status': 'error', 'msg': '请输入导出指令'})

                cmd = parse_voice_command(voice_text)

                export_url = ''
                export_msg = ''
                error_msg = ''

                target_teacher = None
                if cmd.get('teacher_name'):
                    teacher_name = cmd['teacher_name']
                    target_teacher = User.query.filter(
                        or_(
                            User.username.like(f'%{teacher_name}%'),
                            User.employee_id.like(f'%{teacher_name}%')
                        ),
                        User.id.in_([ut.user_id for ut in UserTeam.query.filter_by(team_id=current_team.id).all()])
                    ).first()

                    if not target_teacher:
                        export_msg = f'未找到团队成员：{teacher_name}'
                        return json.dumps({
                            'status': 'warning',
                            'voice_text': voice_text,
                            'export_url': '',
                            'msg': export_msg
                        })

                type_mapping = {
                    '期刊论文': (JournalPaper, 'journal_paper'),
                    '会议论文': (ConferencePaper, 'conference_paper'),
                    '教材': (Textbook, 'textbook'),
                    '专著': (Monograph, 'monograph'),
                    '专利': (Patent, 'patent'),
                    '软著': (SoftwareCopyright, 'software_copyright'),
                    '教学成果获奖': (TeachingAchievementAward, 'teaching_achievement_award'),
                    '教学竞赛获奖': (TeachingCompetitionAward, 'teaching_competition_award'),
                    '指导学生获奖': (StudentGuidanceAward, 'student_guidance_award'),
                    '教研项目': (TeachingProject, 'teaching_project')
                }

                if cmd['action'] == 'export' and cmd['type_name']:
                    if cmd['type_name'] not in type_mapping:
                        export_msg = f'暂不支持导出{cmd["type_name"]}类型成果'
                    else:
                        model, type_key = type_mapping[cmd['type_name']]
                        export_url = f"/team/export_specified?team_id={current_team.id}"
                        export_url += f"&type={type_key}"

                        if target_teacher:
                            export_url += f"&teacher_id={target_teacher.id}"
                            teacher_info = f"{target_teacher.username}（{target_teacher.employee_id}）"
                        else:
                            teacher_info = "所有成员"

                        if cmd['start_date']:
                            export_url += f"&start_date={cmd['start_date']}"
                        if cmd['end_date']:
                            export_url += f"&end_date={cmd['end_date']}"

                        time_info = ""
                        if cmd['start_date'] and cmd['end_date']:
                            time_info = f"{cmd['start_date'][:4]}-{cmd['end_date'][:4]}年"
                        elif cmd['start_date']:
                            time_info = f"{cmd['start_date'][:4]}年"

                        export_msg = f'已识别指令：导出{time_info}{teacher_info}的{cmd["type_name"]}（仅公开给{current_team.name}的成果）'

                else:
                    export_msg = f'未识别有效导出指令，示例：导出张三老师的2024年期刊论文'

                return json.dumps({
                    'status': 'success',
                    'voice_text': voice_text,
                    'export_url': export_url,
                    'msg': export_msg
                })

        except Exception as e:
            return json.dumps({'status': 'error', 'msg': f'处理失败：{str(e)}'})

    # 修复：修正JavaScript模板字符串语法，移除多余的$符号
    form_html = f'''
<h2>团队语音导出成果（{current_team.name}）</h2>
<div class="alert alert-info">
    支持语音指令示例：<br>
    - 导出张三老师的2024年期刊论文<br>
    - 导出李四老师的专利<br>
    - 导出所有成员的教学竞赛获奖<br>
    - 导出王五老师的近三年教研项目
</div>

<div style="margin:20px 0; padding:20px; border:1px solid #eee; border-radius:8px;">
    <h4>网页录音导出</h4>
    <button id="recordBtn" class="btn" style="background:#27ae60;">开始录音</button>
    <button id="stopBtn" class="btn" style="background:#e74c3c; display:none;">停止录音</button>
    <div id="recordStatus" style="margin-top:10px; color:#666;"></div>

    <div id="resultArea" style="margin-top:20px; display:none;">
        <div class="alert alert-info">
            <h5>语音识别结果（可手动修改）：</h5>
            <textarea id="voiceTextInput" style="width:100%; height:100px; margin:10px 0; padding:10px;" placeholder="请输入导出指令..."></textarea>
            <button id="submitTextBtn" class="btn">确认导出</button>
        </div>
        <div id="exportArea"></div>
    </div>
</div>

<script>
    let recorder = null;
    let audioBlob = null;

    document.getElementById('recordBtn').addEventListener('click', async () => {{
        try {{
            const stream = await navigator.mediaDevices.getUserMedia({{ audio: true }});
            recorder = new MediaRecorder(stream);
            const chunks = [];

            recorder.ondataavailable = (e) => chunks.push(e.data);

            recorder.onstop = async () => {{
                audioBlob = new Blob(chunks, {{ type: 'audio/webm' }});
                document.getElementById('recordStatus').textContent = '录音完成，正在识别...';

                const formData = new FormData();
                formData.append('audio_blob', audioBlob, 'record.webm');

                const response = await fetch('/team/voice_export', {{
                    method: 'POST',
                    body: formData
                }});

                const result = await response.json();
                document.getElementById('recordStatus').textContent = '';

                document.getElementById('resultArea').style.display = 'block';
                if (result.status === 'success') {{
                    document.getElementById('voiceTextInput').value = result.voice_text;
                    document.getElementById('exportArea').innerHTML =
                        `<div class="alert alert-success">${{result.msg}}</div>`;
                }} else {{
                    document.getElementById('exportArea').innerHTML =
                        `<div class="alert alert-danger">${{result.msg}}</div>`;
                }}

                stream.getTracks().forEach(track => track.stop());
            }};

            recorder.start();
            document.getElementById('recordBtn').style.display = 'none';
            document.getElementById('stopBtn').style.display = 'inline-block';
            document.getElementById('recordStatus').textContent = '正在录音...（点击停止按钮结束）';

        }} catch (err) {{
            document.getElementById('recordStatus').textContent = `录音权限获取失败：${{err.message}}`;
        }}
    }});

    document.getElementById('stopBtn').addEventListener('click', () => {{
        if (recorder && recorder.state === 'recording') {{
            recorder.stop();
            document.getElementById('recordBtn').style.display = 'inline-block';
            document.getElementById('stopBtn').style.display = 'none';
        }}
    }});

    document.getElementById('submitTextBtn').addEventListener('click', async () => {{
        const voiceText = document.getElementById('voiceTextInput').value.trim();
        if (!voiceText) {{
            alert('请输入导出指令');
            return;
        }}

        document.getElementById('exportArea').innerHTML = '<div class="alert alert-info">正在解析导出指令...</div>';

        const formData = new FormData();
        formData.append('voice_text', voiceText);

        const response = await fetch('/team/voice_export', {{
            method: 'POST',
            body: formData
        }});

        const result = await response.json();
        if (result.status === 'success') {{
            let exportHtml = '';
            if (result.export_url) {{
                exportHtml = `
<div class="alert alert-success">
<p>${{result.msg}}</p>
<a href="${{result.export_url}}" class="btn">点击导出成果</a>
</div>`;
            }} else {{
                exportHtml = `<div class="alert alert-warning">${{result.msg}}</div>`;
            }}
            document.getElementById('exportArea').innerHTML = exportHtml;
        }} else if (result.status === 'warning') {{
            document.getElementById('exportArea').innerHTML =
                `<div class="alert alert-warning">${{result.msg}}</div>`;
        }} else {{
            document.getElementById('exportArea').innerHTML =
                `<div class="alert alert-danger">${{result.msg}}</div>`;
        }}
    }});
</script>
'''

    return render_base_layout(f'团队语音导出成果 - {current_team.name}', form_html, current_user)

@app.route('/team/export_specified')
def team_export_specified():
    """导出团队指定老师的指定类型公开成果"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'team_leader':
        flash('仅团队负责人可使用此功能！', 'danger')
        return redirect(url_for('index'))

    team_id = request.args.get('team_id', type=int)
    type_key = request.args.get('type')
    teacher_id = request.args.get('teacher_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not team_id or not type_key:
        flash('导出参数缺失！', 'danger')
        return redirect(url_for('team_voice_export'))

    team = db.session.get(Team, team_id)
    if not team or team.leader_id != current_user.id:
        flash('无权限导出该团队成果！', 'danger')
        return redirect(url_for('team_voice_export'))
    team_id_str = str(team_id)

    type_model_mapping = {
        'journal_paper': (JournalPaper, '期刊论文', [
            {'name': 'title', 'label': '论文名称'},
            {'name': 'authors', 'label': '论文作者'},
            {'name': 'journal_name', 'label': '期刊名称'},
            {'name': 'publish_year', 'label': '发表年份'},
            {'name': 'publish_date', 'label': '发表日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'conference_paper': (ConferencePaper, '会议论文', [
            {'name': 'title', 'label': '论文名称'},
            {'name': 'authors', 'label': '论文作者'},
            {'name': 'conference_name', 'label': '会议名称'},
            {'name': 'publish_year', 'label': '发表年份'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'textbook': (Textbook, '教材', [
            {'name': 'title', 'label': '教材名称'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'publish_date', 'label': '出版日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'monograph': (Monograph, '专著', [
            {'name': 'title', 'label': '专著名称'},
            {'name': 'isbn', 'label': 'ISBN'},
            {'name': 'publisher', 'label': '出版社'},
            {'name': 'publish_date', 'label': '出版日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'patent': (Patent, '专利', [
            {'name': 'title', 'label': '专利名称'},
            {'name': 'patent_type', 'label': '专利类型'},
            {'name': 'patent_number', 'label': '专利号'},
            {'name': 'apply_date', 'label': '申请日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'software_copyright': (SoftwareCopyright, '软著', [
            {'name': 'title', 'label': '软件名称'},
            {'name': 'copyright_owner', 'label': '著作权人'},
            {'name': 'completion_date', 'label': '开发完成日期'},
            {'name': 'first_publication_date', 'label': '首次发表日期'},
            {'name': 'right_acquisition_method', 'label': '权利取得方式'},
            {'name': 'right_scope', 'label': '权利范围'},
            {'name': 'copyright_number', 'label': '登记号'},
            {'name': 'certificate_number', 'label': '证书号'},
            {'name': 'register_date', 'label': '登记日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'teaching_achievement_award': (TeachingAchievementAward, '教学成果获奖', [
            {'name': 'title', 'label': '获奖名称'},
            {'name': 'award_level', 'label': '获奖等级'},
            {'name': 'award_date', 'label': '获奖日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'teaching_competition_award': (TeachingCompetitionAward, '教学竞赛获奖', [
            {'name': 'title', 'label': '竞赛名称'},
            {'name': 'award_level', 'label': '获奖等级'},
            {'name': 'award_date', 'label': '获奖日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'student_guidance_award': (StudentGuidanceAward, '指导学生获奖', [
            {'name': 'title', 'label': '获奖名称'},
            {'name': 'student_name', 'label': '获奖学生'},
            {'name': 'award_date', 'label': '获奖日期'},
            {'name': 'attachment', 'label': '附件'}
        ]),
        'teaching_project': (TeachingProject, '教研项目', [
            {'name': 'title', 'label': '项目名称'},
            {'name': 'project_type', 'label': '项目类型'},
            {'name': 'start_date', 'label': '开始时间'},
            {'name': 'end_date', 'label': '结束时间'},
            {'name': 'attachment', 'label': '附件'}
        ])
    }

    if type_key not in type_model_mapping:
        flash('不支持的成果类型！', 'danger')
        return redirect(url_for('team_voice_export'))

    model, type_name, fields_config = type_model_mapping[type_key]

    query = model.query.filter(
        func.instr(func.concat(',', model.public_team_ids, ','), func.concat(',', team_id_str, ',')) > 0
    )

    if teacher_id:
        is_team_member = UserTeam.query.filter_by(team_id=team_id, user_id=teacher_id).first()
        if not is_team_member:
            flash('该老师不属于本团队！', 'danger')
            return redirect(url_for('team_voice_export'))
        query = query.filter(model.user_id == teacher_id)

    date_field_map = {
        JournalPaper: 'publish_date',
        ConferencePaper: 'conference_time',
        Textbook: 'publish_date',
        Monograph: 'publish_date',
        TeachingProject: 'start_date',
        Patent: 'apply_date',
        SoftwareCopyright: 'register_date',
        TeachingAchievementAward: 'award_date',
        TeachingCompetitionAward: 'award_date',
        StudentGuidanceAward: 'award_date'
    }
    date_field = date_field_map.get(model)
    if date_field:
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(getattr(model, date_field) >= start_date_obj)
            except:
                pass
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(getattr(model, date_field) <= end_date_obj)
            except:
                pass

    items = query.all()
    if not items:
        flash('暂无符合条件的公开成果！', 'warning')
        return redirect(url_for('team_voice_export'))

    wb = openpyxl.Workbook()
    ws = wb.active

    teacher_name = "所有成员"
    if teacher_id:
        teacher = db.session.get(User, teacher_id)
        teacher_name = teacher.username if teacher else "未知老师"
    ws.title = f"{team.name}-{teacher_name}-{type_name}"

    headers = [f['label'] for f in fields_config]
    ws.append(headers)

    for item in items:
        row = []
        for field in fields_config:
            field_name = field['name']
            value = getattr(item, field_name, '')

            if value is None:
                value = ''
            elif isinstance(value, (date, datetime)):
                value = value.strftime('%Y-%m-%d') if value else ''
            elif field_name == 'attachment' and value:
                value = os.path.basename(value) if value else ''
            row.append(value)
        ws.append(row)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{team.name}_{teacher_name}_{type_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@app.errorhandler(413)
def request_entity_too_large(error):
    user = get_current_user()
    content = '''
    <div class="alert alert-danger">
        <h4>上传失败！</h4>
        <p>文件大小超过限制（当前限制：100MB），请压缩或拆分文件后重新上传。</p>
    </div>
    <a href="/achievement/ocr_import" class="btn">重新上传</a>
    '''
    return render_base_layout('上传失败', content, user), 413


# ---------------------- 使用示例：添加期刊论文并关联多个作者 ----------------------

@app.route('/achievement/journal_paper/add_author', methods=['POST'])
def add_journal_paper_with_authors():
    """添加期刊论文并关联多个系统用户作者"""
    current_user = get_current_user()

    # 获取表单数据
    title = request.form.get('title')
    authors_str = request.form.get('authors')  # 逗号分隔的作者名字符串
    corresponding_authors_str = request.form.get('corresponding_authors')
    journal_name = request.form.get('journal_name')
    year = request.form.get('year', type=int)

    # 创建论文记录
    paper = JournalPaper(
        user_id=current_user.id,
        title=title,
        authors=authors_str,
        corresponding_authors=corresponding_authors_str,
        journal_name=journal_name,
        year=year
    )
    db.session.add(paper)
    db.session.flush()  # 获取 paper.id

    # 解析作者列表并关联系统用户
    import re
    author_names = [name.strip() for name in re.split(r'[;,;,,]', authors_str)]
    corresponding_names = [name.strip() for name in
                           re.split(r'[;,;,,]', corresponding_authors_str or '')] if corresponding_authors_str else []

    db.session.commit()
    flash('期刊论文添加成功！', 'success')
    return redirect(url_for('journal_paper_list'))


# ---------------------- 查询示例：获取某用户参与的所有期刊论文 ----------------------

def get_user_journal_papers(user_id):
    """获取用户参与的所有期刊论文（包含作为作者和关联人）"""
    owned_papers = JournalPaper.query.filter_by(user_id=user_id).all()

    contributed_papers = JournalPaper.query.join(AchievementContributor).filter(
        AchievementContributor.user_id == user_id,
        AchievementContributor.achievement_type == 'journal_paper'
    ).all()

    all_papers = {p.id: p for p in owned_papers}
    for p in contributed_papers:
        all_papers[p.id] = p

    return list(all_papers.values())


# ---------------------- 查询示例：获取论文的详细信息（包含作者信息） ----------------------

def get_paper_detail_with_authors(paper_id):
    """获取期刊论文详情及作者信息"""
    paper = JournalPaper.query.get(paper_id)
    if not paper:
        return None

    # 从 authors 字段解析作者信息（修复：支持多种分隔符）
    import re
    authors_info = []
    if paper.authors:
        author_names = [name.strip() for name in re.split(r'[;,;,,]', paper.authors)]
        corresponding_names = [name.strip() for name in
                               re.split(r'[;,;,,]', paper.corresponding_authors or '')] if paper.corresponding_authors else []

        for idx, name in enumerate(author_names, start=1):
            authors_info.append({
                'name': name,
                'author_order': idx,
                'is_corresponding': name in corresponding_names
            })

    return {
        'paper': paper,
        'authors': authors_info
    }


@app.route('/achievement/journal_paper/submit', methods=['POST'])
def submit_journal_paper():
    """提交期刊论文（支持多作者关联）"""
    current_user = get_current_user()

    title = request.form.get('title')
    authors_str = request.form.get('authors')
    corresponding_authors_str = request.form.get('corresponding_authors')
    journal_name = request.form.get('journal_name')
    year = request.form.get('year', type=int)
    volume = request.form.get('volume')
    issue = request.form.get('issue')
    page_range = request.form.get('page_range')
    doi = request.form.get('doi')

    paper = JournalPaper(
        user_id=current_user.id,
        title=title,
        authors=authors_str,
        corresponding_authors=corresponding_authors_str,
        journal_name=journal_name,
        year=year,
        volume=volume,
        issue=issue,
        page_range=page_range,
        doi=doi
    )
    db.session.add(paper)
    db.session.flush()

    auto_link_contributors(paper, 'journal_paper', authors_str, current_user.id)

    db.session.commit()

    flash('期刊论文添加成功！', 'success')
    return redirect(url_for('journal_paper_list'))


@app.route('/my_achievements')
def my_achievements():
    """我的成果（包含作为作者参与的成果）"""
    current_user = get_current_user()

    # 查询期刊论文（仅查看拥有的）
    papers = JournalPaper.query.filter_by(user_id=current_user.id).all()

    # 标记是否拥有
    papers_data = []
    for paper in papers:
        papers_data.append({
            'paper': paper,
            'is_owner': paper.user_id == current_user.id,
        })

    content = f'''
    <h2>我的期刊论文</h2>
    <div class="alert alert-info">
        显示您拥有的论文（共{len(papers_data)}篇）
    </div>
    <table class="table">
        <thead>
            <tr>
                <th>序号</th>
                <th>论文名称</th>
                <th>期刊名称</th>
                <th>年份</th>
                <th>您的角色</th>
                <th>操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    for idx, item in enumerate(papers_data, start=1):
        paper = item['paper']
        role_label = '<span class="badge badge-success">拥有者</span>'

        action_btns = f'<a href="/achievement/journal_paper/edit?id={paper.id}" class="btn btn-sm">编辑</a>'

        content += f'''
            <tr>
                <td>{idx}</td>
                <td>{paper.title}</td>
                <td>{paper.journal_name}</td>
                <td>{paper.year or '-'}</td>
                <td>{role_label}</td>
                <td>{action_btns}</td>
            </tr>
        '''

    content += '''
        </tbody>
    </table>
    '''

    return render_base_layout('我的期刊论文', content, current_user)


@app.route('/admin/dict_manage/<dict_type>', methods=['GET', 'POST'])
def admin_dict_manage(dict_type):
    """管理员-字典表维护（增删改查）"""
    current_user = get_current_user()
    if not current_user or current_user.role != 'admin':
        flash('无管理员权限！', 'danger')
        return redirect(url_for('index'))

    # 字典表映射
    dict_mapping = {
        'achievement_type': (TeachingAchievementType, '教学成果奖类型', ['type_name']),
        'achievement_level': (AchievementLevel, '成果等级', ['level_name']),
        'award_rank': (AwardRank, '获奖等级', ['rank_name'])
    }

    if dict_type not in dict_mapping:
        flash('无效的字典类型！', 'danger')
        return redirect(url_for('index'))

    model, dict_name, name_fields = dict_mapping[dict_type]

    # 处理操作
    if request.method == 'POST':
        action = request.form.get('action')
        try:
            if action == 'add':
                # 添加新记录
                name_value = request.form.get(name_fields[0])
                sort_order = request.form.get('sort_order', 0, type=int)

                if model.query.filter_by(**{name_fields[0]: name_value}).first():
                    flash(f'{name_fields[0]}已存在！', 'danger')
                else:
                    new_record = model(**{name_fields[0]: name_value, 'sort_order': sort_order})
                    db.session.add(new_record)
                    db.session.commit()
                    flash(f'{dict_name}添加成功！', 'success')

            elif action == 'edit':
                # 编辑记录
                record_id = request.form.get('id')
                record = db.session.get(model, record_id)
                if record:
                    for field in name_fields:
                        setattr(record, field, request.form.get(field))
                    record.sort_order = request.form.get('sort_order', 0, type=int)
                    db.session.commit()
                    flash(f'{dict_name}更新成功！', 'success')

            elif action == 'delete':
                # 删除记录
                record_id = request.form.get('id')
                record = db.session.get(model, record_id)
                if record:
                    db.session.delete(record)
                    db.session.commit()
                    flash(f'{dict_name}删除成功！', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'操作失败：{str(e)}', 'danger')

    # 查询所有记录
    records = model.query.order_by(model.sort_order).all()

    # 渲染页面
    dict_html = f'''
    <h2>{dict_name}管理</h2>

    <!-- 添加记录表单 -->
    <div style="margin-bottom:30px; padding:20px; border:1px solid #eee; border-radius:8px;">
        <h3>添加{dict_name}</h3>
        <form method="POST">
            <input type="hidden" name="action" value="add">
            <div class="form-group">
                <label>{name_fields[0]} <span class="required">*</span></label>
                <input type="text" name="{name_fields[0]}" required>
            </div>
            <div class="form-group">
                <label>排序顺序</label>
                <input type="number" name="sort_order" value="0">
            </div>
            <button type="submit" class="btn">添加</button>
        </form>
    </div>

    <!-- 记录列表 -->
    <table style="width:100%; border-collapse:collapse;">
        <thead>
            <tr style="background:#f5f7fa;">
                <th style="padding:10px; border:1px solid #dee2e6;">{name_fields[0]}</th>
                <th style="padding:10px; border:1px solid #dee2e6;">排序</th>
                <th style="padding:10px; border:1px solid #dee2e6;">状态</th>
                <th style="padding:10px; border:1px solid #dee2e6;">操作</th>
            </tr>
        </thead>
        <tbody>
    '''

    for record in records:
        status_text = '启用' if record.is_active else '禁用'
        status_style = 'color: #27ae60;' if record.is_active else 'color: #95a5a6;'

        dict_html += f'''
        <tr>
            <td style="padding:10px; border:1px solid #dee2e6;">{getattr(record, name_fields[0])}</td>
            <td style="padding:10px; border:1px solid #dee2e6;">{record.sort_order}</td>
            <td style="padding:10px; border:1px solid #dee2e6;"><span style="{status_style}">{status_text}</span></td>
            <td style="padding:10px; border:1px solid #dee2e6;">
                <button onclick="editRecord({record.id}, '{getattr(record, name_fields[0])}', {record.sort_order})" class="btn" style="padding:5px 10px; font-size:12px;">编辑</button>
                <form method="POST" style="display:inline;" onsubmit="return confirm('确定删除？')">
                    <input type="hidden" name="action" value="delete">
                    <input type="hidden" name="id" value="{record.id}">
                    <button type="submit" class="btn" style="padding:5px 10px; font-size:12px; background:#e74c3c;">删除</button>
                </form>
            </td>
        </tr>
        '''

    dict_html += '''
        </tbody>
    </table>

    <!-- 编辑弹窗 -->
    <div id="editModal" style="display:none; position:fixed; top:50%; left:50%; transform:translate(-50%, -50%); background:white; padding:30px; border-radius:8px; box-shadow:0 0 20px rgba(0,0,0,0.3); z-index:1000;">
        <h3>编辑''' + dict_name + '''</h3>
        <form method="POST">
            <input type="hidden" name="action" value="edit">
            <input type="hidden" name="id" id="edit_id">
            <div class="form-group">
                <label>''' + name_fields[0] + '''</label>
                <input type="text" name="''' + name_fields[0] + '''" id="edit_name" required>
            </div>
            <div class="form-group">
                <label>排序顺序</label>
                <input type="number" name="sort_order" id="edit_sort" value="0">
            </div>
            <button type="submit" class="btn">保存</button>
            <button type="button" onclick="closeEditModal()" class="btn" style="background:#95a5a6;">取消</button>
        </form>
    </div>

    <script>
        function editRecord(id, name, sort) {
            document.getElementById('edit_id').value = id;
            document.getElementById('edit_name').value = name;
            document.getElementById('edit_sort').value = sort;
            document.getElementById('editModal').style.display = 'block';
        }
        function closeEditModal() {
            document.getElementById('editModal').style.display = 'none';
        }
    </script>
    '''

    return render_base_layout(f'{dict_name}管理', dict_html, current_user)


# ---------------------- 5. 初始化数据库（强制重建+防重复创建） ----------------------
def init_database():
    """初始化数据库（强制删除旧文件 + 创建新表 + 默认管理员）"""
    with app.app_context():
        if os.path.exists(DB_FILE):
            try:
                os.remove(DB_FILE)
            except Exception as e:
                raise Exception(f'无法删除旧数据库文件，请手动删除 {DB_FILE} 后重试')

        db.create_all()

        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                username='admin',
                employee_id='000000',
                email='admin@example.com',
                role='admin'
            )
            admin.set_password('admin123')
            db.session.add(admin)

        inclusion_data = [
            {'type_name': 'SCI 期刊', 'type_code': 'SCI', 'description': '科学引文索引期刊', 'sort_order': 1},
            {'type_name': 'SSCI 期刊', 'type_code': 'SSCI', 'description': '社会科学引文索引期刊', 'sort_order': 2},
            {'type_name': 'EI 期刊', 'type_code': 'EI', 'description': '工程索引期刊', 'sort_order': 3},
            {'type_name': 'CSSCI 期刊', 'type_code': 'CSSCI', 'description': '中文社会科学引文索引期刊',
             'sort_order': 4},
            {'type_name': 'CSCD 核心库期刊', 'type_code': 'CSCD_CORE',
             'description': '中国科学引文数据库核心库期刊', 'sort_order': 5},
            {'type_name': 'CSCD 扩展库期刊', 'type_code': 'CSCD_EXT', 'description': '中国科学引文数据库扩展库期刊',
             'sort_order': 6},
            {'type_name': '北大中文核心期刊', 'type_code': 'PKU_CORE', 'description': '北京大学中文核心期刊',
             'sort_order': 7},
            {'type_name': '中国科技核心期刊', 'type_code': 'CSTPCD', 'description': '中国科技论文统计源期刊',
             'sort_order': 8},
            {'type_name': '普通期刊', 'type_code': 'GENERAL', 'description': '普通期刊', 'sort_order': 9},
            {'type_name': '其它', 'type_code': 'OTHER', 'description': '其他收录类型', 'sort_order': 10},
        ]

        for data in inclusion_data:
            existing = InclusionType.query.filter_by(type_code=data['type_code']).first()
            if not existing:
                inclusion_type = InclusionType(**data)
                db.session.add(inclusion_type)

        textbook_level_data = [
            {'level_name': '国家级规划教材', 'level_code': 'NATIONAL', 'description': '国家级规划教材',
             'sort_order': 1},
            {'level_name': '全国行业规划教材', 'level_code': 'INDUSTRY', 'description': '全国行业规划教材',
             'sort_order': 2},
            {'level_name': '协编教材', 'level_code': 'COEDIT', 'description': '协编教材', 'sort_order': 3},
            {'level_name': '自编教材', 'level_code': 'SELF', 'description': '自编教材', 'sort_order': 4},
            {'level_name': '其它', 'level_code': 'OTHER', 'description': '其它', 'sort_order': 5},
        ]

        for data in textbook_level_data:
            existing = TextbookLevel.query.filter_by(level_code=data['level_code']).first()
            if not existing:
                textbook_level = TextbookLevel(**data)
                db.session.add(textbook_level)

        db.session.commit()


# ---------------------- 6. 启动应用 ----------------------
if __name__ == '__main__':
    with app.app_context():
        db.create_all()

        # 创建默认管理员账户
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                username='admin',
                employee_id='admin',
                email='admin@hntcm.edu.cn',
                role='admin'
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()

        # 初始化教学成果奖类型
        if TeachingAchievementType.query.count() == 0:
            default_types = [
                TeachingAchievementType(type_name='湖南中医药大学教学成果奖', sort_order=1),
                TeachingAchievementType(type_name='湖南中医药大学研究生教学成果奖', sort_order=2),
                TeachingAchievementType(type_name='湖南省计算机学会高等教育教学成果奖', sort_order=3),
                TeachingAchievementType(type_name='其它', sort_order=4)
            ]
            db.session.add_all(default_types)
            db.session.commit()
            print("初始化教学成果奖类型数据")

        # 初始化成果等级
        if AchievementLevel.query.count() == 0:
            default_levels = [
                AchievementLevel(level_name='国家级', sort_order=1),
                AchievementLevel(level_name='省部级', sort_order=2),
                AchievementLevel(level_name='市厅级', sort_order=3),
                AchievementLevel(level_name='校级', sort_order=4),
                AchievementLevel(level_name='院级', sort_order=5),
                AchievementLevel(level_name='其它', sort_order=6)
            ]
            db.session.add_all(default_levels)
            db.session.commit()
            print("初始化成果等级数据")

        # 初始化获奖等级
        if AwardRank.query.count() == 0:
            default_ranks = [
                AwardRank(rank_name='特等奖', sort_order=1),
                AwardRank(rank_name='一等奖', sort_order=2),
                AwardRank(rank_name='二等奖', sort_order=3),
                AwardRank(rank_name='三等奖', sort_order=4),
                AwardRank(rank_name='优秀奖', sort_order=5),
                AwardRank(rank_name='其它', sort_order=6)
            ]
            db.session.add_all(default_ranks)
            db.session.commit()
            print("初始化获奖等级数据")

        print("数据库初始化完成")

    app.run(debug=True, host='0.0.0.0', port=5000)

